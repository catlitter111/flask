#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
人体特征提取ROS2节点
==================
基于RKNN的人体特征提取服务，整合服装检测和姿态估计
- 提供特征提取服务接口
- 计算16个身体比例特征
- 提取服装颜色信息
- 保存特征数据到Excel文件
- 生成带标注的结果图像

作者: AI Assistant
移植自: obtain_features.py
"""

import rclpy
from rclpy.node import Node
from custom_msgs.srv import FeatureExtraction
from sensor_msgs.msg import Image
from cv_bridge import CvBridge

import cv2
import numpy as np
import time
import traceback
import logging
from datetime import datetime
from pathlib import Path
import openpyxl
from collections import Counter

# 导入现有的检测模块
try:
    from .rknn_colour_detect import Obtain_the_target_color
    CLOTHING_DETECTION_AVAILABLE = True
    print("✅ 服装检测模块导入成功")
except ImportError as e:
    CLOTHING_DETECTION_AVAILABLE = False
    print(f"⚠️ 服装检测模块导入失败: {e}")

try:
    from .yolov8_pose_detector import detect_human_pose, draw_human_pose
    POSE_DETECTION_AVAILABLE = True
    print("✅ 姿态检测模块导入成功")
except ImportError as e:
    POSE_DETECTION_AVAILABLE = False
    print(f"⚠️ 姿态检测模块导入失败: {e}")


class FeatureExtractionNode(Node):
    """人体特征提取ROS2节点类"""
    
    def __init__(self):
        super().__init__('feature_extraction_node')
        
        # 声明参数
        self.declare_parameter('output_dir', 'features-data')
        self.declare_parameter('video_frame_interval', 10)  # 视频帧间隔
        self.declare_parameter('video_detection_limit', 20)  # 视频最大检测帧数
        
        # 初始化CV Bridge
        self.bridge = CvBridge()
        
        # 配置输出目录
        output_dir_param = self.get_parameter('output_dir').get_parameter_value().string_value
        self.output_dir = Path(output_dir_param)
        self.setup_output_directory()
        
        # 创建服务
        self.feature_service = self.create_service(
            FeatureExtraction,
            '/features/extract_features',
            self.extract_features_callback
        )
        
        # 统计信息
        self.extraction_count = 0
        
        self.get_logger().info('✅ 人体特征提取节点初始化完成')
        if CLOTHING_DETECTION_AVAILABLE:
            self.get_logger().info('🤖 服装检测功能已启用')
        else:
            self.get_logger().warn('⚠️ 服装检测功能不可用')
        if POSE_DETECTION_AVAILABLE:
            self.get_logger().info('🦴 姿态检测功能已启用')
        else:
            self.get_logger().warn('⚠️ 姿态检测功能不可用')
        
        # 显示视频处理配置
        frame_interval = self.get_parameter('video_frame_interval').get_parameter_value().integer_value
        detection_limit = self.get_parameter('video_detection_limit').get_parameter_value().integer_value
        self.get_logger().info(f'🎬 视频处理功能已启用: 每{frame_interval}帧检测, 最多{detection_limit}帧')
        
        self.get_logger().info('🔧 服务地址: /features/extract_features')
        self.get_logger().info('📝 使用方法: person_name前缀"VIDEO:"可处理视频文件')

    def setup_output_directory(self):
        """创建输出目录"""
        self.output_dir.mkdir(exist_ok=True)
        self.get_logger().info(f"输出目录已设置: {self.output_dir}")

    def extract_features_callback(self, request, response):
        """特征提取服务回调函数"""
        try:
            self.get_logger().info(f"📞 收到特征提取请求: 人物名称={request.person_name}")
            start_time = time.time()
            
            # 检查是否为视频处理请求（通过特殊标识符）
            if request.person_name.startswith("VIDEO:"):
                # 视频文件路径从person_name中提取
                video_path = request.person_name[6:]  # 移除"VIDEO:"前缀
                self.get_logger().info(f"🎬 检测到视频处理请求: {video_path}")
                
                # 检查视频文件是否存在
                if not Path(video_path).exists():
                    response.success = False
                    response.message = f"视频文件不存在: {video_path}"
                    self.get_logger().error(response.message)
                    return response
                
                # 使用视频处理功能
                result = self.obtain_features_from_video(
                    video_path,
                    Path(video_path).stem,  # 使用文件名作为标识
                    request.save_to_file,
                    request.output_path
                )
                
            else:
                # 常规图像处理
                # 转换ROS图像到OpenCV格式
                try:
                    cv_image = self.bridge.imgmsg_to_cv2(request.image, "bgr8")
                    self.get_logger().info(f"📸 图像转换成功，尺寸: {cv_image.shape}")
                except Exception as e:
                    response.success = False
                    response.message = f"图像转换失败: {str(e)}"
                    self.get_logger().error(response.message)
                    return response
                
                # 提取特征
                result = self.obtain_features(
                    cv_image, 
                    request.person_name,
                    request.save_to_file,
                    request.output_path
                )
            
            if result is None:
                response.success = False
                response.message = "特征提取失败"
                self.get_logger().error("❌ 特征提取失败")
                return response
            
            # 解析结果
            canvas, person_ratios, result_paths = result
            
            # 填充响应
            response.success = True
            response.message = f"特征提取成功，耗时: {time.time() - start_time:.2f}秒"
            response.person_count = 1 if person_ratios else 0
            
            if person_ratios and len(person_ratios) >= 18:
                # 身体比例数据（前16个）
                response.body_ratios = [float(x) for x in person_ratios[:16]]
                
                # 服装颜色数据（第17和18个元素）
                shirt_color = person_ratios[16]
                pants_color = person_ratios[17]
                
                # 转换颜色格式
                response.shirt_color = self.parse_color(shirt_color)
                response.pants_color = self.parse_color(pants_color)
            else:
                response.body_ratios = [0.0] * 16
                response.shirt_color = [0, 0, 0]
                response.pants_color = [0, 0, 0]
            
            # 文件路径
            response.result_image_path = result_paths.get('image', '')
            response.feature_data_path = result_paths.get('excel', '')
            response.result_video_path = result_paths.get('video', '')
            
            self.extraction_count += 1
            self.get_logger().info(f"✅ 特征提取完成，总处理次数: {self.extraction_count}")
            
            return response
            
        except Exception as e:
            self.get_logger().error(f"特征提取服务回调错误: {e}")
            traceback.print_exc()
            response.success = False
            response.message = f"服务处理错误: {str(e)}"
            return response

    def parse_color(self, color_data):
        """解析颜色数据为整数列表"""
        try:
            if isinstance(color_data, (tuple, list)) and len(color_data) >= 3:
                return [int(color_data[0]), int(color_data[1]), int(color_data[2])]
            elif isinstance(color_data, str) and color_data.startswith('(') and color_data.endswith(')'):
                # 处理字符串形式的元组 "(r,g,b)"
                color_tuple = eval(color_data)
                return [int(color_tuple[0]), int(color_tuple[1]), int(color_tuple[2])]
            else:
                return [0, 0, 0]  # 默认黑色
        except Exception as e:
            self.get_logger().warn(f"颜色解析失败: {e}")
            return [0, 0, 0]

    def obtain_features(self, image, name, save_to_file=True, output_path=''):
        """
        从单个图像中提取特征，整合服装检测和姿态估计
        """
        start_time = time.time()
        self.get_logger().info(f"开始处理图像: {name}")

        try:
            # 确保图像格式正确
            if image is None:
                raise ValueError("输入图像为空")

            img = image.copy()
            canvas = img.copy()

            # 获取目标颜色（服装检测）
            shirt_color = (0, 0, 0)  # 默认黑色
            pants_color = (0, 0, 0)  # 默认黑色
            
            if CLOTHING_DETECTION_AVAILABLE:
                self.get_logger().info("正在获取目标颜色...")
                try:
                    pairs = Obtain_the_target_color(img)
                    self.get_logger().info(f"检测到 {len(pairs)} 对服装")

                    for pair in pairs:
                        # 检查是否有上衣颜色
                        if len(pair) > 2 and pair[2]:
                            shirt_color = pair[2]

                        # 检查是否有裤子颜色
                        if len(pair) > 3 and pair[3]:
                            pants_color = pair[3]

                        # 在图像上标记上衣和下装
                        if len(pair[0]) > 1:  # 有上衣
                            x1, y1, x2, y2 = pair[0][:4]
                            cv2.rectangle(canvas, (x1, y1), (x2, y2), (255, 0, 0), 2)
                            cv2.putText(canvas, "Upper", (x1, y1 - 10),
                                      cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 1)

                        if len(pair[1]) > 1:  # 有下装
                            x1, y1, x2, y2 = pair[1][:4]
                            cv2.rectangle(canvas, (x1, y1), (x2, y2), (0, 255, 0), 2)
                            cv2.putText(canvas, "Lower", (x1, y1 - 10),
                                      cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
                except Exception as e:
                    self.get_logger().warn(f"服装检测失败: {e}")

            # 获取人体关键点和计算身体比例
            body_ratios = [0] * 16  # 默认值
            
            if POSE_DETECTION_AVAILABLE:
                self.get_logger().info("正在获取人体关键点...")
                try:
                    pose_results = detect_human_pose(img)
                    
                    if pose_results and len(pose_results) > 0:
                        self.get_logger().info(f"检测到 {len(pose_results)} 个人体")
                        
                        # 绘制姿态检测结果
                        canvas = draw_human_pose(canvas, pose_results)
                        
                        # 计算身体比例（使用第一个检测到的人）
                        keypoints = pose_results[0].keypoints
                        body_ratios = self.calculate_body_ratios(keypoints)
                        
                        if body_ratios is None:
                            body_ratios = [0] * 16
                        
                        self.get_logger().info("身体比例计算成功")
                    else:
                        self.get_logger().warn("未检测到人体关键点")
                except Exception as e:
                    self.get_logger().warn(f"姿态检测失败: {e}")

            # 组合所有特征：身体比例+衣物颜色
            person_ratios = body_ratios + [shirt_color, pants_color]

            # 保存结果
            result_paths = {}
            
            if save_to_file:
                # 确定输出路径
                if output_path:
                    output_dir = Path(output_path)
                    output_dir.mkdir(parents=True, exist_ok=True)
                else:
                    output_dir = self.output_dir

                # 保存结果图像
                result_image_path = output_dir / f"{name}_result.jpg"
                cv2.imwrite(str(result_image_path), canvas)
                result_paths['image'] = str(result_image_path)
                self.get_logger().info(f"结果图像已保存到: {result_image_path}")

                # 保存特征数据到Excel
                excel_path = output_dir / f"{name}_features.xlsx"
                self.save_features_to_excel(body_ratios, shirt_color, pants_color, excel_path, name)
                result_paths['excel'] = str(excel_path)
                self.get_logger().info(f"特征数据已保存到: {excel_path}")

            self.get_logger().info(f"图像处理完成，耗时: {time.time() - start_time:.2f}秒")
            return canvas, person_ratios, result_paths

        except Exception as e:
            self.get_logger().error(f"处理图像时出错: {e}")
            traceback.print_exc()
            return None

    def obtain_features_from_video(self, video_path, name, save_to_file=True, output_path=''):
        """
        从视频文件中提取特征，每隔几帧检测几次，取平均值，并保存完整的标注视频
        """
        start_time = time.time()
        self.get_logger().info(f"开始处理视频: {video_path}")

        try:
            # 打开视频文件
            cap = cv2.VideoCapture(video_path)
            if not cap.isOpened():
                raise ValueError(f"无法打开视频文件: {video_path}")

            # 获取视频信息
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            fps = cap.get(cv2.CAP_PROP_FPS)
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            duration = total_frames / fps if fps > 0 else 0
            
            frame_interval = self.get_parameter('video_frame_interval').get_parameter_value().integer_value
            detection_limit = self.get_parameter('video_detection_limit').get_parameter_value().integer_value
            
            self.get_logger().info(f"视频信息: 总帧数={total_frames}, FPS={fps:.2f}, 分辨率={width}x{height}, 时长={duration:.2f}秒")
            self.get_logger().info(f"检测设置: 每{frame_interval}帧检测一次, 最多检测{detection_limit}帧")

            # 准备视频写入器（如果需要保存）
            video_writer = None
            result_video_path = None
            
            if save_to_file:
                # 确定输出路径
                if output_path:
                    output_dir = Path(output_path)
                    output_dir.mkdir(parents=True, exist_ok=True)
                else:
                    output_dir = self.output_dir
                
                # 设置输出视频路径
                result_video_path = output_dir / f"{name}_video_result.mp4"
                
                # 创建视频写入器
                fourcc = cv2.VideoWriter.fourcc(*'mp4v')
                video_writer = cv2.VideoWriter(str(result_video_path), fourcc, fps, (width, height))
                
                if video_writer.isOpened():
                    self.get_logger().info(f"📹 视频写入器已创建: {result_video_path}")
                else:
                    self.get_logger().error("❌ 视频写入器创建失败")
                    video_writer = None

            # 收集所有检测结果
            all_body_ratios = []
            all_shirt_colors = []
            all_pants_colors = []
            processed_frames = []
            frame_count = 0
            detection_count = 0
            
            # 重置视频到开头
            cap.set(cv2.CAP_PROP_POS_FRAMES, 0)

            while True:
                ret, frame = cap.read()
                if not ret:
                    break

                # 创建当前帧的副本用于标注
                annotated_frame = frame.copy()
                
                # 在帧上添加基本信息
                progress = (frame_count + 1) / total_frames * 100
                info_text = f"Frame: {frame_count + 1}/{total_frames} ({progress:.1f}%)"
                cv2.putText(annotated_frame, info_text, (10, 30), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)

                # 是否需要进行特征提取检测
                should_detect = (frame_count % frame_interval == 0) and (detection_count < detection_limit)
                
                if should_detect:
                    self.get_logger().info(f"处理第 {frame_count} 帧 (检测 {detection_count + 1}/{detection_limit})")
                    
                    # 对当前帧进行特征提取
                    result = self.extract_features_from_frame(frame, f"{name}_frame_{frame_count}")
                    
                    if result:
                        body_ratios, shirt_color, pants_color, detection_frame = result
                        
                        # 使用检测结果的标注帧替换基本标注帧
                        annotated_frame = detection_frame.copy()
                        
                        # 在检测帧上添加进度信息
                        cv2.putText(annotated_frame, info_text, (10, 30), 
                                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                        
                        # 检查结果有效性
                        if self.is_valid_detection(body_ratios, shirt_color, pants_color):
                            all_body_ratios.append(body_ratios)
                            all_shirt_colors.append(shirt_color)
                            all_pants_colors.append(pants_color)
                            processed_frames.append(annotated_frame)
                            
                            # 在帧上标记这是一个有效检测
                            cv2.putText(annotated_frame, "VALID DETECTION", (10, 60), 
                                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
                            
                            self.get_logger().info(f"第 {frame_count} 帧检测有效")
                        else:
                            # 在帧上标记这是一个无效检测
                            cv2.putText(annotated_frame, "INVALID DETECTION", (10, 60), 
                                       cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
                            self.get_logger().info(f"第 {frame_count} 帧检测无效，跳过")
                    else:
                        # 检测失败
                        cv2.putText(annotated_frame, "DETECTION FAILED", (10, 60), 
                                   cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
                    
                    detection_count += 1
                else:
                    # 非检测帧，只显示基本信息
                    cv2.putText(annotated_frame, "SKIPPED FRAME", (10, 60), 
                               cv2.FONT_HERSHEY_SIMPLEX, 0.7, (128, 128, 128), 2)

                # 添加统计信息到帧上
                if all_body_ratios:
                    stats_text = f"Valid detections: {len(all_body_ratios)}"
                    cv2.putText(annotated_frame, stats_text, (10, height - 30), 
                               cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 0), 2)

                # 写入视频帧（无论是否检测）
                if video_writer is not None:
                    video_writer.write(annotated_frame)

                frame_count += 1

                # 每处理100帧输出一次进度
                if frame_count % 100 == 0:
                    elapsed_time = time.time() - start_time
                    estimated_total = elapsed_time * total_frames / frame_count
                    remaining_time = estimated_total - elapsed_time
                    self.get_logger().info(f"进度: {frame_count}/{total_frames} ({progress:.1f}%), "
                                         f"预计剩余: {remaining_time:.1f}秒")

            # 释放资源
            cap.release()
            if video_writer is not None:
                video_writer.release()
                self.get_logger().info(f"✅ 视频写入完成: {result_video_path}")

            if not all_body_ratios:
                self.get_logger().error("视频中未检测到有效特征")
                return None

            # 计算平均特征
            self.get_logger().info(f"共收集到 {len(all_body_ratios)} 个有效检测结果，开始计算平均值")
            avg_body_ratios = self.calculate_average_body_ratios(all_body_ratios)
            avg_shirt_color = self.calculate_average_color(all_shirt_colors)
            avg_pants_color = self.calculate_average_color(all_pants_colors)

            # 使用最后一个有效帧作为结果图像
            canvas = processed_frames[-1] if processed_frames else None
            
            if canvas is not None:
                # 在图像上添加统计信息
                cv2.putText(canvas, f"Total Frames: {frame_count}", 
                           (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                cv2.putText(canvas, f"Valid Detections: {len(all_body_ratios)}/{detection_count}", 
                           (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
                cv2.putText(canvas, f"Avg Features Extracted", 
                           (10, 90), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

            # 组合平均特征
            person_ratios = avg_body_ratios + [avg_shirt_color, avg_pants_color]

            # 保存结果
            result_paths = {}
            
            if save_to_file:
                # 确定输出路径
                if output_path:
                    output_dir = Path(output_path)
                    output_dir.mkdir(parents=True, exist_ok=True)
                else:
                    output_dir = self.output_dir

                # 保存结果图像（最后一帧）
                if canvas is not None:
                    result_image_path = output_dir / f"{name}_video_result.jpg"
                    cv2.imwrite(str(result_image_path), canvas)
                    result_paths['image'] = str(result_image_path)
                    self.get_logger().info(f"视频结果图像已保存到: {result_image_path}")

                # 保存完整视频路径
                if result_video_path and result_video_path.exists():
                    result_paths['video'] = str(result_video_path)
                    video_size = result_video_path.stat().st_size / (1024 * 1024)  # MB
                    self.get_logger().info(f"完整标注视频已保存到: {result_video_path} (大小: {video_size:.2f}MB)")

                # 保存特征数据到Excel
                excel_path = output_dir / f"{name}_video_features.xlsx"
                self.save_video_features_to_excel(
                    all_body_ratios, avg_body_ratios, 
                    all_shirt_colors, avg_shirt_color,
                    all_pants_colors, avg_pants_color,
                    excel_path, name, len(all_body_ratios), detection_count
                )
                result_paths['excel'] = str(excel_path)
                self.get_logger().info(f"视频特征数据已保存到: {excel_path}")

            processing_time = time.time() - start_time
            self.get_logger().info(f"视频处理完成，耗时: {processing_time:.2f}秒，有效检测: {len(all_body_ratios)}/{detection_count}")
            self.get_logger().info(f"处理速度: {frame_count / processing_time:.1f} FPS")
            
            return canvas, person_ratios, result_paths

        except Exception as e:
            self.get_logger().error(f"处理视频时出错: {e}")
            traceback.print_exc()
            return None

    def extract_features_from_frame(self, frame, frame_name):
        """
        从单帧图像中提取特征（不保存文件）
        """
        try:
            # 获取目标颜色（服装检测）
            shirt_color = (0, 0, 0)
            pants_color = (0, 0, 0)
            canvas = frame.copy()
            
            if CLOTHING_DETECTION_AVAILABLE:
                try:
                    pairs = Obtain_the_target_color(frame)
                    for pair in pairs:
                        if len(pair) > 2 and pair[2]:
                            shirt_color = pair[2]
                        if len(pair) > 3 and pair[3]:
                            pants_color = pair[3]
                        
                        # 标记检测结果
                        if len(pair[0]) > 1:
                            x1, y1, x2, y2 = pair[0][:4]
                            cv2.rectangle(canvas, (x1, y1), (x2, y2), (255, 0, 0), 2)
                        if len(pair[1]) > 1:
                            x1, y1, x2, y2 = pair[1][:4]
                            cv2.rectangle(canvas, (x1, y1), (x2, y2), (0, 255, 0), 2)
                except Exception as e:
                    self.get_logger().debug(f"帧 {frame_name} 服装检测失败: {e}")

            # 获取人体关键点和计算身体比例
            body_ratios = [0] * 16
            
            if POSE_DETECTION_AVAILABLE:
                try:
                    pose_results = detect_human_pose(frame)
                    if pose_results and len(pose_results) > 0:
                        canvas = draw_human_pose(canvas, pose_results)
                        keypoints = pose_results[0].keypoints
                        body_ratios = self.calculate_body_ratios(keypoints)
                        if body_ratios is None:
                            body_ratios = [0] * 16
                except Exception as e:
                    self.get_logger().debug(f"帧 {frame_name} 姿态检测失败: {e}")

            return body_ratios, shirt_color, pants_color, canvas

        except Exception as e:
            self.get_logger().debug(f"处理帧 {frame_name} 时出错: {e}")
            return None

    def is_valid_detection(self, body_ratios, shirt_color, pants_color):
        """
        检查检测结果是否有效
        """
        # 检查身体比例是否有效（至少有一半的比例值不为0）
        valid_ratios = sum(1 for ratio in body_ratios if ratio > 0)
        ratio_threshold = len(body_ratios) * 0.3  # 至少30%的比例有效
        
        # 检查颜色是否有效（不是纯黑色）
        shirt_valid = shirt_color != (0, 0, 0) if isinstance(shirt_color, tuple) else any(c > 0 for c in shirt_color)
        pants_valid = pants_color != (0, 0, 0) if isinstance(pants_color, tuple) else any(c > 0 for c in pants_color)
        
        # 至少需要身体比例有效或者颜色检测有效
        return valid_ratios >= ratio_threshold or shirt_valid or pants_valid

    def calculate_average_body_ratios(self, all_ratios):
        """
        计算身体比例的平均值，排除无效值
        """
        if not all_ratios:
            return [0.0] * 16
        
        averaged_ratios = []
        num_ratios = len(all_ratios[0])
        
        for i in range(num_ratios):
            valid_values = [ratios[i] for ratios in all_ratios if ratios[i] > 0]
            if valid_values:
                avg_value = sum(valid_values) / len(valid_values)
                averaged_ratios.append(avg_value)
            else:
                averaged_ratios.append(0.0)
        
        return averaged_ratios

    def calculate_average_color(self, all_colors):
        """
        计算颜色的平均值，排除无效值
        """
        if not all_colors:
            return (0, 0, 0)
        
        valid_colors = []
        for color in all_colors:
            if isinstance(color, tuple) and color != (0, 0, 0):
                valid_colors.append(color)
            elif isinstance(color, (list, tuple)) and len(color) >= 3 and any(c > 0 for c in color[:3]):
                valid_colors.append(tuple(color[:3]))
        
        if not valid_colors:
            return (0, 0, 0)
        
        # 计算RGB平均值
        avg_r = sum(color[0] for color in valid_colors) / len(valid_colors)
        avg_g = sum(color[1] for color in valid_colors) / len(valid_colors)
        avg_b = sum(color[2] for color in valid_colors) / len(valid_colors)
        
        return (int(avg_r), int(avg_g), int(avg_b))

    def calculate_body_ratios(self, keypoints):
        """
        根据人体关键点计算身体比例
        """
        if keypoints is None or len(keypoints) == 0:
            self.get_logger().error("没有有效的关键点数据")
            return None

        try:
            # keypoints是numpy数组，形状为 [17, 3] (x, y, confidence)
            kpts = keypoints

            # 检查关键点有效性的函数
            def is_valid(p1_idx, p2_idx):
                return (kpts[p1_idx][2] > 0.5 and kpts[p2_idx][2] > 0.5 and
                        kpts[p1_idx][0] != 0 and kpts[p1_idx][1] != 0 and
                        kpts[p2_idx][0] != 0 and kpts[p2_idx][1] != 0)

            # 计算两点之间的距离
            def distance(p1_idx, p2_idx):
                if not is_valid(p1_idx, p2_idx):
                    return 0
                return np.sqrt(((kpts[p1_idx][0] - kpts[p2_idx][0]) ** 2) +
                               ((kpts[p1_idx][1] - kpts[p2_idx][1]) ** 2))

            # 计算一系列有意义的身体比例
            ratios = []

            # YOLOv8 Pose关键点索引:
            # 0: 鼻子, 1: 左眼, 2: 右眼, 3: 左耳, 4: 右耳, 5: 左肩, 6: 右肩
            # 7: 左肘, 8: 右肘, 9: 左腕, 10: 右腕, 11: 左髋, 12: 右髋
            # 13: 左膝, 14: 右膝, 15: 左踝, 16: 右踝

            # 1. 上肢与下肢比例
            upper_limb = (distance(5, 7) + distance(7, 9) + distance(6, 8) + distance(8, 10)) / 4
            lower_limb = (distance(11, 13) + distance(13, 15) + distance(12, 14) + distance(14, 16)) / 4
            if upper_limb > 0 and lower_limb > 0:
                ratios.append(upper_limb / lower_limb)
            else:
                ratios.append(0)

            # 2. 躯干与身高比例
            torso_height = (distance(5, 11) + distance(6, 12)) / 2
            body_height = ((distance(0, 15) + distance(0, 16)) / 2)
            if torso_height > 0 and body_height > 0:
                ratios.append(torso_height / body_height)
            else:
                ratios.append(0)

            # 3. 肩宽与身高比例
            shoulder_width = distance(5, 6)
            if shoulder_width > 0 and body_height > 0:
                ratios.append(shoulder_width / body_height)
            else:
                ratios.append(0)

            # 4. 臀宽与肩宽比例
            hip_width = distance(11, 12)
            if hip_width > 0 and shoulder_width > 0:
                ratios.append(hip_width / shoulder_width)
            else:
                ratios.append(0)

            # 5. 头部与躯干比例
            head_height = (distance(0, 5) + distance(0, 6)) / 2
            if head_height > 0 and torso_height > 0:
                ratios.append(head_height / torso_height)
            else:
                ratios.append(0)

            # 6. 手臂与身高比例
            arm_length = (distance(5, 9) + distance(6, 10)) / 2
            if arm_length > 0 and body_height > 0:
                ratios.append(arm_length / body_height)
            else:
                ratios.append(0)

            # 7. 腿长与身高比例
            leg_length = (distance(11, 15) + distance(12, 16)) / 2
            if leg_length > 0 and body_height > 0:
                ratios.append(leg_length / body_height)
            else:
                ratios.append(0)

            # 8. 上臂与下臂比例
            upper_arm = (distance(5, 7) + distance(6, 8)) / 2
            lower_arm = (distance(7, 9) + distance(8, 10)) / 2
            if upper_arm > 0 and lower_arm > 0:
                ratios.append(upper_arm / lower_arm)
            else:
                ratios.append(0)

            # 9. 大腿与小腿比例
            thigh = (distance(11, 13) + distance(12, 14)) / 2
            calf = (distance(13, 15) + distance(14, 16)) / 2
            if thigh > 0 and calf > 0:
                ratios.append(thigh / calf)
            else:
                ratios.append(0)

            # 10. 躯干与腿长比例
            if torso_height > 0 and leg_length > 0:
                ratios.append(torso_height / leg_length)
            else:
                ratios.append(0)

            # 11. 手臂与腿长比例
            if arm_length > 0 and leg_length > 0:
                ratios.append(arm_length / leg_length)
            else:
                ratios.append(0)

            # 12. 肩宽与髋宽比例
            if shoulder_width > 0 and hip_width > 0:
                ratios.append(shoulder_width / hip_width)
            else:
                ratios.append(0)

            # 13. 头围与身高比例（估算头围）
            head_width = (distance(3, 4) * 1.2)  # 估算头围
            if head_width > 0 and body_height > 0:
                ratios.append(head_width / body_height)
            else:
                ratios.append(0)

            # 14. 脚长与身高比例（估算脚长）
            foot_length = (distance(15, 16) * 0.7)  # 估算脚长
            if foot_length > 0 and body_height > 0:
                ratios.append(foot_length / body_height)
            else:
                ratios.append(0)

            # 15. 脚踝宽与身高比例
            ankle_width = distance(15, 16)
            if ankle_width > 0 and body_height > 0:
                ratios.append(ankle_width / body_height)
            else:
                ratios.append(0)

            # 16. 腰围与身高比例（估算腰围）
            waist = (hip_width * 0.85)  # 估算腰围
            if waist > 0 and body_height > 0:
                ratios.append(waist / body_height)
            else:
                ratios.append(0)

            # 确保我们有16个比例
            while len(ratios) < 16:
                ratios.append(0)

            return ratios[:16]  # 只返回前16个比例

        except Exception as e:
            self.get_logger().error(f"计算身体比例时出错: {e}")
            traceback.print_exc()
            return None

    def save_features_to_excel(self, body_ratios, shirt_color, pants_color, excel_path, name):
        """保存特征数据到Excel文件"""
        try:
            wb = openpyxl.Workbook()
            sheet = wb.active
            if sheet is None:
                # 如果 active 为 None，创建一个新的工作表
                sheet = wb.create_sheet()
                wb.remove(wb.worksheets[0])  # 删除默认空工作表
            
            sheet.title = self.sanitize_sheet_name(f'特征_{name}')

            # 添加比例数据 - 直接将数据保存到A列
            for i, ratio in enumerate(body_ratios):
                sheet.cell(row=i + 1, column=1).value = float(ratio)

            # 添加颜色数据 - 直接添加到A17和A18单元格
            sheet.cell(row=17, column=1).value = str(shirt_color)
            sheet.cell(row=18, column=1).value = str(pants_color)

            wb.save(str(excel_path))
            self.get_logger().info(f"特征数据已保存到: {excel_path}")

        except Exception as e:
            self.get_logger().error(f"保存Excel文件失败: {e}")
            traceback.print_exc()

    def save_video_features_to_excel(self, all_body_ratios, avg_body_ratios, 
                                   all_shirt_colors, avg_shirt_color,
                                   all_pants_colors, avg_pants_color,
                                   excel_path, name, valid_frames, total_frames):
        """保存视频特征数据到Excel文件，包含每帧数据和平均值"""
        try:
            wb = openpyxl.Workbook()
            
            # 删除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

            # 创建平均值工作表
            avg_sheet = wb.create_sheet(title=self.sanitize_sheet_name(f'平均特征_{name}'))
            
            # 添加平均比例数据 - 直接将数据保存到A列
            for i, ratio in enumerate(avg_body_ratios):
                avg_sheet.cell(row=i + 1, column=1).value = float(ratio)

            # 添加平均颜色数据 - 直接添加到A17和A18单元格
            avg_sheet.cell(row=17, column=1).value = str(avg_shirt_color)
            avg_sheet.cell(row=18, column=1).value = str(avg_pants_color)

            # 添加统计信息
            avg_sheet.cell(row=19, column=1).value = f"有效帧数: {valid_frames}"
            avg_sheet.cell(row=20, column=1).value = f"总检测帧数: {total_frames}"

            # 创建详细数据工作表
            if all_body_ratios:
                detail_sheet = wb.create_sheet(title=self.sanitize_sheet_name(f'详细数据_{name}'))
                
                # 添加标题行
                headers = [f"比例{i+1}" for i in range(16)] + ["上装颜色", "下装颜色"]
                for col, header in enumerate(headers, 1):
                    detail_sheet.cell(row=1, column=col).value = header

                # 添加每帧的数据
                for row_idx, (body_ratios, shirt_color, pants_color) in enumerate(
                    zip(all_body_ratios, all_shirt_colors, all_pants_colors), 2):
                    
                    # 身体比例数据
                    for col_idx, ratio in enumerate(body_ratios, 1):
                        detail_sheet.cell(row=row_idx, column=col_idx).value = float(ratio)
                    
                    # 颜色数据
                    detail_sheet.cell(row=row_idx, column=17).value = str(shirt_color)
                    detail_sheet.cell(row=row_idx, column=18).value = str(pants_color)

            wb.save(str(excel_path))
            self.get_logger().info(f"视频特征数据已保存到: {excel_path}")

        except Exception as e:
            self.get_logger().error(f"保存视频Excel文件失败: {e}")
            traceback.print_exc()

    def sanitize_sheet_name(self, name, replacement='_'):
        """
        清理Excel工作表名称，替换不允许的字符
        """
        # Excel工作表名称不允许包含的字符
        invalid_chars = [':', '/', '\\', '?', '*', '[', ']', "'", '<', '>']

        # 替换所有不合法字符
        result = name
        for char in invalid_chars:
            result = result.replace(char, replacement)

        # 确保名称不超过31个字符（Excel工作表名称的最大长度）
        if len(result) > 31:
            result = result[:31]

        return result


def main(args=None):
    """主函数"""
    try:
        rclpy.init(args=args)
        node = FeatureExtractionNode()
        
        try:
            rclpy.spin(node)
        except KeyboardInterrupt:
            pass
        finally:
            node.destroy_node()
            rclpy.shutdown()
            
    except Exception as e:
        print(f"主函数错误: {e}")
        traceback.print_exc()


if __name__ == '__main__':
    main() 