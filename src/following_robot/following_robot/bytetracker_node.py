#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简化版ByteTracker ROS2节点
========================
专注于跟踪算法，通过订阅消息接收特征数据

功能特性：
- 多目标跟踪模式：跟踪所有检测到的人物
- 单目标跟踪模式：锁定并跟踪特定目标人物  
- 融合颜色特征和身体比例特征的匹配算法
- 卡尔曼滤波预测目标运动
- 通过ROS2消息接收特征数据

作者: AI Assistant
简化版本: v2.0.0
"""

import rclpy
from rclpy.node import Node
from rclpy.qos import QoSProfile, ReliabilityPolicy, HistoryPolicy
import numpy as np
import time
import traceback
from collections import OrderedDict, deque
from enum import Enum
import threading
import scipy.linalg
import os
import json
import cv2

# ROS2消息类型
from geometry_msgs.msg import Point, Twist
from std_msgs.msg import String, Float32, Bool, Header
from sensor_msgs.msg import Image
from cv_bridge import CvBridge

# 自定义消息
from custom_msgs.msg import TrackedPerson, TrackedPersonArray, TrackingMode, Position, TrackingResult

# ==================== ByteTracker核心类 ====================

class TrackState:
    """跟踪状态枚举类"""
    NEW = 0        # 新建
    TRACKED = 1    # 跟踪中
    LOST = 2       # 丢失
    REMOVED = 3    # 已移除

class BaseTrack:
    """基础跟踪类"""
    _count = 0

    def __init__(self):
        self.track_id = 0
        self.is_activated = False
        self.state = TrackState.NEW
        self.history = OrderedDict()
        self.features = []
        self.curr_feature = None
        self.score = 0
        self.start_frame = 0
        self.frame_id = 0
        self.time_since_update = 0
        self.upper_color = None
        self.lower_color = None
        self.location = (np.inf, np.inf)

    @property
    def end_frame(self):
        return self.frame_id

    @staticmethod
    def next_id():
        BaseTrack._count += 1
        return BaseTrack._count

    def activate(self, *args):
        raise NotImplementedError("需要在子类中实现")

    def predict(self):
        raise NotImplementedError("需要在子类中实现")

    def update(self, *args, **kwargs):
        raise NotImplementedError("需要在子类中实现")

    def mark_lost(self):
        self.state = TrackState.LOST

    def mark_removed(self):
        self.state = TrackState.REMOVED

class KalmanFilter:
    """卡尔曼滤波器"""

    def __init__(self):
        self.ndim = 4
        self.dt = 1.0
        self._motion_mat = np.eye(2 * self.ndim, 2 * self.ndim)
        for i in range(self.ndim):
            self._motion_mat[i, self.ndim + i] = self.dt
        self._update_mat = np.eye(self.ndim, 2 * self.ndim)
        self._std_weight_position = 1. / 20
        self._std_weight_velocity = 1. / 160
        self.chi2inv95 = {
            1: 3.8415, 2: 5.9915, 3: 7.8147, 4: 9.4877,
            5: 11.070, 6: 12.592, 7: 14.067, 8: 15.507, 9: 16.919
        }

    def initiate(self, measurement):
        """
        参数: measurement (np.array): [cx, cy, a, h] 格式的观测值
        返回: tuple: (mean, covariance) 初始状态和协方差
        """
        mean_pos = measurement
        mean_vel = np.zeros_like(mean_pos)
        mean = np.r_[mean_pos, mean_vel]
        std = [
            2 * self._std_weight_position * measurement[3],
            2 * self._std_weight_position * measurement[3],
            1e-2,
            2 * self._std_weight_position * measurement[3],
            10 * self._std_weight_velocity * measurement[3],
            10 * self._std_weight_velocity * measurement[3],
            1e-5,
            10 * self._std_weight_velocity * measurement[3]
        ]
        covariance = np.diag(np.square(std))
        return mean, covariance

    def predict(self, mean, covariance):
        """
        参数: mean, covariance - 当前状态
        返回: tuple: (predicted_mean, predicted_covariance) 预测状态
        """
        std_pos = [
            self._std_weight_position * mean[3],
            self._std_weight_position * mean[3],
            1e-2,
            self._std_weight_position * mean[3]
        ]
        std_vel = [
            self._std_weight_velocity * mean[3],
            self._std_weight_velocity * mean[3],
            1e-5,
            self._std_weight_velocity * mean[3]
        ]
        motion_cov = np.diag(np.square(np.r_[std_pos, std_vel]))
        mean = np.dot(mean, self._motion_mat.T)
        covariance = np.linalg.multi_dot((
            self._motion_mat, covariance, self._motion_mat.T)) + motion_cov
        return mean, covariance

    def multi_predict(self, means, covariances):
        """
        参数: means, covariances - 多个状态
        返回: tuple: (new_means, new_covariances) 预测后的状态
        """
        if len(means) == 0:
            return means, covariances
        std_pos = [
            self._std_weight_position * means[:, 3],
            self._std_weight_position * means[:, 3],
            1e-2 * np.ones_like(means[:, 3]),
            self._std_weight_position * means[:, 3]
        ]
        std_vel = [
            self._std_weight_velocity * means[:, 3],
            self._std_weight_velocity * means[:, 3],
            1e-5 * np.ones_like(means[:, 3]),
            self._std_weight_velocity * means[:, 3]
        ]
        sqr = np.square(np.r_[std_pos, std_vel]).T
        motion_covs = []
        for i in range(len(means)):
            motion_covs.append(np.diag(sqr[i]))
        motion_covs = np.asarray(motion_covs)
        new_means = np.dot(means, self._motion_mat.T)
        new_covariances = []
        for i, (mean, cov) in enumerate(zip(means, covariances)):
            new_cov = np.linalg.multi_dot((
                self._motion_mat, cov, self._motion_mat.T)) + motion_covs[i]
            new_covariances.append(new_cov)
        return new_means, np.asarray(new_covariances)

    def update(self, mean, covariance, measurement):
        """
        参数: mean, covariance, measurement - 状态和观测
        返回: tuple: (new_mean, new_covariance) 更新后的状态
        """
        projected_mean, projected_cov = self.project(mean, covariance)
        chol_factor, lower = scipy.linalg.cho_factor(
            projected_cov, lower=True, check_finite=False)
        kalman_gain = scipy.linalg.cho_solve(
            (chol_factor, lower), np.dot(covariance, self._update_mat.T).T,
            check_finite=False).T
        innovation = measurement - projected_mean
        new_mean = mean + np.dot(innovation, kalman_gain.T)
        new_covariance = covariance - np.linalg.multi_dot((
            kalman_gain, projected_cov, kalman_gain.T))
        return new_mean, new_covariance

    def project(self, mean, covariance):
        """
        参数: mean, covariance - 状态
        返回: tuple: (projected_mean, projected_covariance) 投影后的观测
        """
        std = [
            self._std_weight_position * mean[3],
            self._std_weight_position * mean[3],
            1e-1,
            self._std_weight_position * mean[3]
        ]
        innovation_cov = np.diag(np.square(std))
        mean = np.dot(self._update_mat, mean)
        covariance = np.linalg.multi_dot((
            self._update_mat, covariance, self._update_mat.T))
        return mean, covariance + innovation_cov

    def gating_distance(self, mean, covariance, measurements, only_position=False, metric='maha'):
        """
        参数: mean, covariance, measurements, only_position, metric
        返回: np.array: 门控距离数组
        """
        mean, covariance = self.project(mean, covariance)
        if only_position:
            mean, covariance = mean[:2], covariance[:2, :2]
            measurements = measurements[:, :2]
        d = measurements - mean
        if metric == 'gaussian':
            return np.sum(d * d, axis=1)
        elif metric == 'maha':
            cholesky_factor = np.linalg.cholesky(covariance)
            z = scipy.linalg.solve_triangular(
                cholesky_factor, d.T, lower=True, check_finite=False,
                overwrite_b=True)
            squared_maha = np.sum(z * z, axis=0)
            return squared_maha
        else:
            raise ValueError('无效的距离度量')

class STrack(BaseTrack):
    """单个目标跟踪器实现"""

    shared_kalman = KalmanFilter()

    def __init__(self, tlwh, score, temp_feat=None, upper_color=None, lower_color=None, body_ratios=None):
        """
        参数: tlwh, score, temp_feat, upper_color, lower_color, body_ratios
        """
        super(STrack, self).__init__()
        self._tlwh = np.asarray(tlwh, dtype=np.float64)
        self.kalman_filter = None
        self.mean, self.covariance = None, None
        self.predicted_mean, self.predicted_covariance = None, None
        self.is_activated = False
        self.score = score
        self.tracklet_len = 0
        self.upper_color = upper_color
        self.lower_color = lower_color
        self.body_ratios = body_ratios if body_ratios is not None else [0.0] * 16

    def predict(self):
        """预测下一个状态"""
        if self.kalman_filter is None or self.mean is None or self.covariance is None:
            return
        if self.state != TrackState.TRACKED:
            mean_state = self.mean.copy()
            mean_state[7] = 0
            self.mean, self.covariance = self.kalman_filter.predict(mean_state, self.covariance)
        else:
            self.mean, self.covariance = self.kalman_filter.predict(self.mean, self.covariance)
        self.predicted_mean = self.mean.copy()
        self.predicted_covariance = self.covariance.copy()

    @staticmethod
    def multi_predict(stracks):
        """
        参数: stracks (list): STrack对象列表
        """
        if len(stracks) > 0:
            multi_mean = np.asarray([st.mean.copy() for st in stracks])
            multi_covariance = np.asarray([st.covariance for st in stracks])
            for i, st in enumerate(stracks):
                if st.state != TrackState.TRACKED:
                    multi_mean[i][7] = 0
            multi_mean, multi_covariance = STrack.shared_kalman.multi_predict(multi_mean, multi_covariance)
            for i, (mean, cov) in enumerate(zip(multi_mean, multi_covariance)):
                stracks[i].mean = mean
                stracks[i].covariance = cov
                stracks[i].predicted_mean = mean.copy()
                stracks[i].predicted_covariance = cov.copy()

    def activate(self, kalman_filter, frame_id):
        """
        参数: kalman_filter, frame_id
        """
        self.kalman_filter = kalman_filter
        self.track_id = self.next_id()
        self.mean, self.covariance = self.kalman_filter.initiate(self.tlwh_to_xyah(self._tlwh))
        self.tracklet_len = 0
        self.state = TrackState.TRACKED
        if frame_id == 1:
            self.is_activated = True
        self.frame_id = frame_id
        self.start_frame = frame_id

    def re_activate(self, new_track, frame_id, new_id=False):
        """
        参数: new_track, frame_id, new_id
        """
        if self.kalman_filter is None:
            return
        self.mean, self.covariance = self.kalman_filter.update(
            self.mean, self.covariance, self.tlwh_to_xyah(new_track.tlwh)
        )
        self.tracklet_len = 0
        self.state = TrackState.TRACKED
        self.is_activated = True
        self.frame_id = frame_id
        if new_track.upper_color is not None and all(c >= 0 for c in new_track.upper_color):
            self.upper_color = new_track.upper_color
        if new_track.lower_color is not None and all(c >= 0 for c in new_track.lower_color):
            self.lower_color = new_track.lower_color
        if hasattr(new_track, 'body_ratios') and new_track.body_ratios:
            if len(new_track.body_ratios) == len(self.body_ratios):
                self.body_ratios = new_track.body_ratios
        if new_id:
            self.track_id = self.next_id()
        self.score = new_track.score

    def update(self, new_track, frame_id):
        """
        参数: new_track, frame_id
        """
        self.frame_id = frame_id
        self.tracklet_len += 1
        if self.kalman_filter is not None:
            new_tlwh = new_track.tlwh
            self.mean, self.covariance = self.kalman_filter.update(
                self.mean, self.covariance, self.tlwh_to_xyah(new_tlwh)
            )
        self.state = TrackState.TRACKED
        self.is_activated = True
        self.score = new_track.score
        if new_track.upper_color is not None and all(c >= 0 for c in new_track.upper_color):
            if self.upper_color is not None:
                weight = min(0.3, 1.0 / self.tracklet_len)
                self.upper_color = tuple(int((1 - weight) * c1 + weight * c2)
                                         for c1, c2 in zip(self.upper_color, new_track.upper_color))
            else:
                self.upper_color = new_track.upper_color
        if new_track.lower_color is not None and all(c >= 0 for c in new_track.lower_color):
            if self.lower_color is not None:
                weight = min(0.3, 1.0 / self.tracklet_len)
                self.lower_color = tuple(int((1 - weight) * c1 + weight * c2)
                                         for c1, c2 in zip(self.lower_color, new_track.lower_color))
            else:
                self.lower_color = new_track.lower_color
        if hasattr(new_track, 'body_ratios') and new_track.body_ratios:
            if len(new_track.body_ratios) == len(self.body_ratios):
                weight = min(0.2, 1.0 / self.tracklet_len)
                for i in range(len(self.body_ratios)):
                    if new_track.body_ratios[i] > 0:
                        self.body_ratios[i] = (1 - weight) * self.body_ratios[i] + weight * new_track.body_ratios[i]

    @property
    def tlwh(self):
        """返回: np.array: [x,y,w,h] 格式的边界框"""
        if self.mean is None:
            return self._tlwh.copy()
        ret = self.mean[:4].copy()
        ret[2] *= ret[3]
        ret[:2] -= ret[2:] / 2
        return ret

    @property
    def predicted_tlwh(self):
        """返回: np.array: 预测的边界框"""
        if self.predicted_mean is None:
            return self.tlwh.copy()
        ret = self.predicted_mean[:4].copy()
        ret[2] *= ret[3]
        ret[:2] -= ret[2:] / 2
        return ret

    @property
    def predicted_tlbr(self):
        """返回: np.array: 预测的边界框，左上角和右下角坐标"""
        ret = self.predicted_tlwh.copy()
        ret[2:] += ret[:2]
        return ret

    @property
    def tlbr(self):
        """返回: np.array: [x1,y1,x2,y2] 格式的边界框"""
        ret = self.tlwh.copy()
        ret[2:] += ret[:2]
        return ret

    @staticmethod
    def tlwh_to_xyah(tlwh):
        """
        参数: tlwh (np.array): [x,y,w,h] 格式边界框
        返回: np.array: [cx,cy,aspect,h] 格式
        """
        ret = np.asarray(tlwh).copy()
        ret[:2] += ret[2:] / 2
        ret[2] /= ret[3]
        return ret

    def to_xyah(self):
        """返回: np.array: [cx,cy,aspect,h] 格式的状态"""
        return self.tlwh_to_xyah(self.tlwh)

    @staticmethod
    def tlbr_to_tlwh(tlbr):
        """
        参数: tlbr (np.array): [x1,y1,x2,y2] 格式边界框
        返回: np.array: [x,y,w,h] 格式边界框
        """
        ret = np.asarray(tlbr).copy()
        ret[2:] -= ret[:2]
        return ret

    @staticmethod
    def tlwh_to_tlbr(tlwh):
        """
        参数: tlwh (np.array): [x,y,w,h] 格式边界框
        返回: np.array: [x1,y1,x2,y2] 格式边界框
        """
        ret = np.asarray(tlwh).copy()
        ret[2:] += ret[:2]
        return ret

    def __repr__(self):
        """返回: str: 轨迹的字符串描述"""
        return f'OT_{self.track_id}({self.start_frame}-{self.end_frame})'

# ==================== 辅助函数 ====================

def iou_distance(atracks, btracks):
    """
    参数: atracks, btracks - 轨迹或边界框列表
    返回: np.array: IOU距离矩阵
    """
    if (len(atracks) > 0 and isinstance(atracks[0], np.ndarray)) or \
            (len(btracks) > 0 and isinstance(btracks[0], np.ndarray)):
        atlbrs = atracks
        btlbrs = btracks
    else:
        atlbrs = [track.tlbr for track in atracks]
        btlbrs = [track.tlbr for track in btracks]
    ious = optimized_bbox_ious(np.array(atlbrs), np.array(btlbrs))
    cost_matrix = 1 - ious
    return cost_matrix

def color_distance(atracks, btracks, alpha=0.5):
    """
    参数: atracks, btracks - 轨迹列表, alpha - 上衣颜色权重
    返回: np.array: 颜色距离矩阵
    """
    cost_matrix = np.zeros((len(atracks), len(btracks)), dtype=np.float64)
    for i, atrack in enumerate(atracks):
        for j, btrack in enumerate(btracks):
            upper_dist = 1.0
            if hasattr(atrack, 'upper_color') and hasattr(btrack, 'upper_color') and \
                    atrack.upper_color is not None and btrack.upper_color is not None:
                upper_dist = np.sqrt(np.sum([(a - b) ** 2 for a, b in zip(atrack.upper_color, btrack.upper_color)]))
                upper_dist = min(upper_dist / 442.0, 1.0)
            lower_dist = 1.0
            if hasattr(atrack, 'lower_color') and hasattr(btrack, 'lower_color') and \
                    atrack.lower_color is not None and btrack.lower_color is not None:
                lower_dist = np.sqrt(np.sum([(a - b) ** 2 for a, b in zip(atrack.lower_color, btrack.lower_color)]))
                lower_dist = min(lower_dist / 442.0, 1.0)
            if upper_dist < 1.0 and lower_dist < 1.0:
                cost_matrix[i, j] = alpha * upper_dist + (1 - alpha) * lower_dist
            elif upper_dist < 1.0:
                cost_matrix[i, j] = upper_dist
            elif lower_dist < 1.0:
                cost_matrix[i, j] = lower_dist
            else:
                cost_matrix[i, j] = 1.0
    return cost_matrix

def fuse_motion(kf, cost_matrix, tracks, detections, lambda_=0.98):
    """
    参数: kf, cost_matrix, tracks, detections, lambda_
    返回: np.array: 融合运动信息后的代价矩阵
    """
    if cost_matrix.size == 0:
        return cost_matrix
    gating_threshold = kf.chi2inv95[4]
    measurements = np.asarray([det.tlwh_to_xyah(det.tlwh) for det in detections])
    for row, track in enumerate(tracks):
        gating_distance = kf.gating_distance(
            track.mean, track.covariance, measurements)
        cost_matrix[row, gating_distance > gating_threshold] = np.inf
        cost_matrix[row] = lambda_ * cost_matrix[row] + (1 - lambda_) * gating_distance
    return cost_matrix

def fuse_iou_with_color(iou_cost, color_cost, detections, w_iou=0.7, w_color=0.3):
    """
    参数: iou_cost, color_cost, detections, w_iou, w_color
    返回: np.array: 融合后的代价矩阵
    """
    if iou_cost.size == 0:
        return iou_cost
    if color_cost.size == 0 or color_cost.shape != iou_cost.shape:
        w_iou, w_color = 1.0, 0.0
    iou_sim = 1 - iou_cost
    color_sim = 1 - color_cost
    det_scores = np.array([det.score for det in detections])
    det_scores = np.expand_dims(det_scores, axis=0).repeat(iou_cost.shape[0], axis=0)
    if w_color > 0:
        fuse_sim = w_iou * iou_sim + w_color * color_sim
        fuse_sim = fuse_sim * (1 + det_scores) / 2
    else:
        fuse_sim = iou_sim * (1 + det_scores) / 2
    fuse_cost = 1 - fuse_sim
    return fuse_cost

def linear_assignment(cost_matrix, thresh=0.7):
    """
    参数: cost_matrix, thresh
    返回: tuple: (matches, unmatched_a, unmatched_b)
    """
    try:
        import lap
        if cost_matrix.size == 0:
            return [], list(range(cost_matrix.shape[0])), list(range(cost_matrix.shape[1]))
        matches, unmatched_a, unmatched_b = [], [], []
        cost, x, y = lap.lapjv(cost_matrix, extend_cost=True, cost_limit=thresh)
        for ix, mx in enumerate(x):
            if mx >= 0:
                matches.append([ix, mx])
        unmatched_a = np.where(x < 0)[0].tolist()
        unmatched_b = np.where(y < 0)[0].tolist()
        return matches, unmatched_a, unmatched_b
    except ImportError:
        from scipy.optimize import linear_sum_assignment
        cost_matrix = np.copy(cost_matrix)
        cost_matrix[cost_matrix > thresh] = thresh + 1e-4
        row_ind, col_ind = linear_sum_assignment(cost_matrix)
        matches = []
        unmatched_a = list(range(cost_matrix.shape[0]))
        unmatched_b = list(range(cost_matrix.shape[1]))
        for r, c in zip(row_ind, col_ind):
            if cost_matrix[r, c] <= thresh:
                matches.append([r, c])
                unmatched_a.remove(r)
                unmatched_b.remove(c)
        return matches, unmatched_a, unmatched_b

def optimized_bbox_ious(atlbrs, btlbrs):
    """
    参数: atlbrs, btlbrs - 边界框数组
    返回: np.array: IOU矩阵
    """
    atlbrs = np.asarray(atlbrs)
    btlbrs = np.asarray(btlbrs)
    alen, blen = len(atlbrs), len(btlbrs)
    if alen == 0 or blen == 0:
        return np.zeros((alen, blen), dtype=np.float64)
    xx1 = np.maximum(atlbrs[:, 0].reshape(alen, 1), btlbrs[:, 0].reshape(1, blen))
    yy1 = np.maximum(atlbrs[:, 1].reshape(alen, 1), btlbrs[:, 1].reshape(1, blen))
    xx2 = np.minimum(atlbrs[:, 2].reshape(alen, 1), btlbrs[:, 2].reshape(1, blen))
    yy2 = np.minimum(atlbrs[:, 3].reshape(alen, 1), btlbrs[:, 3].reshape(1, blen))
    w = np.maximum(0, xx2 - xx1)
    h = np.maximum(0, yy2 - yy1)
    inter = w * h
    area_a = ((atlbrs[:, 2] - atlbrs[:, 0]) * (atlbrs[:, 3] - atlbrs[:, 1])).reshape(alen, 1)
    area_b = ((btlbrs[:, 2] - btlbrs[:, 0]) * (btlbrs[:, 3] - btlbrs[:, 1])).reshape(1, blen)
    union = area_a + area_b - inter
    ious = np.where(union > 0, inter / union, 0)
    return ious

def joint_stracks(tlista, tlistb):
    """
    参数: tlista, tlistb - 轨迹列表
    返回: list: 合并后的轨迹列表
    """
    exists = {}
    res = []
    for t in tlista:
        exists[t.track_id] = 1
        res.append(t)
    for t in tlistb:
        tid = t.track_id
        if not exists.get(tid, 0):
            exists[tid] = 1
            res.append(t)
    return res

def sub_stracks(tlista, tlistb):
    """
    参数: tlista, tlistb - 轨迹列表
    返回: list: 差集轨迹列表
    """
    stracks = {}
    for t in tlista:
        stracks[t.track_id] = t
    for t in tlistb:
        tid = t.track_id
        if stracks.get(tid, 0):
            del stracks[tid]
    return list(stracks.values())

def remove_duplicate_stracks(stracksa, stracksb):
    """
    参数: stracksa, stracksb - 轨迹列表
    返回: tuple: (resa, resb) - 去重后的轨迹列表
    """
    pdist = iou_distance(stracksa, stracksb)
    pairs = np.where(pdist < 0.15)
    dupa, dupb = list(), list()
    for p, q in zip(*pairs):
        timep = stracksa[p].frame_id - stracksa[p].start_frame
        timeq = stracksb[q].frame_id - stracksb[q].start_frame
        if timep > timeq:
            dupb.append(q)
        else:
            dupa.append(p)
    resa = [t for i, t in enumerate(stracksa) if i not in dupa]
    resb = [t for i, t in enumerate(stracksb) if i not in dupb]
    return resa, resb

# ==================== ByteTracker核心算法类 ====================

class ColorByteTracker:
    """融合颜色特征的ByteTracker算法实现"""

    def __init__(self, args, frame_rate=30):
        """
        参数: args - 配置字典, frame_rate - 帧率
        """
        self.tracked_stracks = []
        self.lost_stracks = []
        self.removed_stracks = []
        self.frame_id = 0
        self.args = args
        self.det_thresh = args.get('track_thresh', 0.5) + 0.1
        self.buffer_size = int(frame_rate / 30.0 * args.get('track_buffer', 30))
        self.max_time_lost = self.buffer_size
        self.kalman_filter = KalmanFilter()
        self.color_weight = args.get('color_weight', 0.3)

    def update(self, detection_results):
        """
        参数: detection_results - 检测结果列表
        返回: list: 当前活跃的跟踪轨迹
        """
        self.frame_id += 1
        activated_starcks = []
        refind_stracks = []
        lost_stracks = []
        removed_stracks = []

        if not detection_results:
            for track in self.tracked_stracks:
                if track.state == TrackState.TRACKED:
                    track.mark_lost()
                    lost_stracks.append(track)
            self._update_status(activated_starcks, refind_stracks, lost_stracks, removed_stracks)
            return [track for track in self.tracked_stracks if track.is_activated]

        detections = []
        low_detections = []

        for result in detection_results:
            try:
                if len(result) >= 5:
                    full_bbox, upper_color, lower_color, confidence, body_ratios = result
                else:
                    full_bbox, upper_color, lower_color, confidence = result
                    body_ratios = None

                # 直接使用消息中提供的完整边界框和置信度
                strack = STrack(
                    STrack.tlbr_to_tlwh(np.array(full_bbox)),
                    confidence,
                    upper_color=upper_color,
                    lower_color=lower_color,
                    body_ratios=body_ratios
                )

                if confidence >= self.args.get('track_thresh', 0.5):
                    detections.append(strack)
                elif confidence >= 0.1:
                    low_detections.append(strack)

            except Exception as e:
                print(f"update error {result}")
                continue

        unconfirmed = []
        tracked_stracks = []

        for track in self.tracked_stracks:
            if not track.is_activated:
                unconfirmed.append(track)
            else:
                tracked_stracks.append(track)

        strack_pool = joint_stracks(tracked_stracks, self.lost_stracks)
        STrack.multi_predict(strack_pool)

        # 第一阶段关联
        iou_dists = iou_distance(strack_pool, detections)
        if self.color_weight > 0:
            color_dists = color_distance(strack_pool, detections)
            dists = fuse_iou_with_color(
                iou_dists, color_dists, detections,
                w_iou=1 - self.color_weight, w_color=self.color_weight
            )
        else:
            dists = iou_dists

        matches, u_track, u_detection = linear_assignment(
            dists, thresh=self.args.get('match_thresh', 0.8)
        )

        for i_track, i_det in matches:
            track = strack_pool[i_track]
            det = detections[i_det]
            if track.state == TrackState.TRACKED:
                track.update(det, self.frame_id)
                activated_starcks.append(track)
            else:
                track.re_activate(det, self.frame_id, new_id=False)
                refind_stracks.append(track)

        # 第二阶段关联
        if len(low_detections) > 0:
            r_tracked_stracks = [strack_pool[i] for i in u_track if strack_pool[i].state == TrackState.TRACKED]
            iou_dists = iou_distance(r_tracked_stracks, low_detections)
            if self.color_weight > 0:
                color_dists = color_distance(r_tracked_stracks, low_detections)
                dists = fuse_iou_with_color(
                    iou_dists, color_dists, low_detections,
                    w_iou=1 - self.color_weight, w_color=self.color_weight
                )
            else:
                dists = iou_dists

            matches, u_tracks, u_low_detection = linear_assignment(dists, thresh=0.5)

            for i_track, i_det in matches:
                track = r_tracked_stracks[i_track]
                det = low_detections[i_det]
                if track.state == TrackState.TRACKED:
                    track.update(det, self.frame_id)
                    activated_starcks.append(track)
                else:
                    track.re_activate(det, self.frame_id, new_id=False)
                    refind_stracks.append(track)

            for i in u_tracks:
                track = r_tracked_stracks[i]
                if track.state != TrackState.LOST:
                    track.mark_lost()
                    lost_stracks.append(track)
        else:
            for i in u_track:
                track = strack_pool[i]
                if track.state != TrackState.LOST:
                    track.mark_lost()
                    lost_stracks.append(track)

        # 处理未确认轨迹
        u_detection_track = [detections[i] for i in u_detection]
        iou_dists = iou_distance(unconfirmed, u_detection_track)
        if self.color_weight > 0:
            color_dists = color_distance(unconfirmed, u_detection_track)
            dists = fuse_iou_with_color(
                iou_dists, color_dists, u_detection_track,
                w_iou=1 - self.color_weight, w_color=self.color_weight
            )
        else:
            dists = iou_dists

        matches, u_unconfirmed, u_detection = linear_assignment(dists, thresh=0.7)
        for i_track, i_det in matches:
            unconfirmed[i_track].update(u_detection_track[i_det], self.frame_id)
            activated_starcks.append(unconfirmed[i_track])

        for i in u_unconfirmed:
            track = unconfirmed[i]
            track.mark_removed()
            removed_stracks.append(track)

        # 创建新轨迹
        for i in u_detection:
            track = u_detection_track[i]
            if track.score < self.det_thresh:
                continue
            track.activate(self.kalman_filter, self.frame_id)
            activated_starcks.append(track)

        for track in self.lost_stracks:
            if self.frame_id - track.end_frame > self.max_time_lost:
                track.mark_removed()
                removed_stracks.append(track)

        self._update_status(activated_starcks, refind_stracks, lost_stracks, removed_stracks)
        return [track for track in self.tracked_stracks if track.is_activated]

    def _update_status(self, activated_starcks, refind_stracks, lost_stracks, removed_stracks):
        """更新跟踪器内部状态"""
        self.tracked_stracks = [t for t in self.tracked_stracks if t.state == TrackState.TRACKED]
        self.tracked_stracks = joint_stracks(self.tracked_stracks, activated_starcks)
        self.tracked_stracks = joint_stracks(self.tracked_stracks, refind_stracks)
        self.lost_stracks = sub_stracks(self.lost_stracks, self.tracked_stracks)
        self.lost_stracks.extend(lost_stracks)
        self.lost_stracks = sub_stracks(self.lost_stracks, self.removed_stracks)
        self.removed_stracks.extend(removed_stracks)
        self.tracked_stracks, self.lost_stracks = remove_duplicate_stracks(
            self.tracked_stracks, self.lost_stracks)

# ==================== 单目标跟踪器 ====================

class SingleTargetTracker:
    """单目标跟踪器"""

    MODE_SELECTING = 0  # 目标选择模式
    MODE_TRACKING = 1   # 目标跟踪模式

    def __init__(self, tracker_args, target_features, max_lost_time=30):
        """
        参数: tracker_args, target_features, max_lost_time
        """
        self.base_tracker = ColorByteTracker(tracker_args)
        self.target_body_ratios, self.target_shirt_color, self.target_pants_color = target_features
        self.mode = self.MODE_SELECTING
        self.target_id = None
        self.lost_frames = 0
        self.max_lost_frames = max_lost_time
        self.frame_count = 0
        self.target_score_history = []
        self.confirmed_target_position = None
        self.target_trajectory = deque(maxlen=30)

    def body_ratio_similarity(self, detected_ratios):
        """
        参数: detected_ratios - 检测到的身体比例
        返回: float: 相似度，0-1之间
        """
        if not detected_ratios or len(detected_ratios) != len(self.target_body_ratios):
            return 0.0
        try:
            target = np.array(self.target_body_ratios)
            detected = np.array(detected_ratios)
            if np.all(target == 0) or np.all(detected == 0):
                return 0.0
            dot_product = np.dot(target, detected)
            norm_product = np.linalg.norm(target) * np.linalg.norm(detected)
            if norm_product == 0:
                return 0.0
            similarity = dot_product / norm_product
            similarity = (similarity + 1) / 2
            return similarity
        except Exception:
            return 0.0

    def color_similarity(self, color1, color2):
        """
        参数: color1, color2 - BGR颜色元组
        返回: float: 相似度，0-1之间
        """
        if not color1 or not color2:
            return 0.0
        try:
            c1 = np.array(color1)
            c2 = np.array(color2)
            distance = np.sqrt(np.sum((c1 - c2) ** 2))
            max_distance = 442.0
            similarity = 1.0 - min(distance / max_distance, 1.0)
            return similarity
        except Exception:
            return 0.0

    def calculate_target_score(self, track):
        """
        参数: track - 跟踪对象
        返回: float: 得分，0-1之间
        """
        shirt_color = track.upper_color if hasattr(track, 'upper_color') and track.upper_color else None
        pants_color = track.lower_color if hasattr(track, 'lower_color') and track.lower_color else None
        shirt_sim = self.color_similarity(self.target_shirt_color, shirt_color) if shirt_color else 0.0
        pants_sim = self.color_similarity(self.target_pants_color, pants_color) if pants_color else 0.0
        body_ratio_sim = 0.0
        has_body_ratios = False
        if hasattr(track, 'body_ratios') and track.body_ratios:
            valid_ratios = sum(1 for r in track.body_ratios if r > 0)
            if valid_ratios > 4:
                body_ratio_sim = self.body_ratio_similarity(track.body_ratios)
                has_body_ratios = True
        feature_count = 0
        weighted_sum = 0.0
        if shirt_color:
            weighted_sum += 0.4 * shirt_sim
            feature_count += 1
        if pants_color:
            weighted_sum += 0.3 * pants_sim
            feature_count += 1
        if has_body_ratios:
            weighted_sum += 0.3 * body_ratio_sim
            feature_count += 1
        if feature_count > 0:
            base_score = weighted_sum / feature_count * (0.8 + 0.2 * feature_count)
            if feature_count == 3:
                base_score *= 1.05
            if feature_count == 1:
                base_score *= 0.9
        else:
            base_score = 0.0
        if self.mode == self.MODE_TRACKING and self.target_id == track.track_id:
            base_score = base_score * 1.2
        return min(base_score, 1.0)

    def select_target(self, tracks):
        """
        参数: tracks - 当前跟踪的轨迹列表
        返回: tuple: (best_track, score)
        """
        if not tracks:
            return None, 0.0
        best_track = None
        best_score = 0.0
        for track in tracks:
            score = self.calculate_target_score(track)
            if score > best_score:
                best_score = score
                best_track = track
        min_score_threshold = 0.4
        if best_score >= min_score_threshold:
            return best_track, best_score
        else:
            return None, best_score

    def update(self, detection_results):
        """
        参数: detection_results - 检测结果列表
        返回: tuple: (track_results, target_track, mode)
        """
        self.frame_count += 1
        all_tracks = self.base_tracker.update(detection_results)
        if not all_tracks:
            if self.mode == self.MODE_TRACKING:
                self.lost_frames += 1
                if self.lost_frames > self.max_lost_frames:
                    self.mode = self.MODE_SELECTING
                    self.target_id = None
                    self.target_trajectory.clear()
            return all_tracks, None, self.mode

        if self.mode == self.MODE_SELECTING:
            best_track, score = self.select_target(all_tracks)
            self.target_score_history.append(score)
            if len(self.target_score_history) > 10:
                self.target_score_history.pop(0)
            if best_track and len(self.target_score_history) >= 3:
                recent_scores = self.target_score_history[-3:]
                if all(s >= 0.4 for s in recent_scores):
                    self.target_id = best_track.track_id
                    self.mode = self.MODE_TRACKING
                    self.lost_frames = 0
                    self.target_trajectory.clear()
                    center_x = (best_track.tlbr[0] + best_track.tlbr[2]) / 2
                    center_y = (best_track.tlbr[1] + best_track.tlbr[3]) / 2
                    self.target_trajectory.append((int(center_x), int(center_y)))
            return all_tracks, best_track, self.mode
        else:
            target_track = None
            for track in all_tracks:
                if track.track_id == self.target_id:
                    target_track = track
                    self.lost_frames = 0
                    center_x = (track.tlbr[0] + track.tlbr[2]) / 2
                    center_y = (track.tlbr[1] + track.tlbr[3]) / 2
                    self.target_trajectory.append((int(center_x), int(center_y)))
                    break
            if target_track is None:
                self.lost_frames += 1
                if self.lost_frames > self.max_lost_frames:
                    self.mode = self.MODE_SELECTING
                    self.target_id = None
                    self.target_trajectory.clear()
            return all_tracks, target_track, self.mode

# ==================== ByteTracker ROS2节点 ====================

class ByteTrackerNode(Node):
    """简化版ByteTracker ROS2节点"""

    def __init__(self):
        super().__init__('bytetracker_node')
        self.lock = threading.Lock()
        
        # 图像处理相关
        self.cv_bridge = CvBridge()
        self.current_image = None
        self.image_lock = threading.Lock()
        
        # 跟踪器参数
        self.tracker_args = {
            'track_thresh': 0.5,
            'track_buffer': 100,
            'match_thresh': 0.8,
            'color_weight': 0.5
        }

        # 跟踪模式
        self.tracking_mode = 'multi'
        self.single_target_tracker = None
        self.multi_target_tracker = None
        self.target_person_name = ""

        # 初始化多目标跟踪器
        self.multi_target_tracker = ColorByteTracker(self.tracker_args)

        # 当前数据和处理结果
        self.current_detection_data = []
        self.current_tracks = []
        self.current_target_track = None
        self.processing = False

        # 统计信息
        self.frame_count = 0
        self.processing_time = 0
        self.current_fps = 0.0

        # 设置参数
        self.setup_parameters()

        # 创建发布者和订阅者
        self.setup_publishers()
        self.setup_subscribers()

        # 创建定时器
        self.setup_timers()

        self.get_logger().info('✅ 简化版ByteTracker节点初始化完成')
        self.get_logger().info(f'🎮 当前跟踪模式: {self.tracking_mode}')

        if self.tracking_mode == 'single':
            self.get_logger().info('🎯 检测到单目标模式，正在初始化单目标跟踪器...')
            self.init_single_target_tracker()

    def setup_parameters(self):
        """设置节点参数"""
        self.declare_parameter('tracking_mode', 'multi')
        self.declare_parameter('track_thresh', 0.5)
        self.declare_parameter('track_buffer', 100)
        self.declare_parameter('match_thresh', 0.8)
        self.declare_parameter('color_weight', 0.5)
        self.declare_parameter('target_features_file', '')
        self.declare_parameter('image_topic', '/camera/color/image_raw')
        self.declare_parameter('enable_display', True)

        self.tracking_mode = self.get_parameter('tracking_mode').value or 'multi'
        self.tracker_args['track_thresh'] = self.get_parameter('track_thresh').value
        self.tracker_args['track_buffer'] = self.get_parameter('track_buffer').value
        self.tracker_args['match_thresh'] = self.get_parameter('match_thresh').value
        self.tracker_args['color_weight'] = self.get_parameter('color_weight').value
        self.target_features_file = self.get_parameter('target_features_file').value or ''
        self.image_topic = self.get_parameter('image_topic').value
        self.enable_display = self.get_parameter('enable_display').value

    def setup_publishers(self):
        """设置发布者"""
        qos = QoSProfile(
            reliability=ReliabilityPolicy.RELIABLE,
            history=HistoryPolicy.KEEP_LAST,
            depth=10
        )

        self.tracked_persons_pub = self.create_publisher(
            TrackedPersonArray, '/bytetracker/tracked_persons', qos)
        self.tracking_result_pub = self.create_publisher(
            TrackingResult, '/bytetracker/tracking_result', qos)
        self.detailed_tracking_pub = self.create_publisher(
            String, '/bytetracker/detailed_tracking_data', qos)
        self.status_pub = self.create_publisher(
            String, '/bytetracker/status', qos)

    def setup_subscribers(self):
        """设置订阅者"""
        qos = QoSProfile(
            reliability=ReliabilityPolicy.BEST_EFFORT,
            history=HistoryPolicy.KEEP_LAST,
            depth=10
        )

        # 人员检测数据订阅者 - 订阅integrated_person_detection_node的输出
        self.person_positions_sub = self.create_subscription(
            String, '/person_detection/person_positions', 
            self.person_positions_callback, qos)
        
        # 图像订阅者
        if hasattr(self, 'image_topic') and self.image_topic:
            self.image_sub = self.create_subscription(
                Image, self.image_topic, 
                self.image_callback, qos)
        
        # 模式控制订阅者
        self.mode_sub = self.create_subscription(
            String, '/bytetracker/set_mode', self.mode_callback, qos)

        # 目标人物设置订阅者
        self.target_sub = self.create_subscription(
            String, '/bytetracker/set_target', self.target_callback, qos)

    def setup_timers(self):
        """设置定时器"""
        self.process_timer = self.create_timer(0.05, self.process_frame)
        self.status_timer = self.create_timer(0.5, self.publish_status)

    def person_positions_callback(self, msg):
        """
        参数: msg (String): 人员检测数据消息
        """
        with self.lock:
            self.current_detection_data = self.parse_detection_message(msg)
    
    def image_callback(self, msg):
        """
        参数: msg (Image): 图像消息
        """
        try:
            with self.image_lock:
                self.current_image = self.cv_bridge.imgmsg_to_cv2(msg, "bgr8")
        except Exception as e:
            self.get_logger().error(f"图像转换错误: {e}")

    def parse_detection_message(self, msg):
        """
        参数: msg (String): 人员检测数据消息
        返回: list: 解析后的检测结果列表，格式为(full_bbox, upper_color, lower_color, confidence, body_ratios)
        """
        detection_results = []
        try:
            # 解析JSON数据
            data = json.loads(msg.data)
            
            if 'persons' not in data:
                self.get_logger().warn("JSON数据中未找到'persons'字段")
                return detection_results
            
            for person in data['persons']:
                # 提取边界框 - 格式为[x1, y1, x2, y2]
                if 'bbox' in person and len(person['bbox']) >= 4:
                    full_bbox = person['bbox']  # 已经是[x1, y1, x2, y2]格式
                else:
                    self.get_logger().warn(f"人员 {person.get('id', '未知')} 缺少有效的边界框信息")
                    continue
                
                # 提取服装颜色信息
                upper_color = None
                lower_color = None
                upper_confidence = 0.0
                lower_confidence = 0.0
                
                if 'clothing' in person:
                    clothing = person['clothing']
                    
                    # 提取上衣颜色和置信度
                    if 'upper' in clothing:
                        upper_data = clothing['upper']
                        # 直接使用RGB值数组 [R, G, B]，转换为BGR元组用于OpenCV
                        if 'color_rgb' in upper_data and len(upper_data['color_rgb']) >= 3:
                            rgb = upper_data['color_rgb']
                            upper_color = (rgb[2], rgb[1], rgb[0])  # RGB转BGR
                        upper_confidence = float(upper_data.get('confidence', 0.0))
                    
                    # 提取下装颜色和置信度
                    if 'lower' in clothing:
                        lower_data = clothing['lower']
                        # 直接使用RGB值数组 [R, G, B]，转换为BGR元组用于OpenCV
                        if 'color_rgb' in lower_data and len(lower_data['color_rgb']) >= 3:
                            rgb = lower_data['color_rgb']
                            lower_color = (rgb[2], rgb[1], rgb[0])  # RGB转BGR
                        lower_confidence = float(lower_data.get('confidence', 0.0))
                
                # 计算总置信度（上衣和下衣置信度之和，如用户要求）
                total_confidence = upper_confidence + lower_confidence
                
                # 提取身体比例数据
                body_ratios = None
                if 'body_ratios' in person and person['body_ratios']:
                    body_ratios = [float(ratio) for ratio in person['body_ratios']]
                    # 确保有16个比例值
                    if len(body_ratios) < 16:
                        body_ratios.extend([0.0] * (16 - len(body_ratios)))
                    elif len(body_ratios) > 16:
                        body_ratios = body_ratios[:16]
                
                detection_results.append((
                    full_bbox, upper_color, lower_color, total_confidence, body_ratios
                ))
                
                self.get_logger().debug(f"解析人员 {person.get('id', '未知')}: "
                                      f"边界框={full_bbox}, "
                                      f"上衣RGB={upper_color}, 下装RGB={lower_color}, "
                                      f"总置信度={total_confidence:.2f}")
                
        except json.JSONDecodeError as e:
            self.get_logger().error(f"JSON解析错误: {e}")
        except Exception as e:
            self.get_logger().error(f"解析人员检测消息错误: {e}")
            import traceback
            self.get_logger().error(f"详细错误信息: {traceback.format_exc()}")
            
        return detection_results

    def process_frame(self):
        """帧处理主循环"""
        with self.lock:
            if not self.current_detection_data or self.processing:
                return
            self.processing = True
            detection_data = self.current_detection_data.copy()
            self.frame_count += 1

        start_time = time.time()
        
        try:
            if self.tracking_mode == 'multi':
                if self.multi_target_tracker is not None:
                    tracks = self.multi_target_tracker.update(detection_data)
                else:
                    tracks = []
                target_track = None
                mode = 'multi'
            else:
                if self.single_target_tracker is not None:
                    all_tracks, target_track, mode = self.single_target_tracker.update(detection_data)
                    tracks = all_tracks
                else:
                    tracks = []
                    target_track = None
                    mode = 'single_not_initialized'

            self.current_tracks = tracks
            self.current_target_track = target_track

            self.publish_tracking_results(tracks, target_track)
            self.publish_tracking_result(tracks, target_track, mode)
            self.publish_detailed_tracking_data(tracks, target_track, mode)
            
            # 显示带有跟踪结果的图像
            if self.enable_display:
                self.display_tracking_results(tracks, target_track)

            total_time = time.time() - start_time
            fps = 1.0 / total_time if total_time > 0 else 0
            self.processing_time = total_time
            
            if not hasattr(self, 'fps_window'):
                self.fps_window = []
                self.fps_window_size = 10
                
            self.fps_window.append(fps)
            if len(self.fps_window) > self.fps_window_size:
                self.fps_window.pop(0)
            
            self.current_fps = sum(self.fps_window) / len(self.fps_window)

        except Exception as e:
            self.get_logger().error(f"❌ 处理第{self.frame_count}帧时发生错误: {e}")
            traceback.print_exc()

        finally:
            with self.lock:
                self.processing = False
    
    def display_tracking_results(self, tracks, target_track):
        """
        显示带有跟踪结果的图像
        参数: tracks - 所有跟踪对象, target_track - 目标跟踪对象
        """
        try:
            with self.image_lock:
                if self.current_image is None:
                    return
                
                # 复制图像以避免修改原始图像
                display_image = self.current_image.copy()
            
            # 绘制所有跟踪框
            for track in tracks:
                if track.state == TrackState.TRACKED:
                    # 判断是否为目标
                    is_target = target_track is not None and track.track_id == target_track.track_id
                    
                    # 目标用红框，其他用绿框
                    color = (0, 0, 255) if is_target else (0, 255, 0)  # BGR格式：红色或绿色
                    thickness = 3 if is_target else 2
                    
                    # 获取边界框
                    tlbr = track.tlbr
                    x1, y1, x2, y2 = int(tlbr[0]), int(tlbr[1]), int(tlbr[2]), int(tlbr[3])
                    
                    # 绘制边界框
                    cv2.rectangle(display_image, (x1, y1), (x2, y2), color, thickness)
                    
                    # 绘制标签
                    label = f"{'TARGET ' if is_target else ''}ID:{track.track_id}"
                    if hasattr(track, 'score'):
                        label += f" ({track.score:.2f})"
                    
                    # 计算标签位置
                    label_size = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)[0]
                    label_y = max(y1 - 10, label_size[1] + 10)
                    
                    # 绘制标签背景
                    cv2.rectangle(display_image, 
                                 (x1, label_y - label_size[1] - 5), 
                                 (x1 + label_size[0] + 5, label_y + 5), 
                                 color, -1)
                    
                    # 绘制标签文本
                    cv2.putText(display_image, label, (x1 + 2, label_y), 
                               cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
                    
                    # 绘制中心点
                    center_x = int((x1 + x2) / 2)
                    center_y = int((y1 + y2) / 2)
                    cv2.circle(display_image, (center_x, center_y), 4, color, -1)
                    
                    # 显示颜色信息（如果有）
                    if hasattr(track, 'upper_color') and track.upper_color:
                        upper_color = track.upper_color
                        cv2.rectangle(display_image, (x1, y2 + 5), (x1 + 30, y2 + 25), upper_color, -1)
                        cv2.putText(display_image, "U", (x1 + 8, y2 + 20), 
                                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
                    
                    if hasattr(track, 'lower_color') and track.lower_color:
                        lower_color = track.lower_color
                        cv2.rectangle(display_image, (x1 + 35, y2 + 5), (x1 + 65, y2 + 25), lower_color, -1)
                        cv2.putText(display_image, "L", (x1 + 43, y2 + 20), 
                                   cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
            
            # 在图像上显示统计信息
            info_text = [
                f"Frame: {self.frame_count}",
                f"Tracks: {len(tracks)}",
                f"Mode: {self.tracking_mode}",
                f"FPS: {self.current_fps:.1f}"
            ]
            
            for i, text in enumerate(info_text):
                y_pos = 30 + i * 25
                cv2.putText(display_image, text, (10, y_pos), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
                cv2.putText(display_image, text, (10, y_pos), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 1)
            
            # 显示目标信息
            if target_track is not None:
                target_info = f"Target ID: {target_track.track_id} (Confidence: {target_track.score:.2f})"
                cv2.putText(display_image, target_info, (10, display_image.shape[0] - 20), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
            
            # 显示图像
            cv2.imshow("ByteTracker - Person Tracking", display_image)
            
            # 检查键盘输入
            key = cv2.waitKey(1) & 0xFF
            if key == ord('q') or key == 27:  # 'q' 或 ESC 键退出
                self.get_logger().info("用户请求退出，正在关闭...")
                rclpy.shutdown()
                
        except Exception as e:
            self.get_logger().error(f"显示跟踪结果错误: {e}")

    def mode_callback(self, msg):
        """
        参数: msg (String): 模式切换消息
        """
        new_mode = msg.data.lower()
        if new_mode in ['multi', 'single']:
            self.tracking_mode = new_mode
            self.get_logger().info(f"🎮 切换到{new_mode}目标跟踪模式")
            if new_mode == 'single' and self.single_target_tracker is None:
                self.init_single_target_tracker()
        else:
            self.get_logger().warn(f"无效的跟踪模式: {new_mode}")

    def target_callback(self, msg):
        """
        参数: msg (String): 目标设置消息
        """
        self.target_person_name = msg.data
        self.get_logger().info(f"🎯 设置目标人物: {self.target_person_name}")
        if self.tracking_mode == 'single':
            self.init_single_target_tracker()

    def init_single_target_tracker(self):
        """
        返回: bool: 初始化是否成功
        """
        try:
            self.get_logger().info('🔧 开始初始化单目标跟踪器...')
            if self.target_features_file:
                self.get_logger().info(f'📖 正在读取目标特征文件: {self.target_features_file}')
                target_features = self.read_target_features(self.target_features_file)
                if target_features:
                    self.single_target_tracker = SingleTargetTracker(
                        self.tracker_args, target_features, max_lost_time=60)
                    self.get_logger().info("✅ 单目标跟踪器初始化成功")
                    return True
                else:
                    self.get_logger().error("❌ 读取目标特征失败")
                    return False
            else:
                self.get_logger().warn("⚠️ 未指定目标特征文件，无法初始化单目标跟踪器")
                return False
        except Exception as e:
            self.get_logger().error(f"❌ 初始化单目标跟踪器错误: {e}")
            traceback.print_exc()
            return False

    def read_target_features(self, xlsx_path):
        """
        参数: xlsx_path (str): Excel文件路径
        返回: tuple: (body_ratios, shirt_color, pants_color) 或 None
        """
        try:
            import openpyxl
            self.get_logger().info(f'🔍 检查特征文件是否存在: {xlsx_path}')
            if not os.path.exists(xlsx_path):
                self.get_logger().error(f"❌ 目标特征文件不存在: {xlsx_path}")
                return None
            
            wb = openpyxl.load_workbook(xlsx_path)
            if wb.active is None:
                self.get_logger().error("❌ Excel文件没有活动工作表")
                return None
                
            sheet = wb.active
            body_ratios = []
            for i in range(1, 17):
                try:
                    cell_value = sheet.cell(row=i, column=1).value
                    if cell_value is not None and str(cell_value).replace('.', '').replace('-', '').isdigit():
                        value = float(str(cell_value))
                    else:
                        value = 0.0
                except (ValueError, TypeError, AttributeError):
                    value = 0.0
                body_ratios.append(value)

            try:
                shirt_color_str = sheet.cell(row=17, column=1).value
            except:
                shirt_color_str = None
                
            try:
                pants_color_str = sheet.cell(row=18, column=1).value
            except:
                pants_color_str = None

            def parse_color(color_str):
                if not color_str:
                    return (0, 0, 0)
                try:
                    if isinstance(color_str, str):
                        color_str = color_str.strip()
                        if color_str.startswith('(') and color_str.endswith(')'):
                            return eval(color_str)
                        else:
                            color_str = color_str.replace('(', '').replace(')', '')
                            values = [int(x.strip()) for x in color_str.split(',')]
                            return tuple(values[:3])
                    elif isinstance(color_str, tuple):
                        return color_str
                    else:
                        return (0, 0, 0)
                except Exception:
                    return (0, 0, 0)

            shirt_color = parse_color(shirt_color_str)
            pants_color = parse_color(pants_color_str)
            self.get_logger().info(f"✅ 成功读取目标特征:")
            self.get_logger().info(f"   📏 身体比例: {len(body_ratios)} 个数值")
            self.get_logger().info(f"   👕 上衣颜色: {shirt_color}")
            self.get_logger().info(f"   👖 下装颜色: {pants_color}")
            return body_ratios, shirt_color, pants_color
        except Exception as e:
            self.get_logger().error(f"❌ 读取目标特征失败: {str(e)}")
            traceback.print_exc()
            return None

    def publish_tracking_results(self, tracks, target_track):
        """
        参数: tracks, target_track - 跟踪结果
        """
        try:
            msg = TrackedPersonArray()
            msg.header.stamp = self.get_clock().now().to_msg()
            msg.header.frame_id = "camera"
            
            # 创建persons列表
            persons_list = []
            
            for track in tracks:
                person = TrackedPerson()
                person.track_id = track.track_id
                
                # 设置边界框（使用sensor_msgs/RegionOfInterest格式）
                person.bbox.x_offset = int(max(0, track.tlbr[0]))
                person.bbox.y_offset = int(max(0, track.tlbr[1]))
                person.bbox.width = int(max(1, track.tlbr[2] - track.tlbr[0]))
                person.bbox.height = int(max(1, track.tlbr[3] - track.tlbr[1]))
                person.bbox.do_rectify = False
                
                # 设置置信度
                person.confidence = float(track.score)
                
                # 设置上装颜色（BGR格式）
                if track.upper_color:
                    person.upper_color = [int(c) for c in track.upper_color]
                else:
                    person.upper_color = []
                
                # 设置下装颜色（BGR格式）
                if track.lower_color:
                    person.lower_color = [int(c) for c in track.lower_color]
                else:
                    person.lower_color = []
                
                # 设置身体比例
                if hasattr(track, 'body_ratios') and track.body_ratios:
                    person.body_ratios = [float(r) for r in track.body_ratios]
                else:
                    person.body_ratios = []
                
                # 设置是否为目标
                person.is_target = (target_track is not None and track.track_id == target_track.track_id)
                
                persons_list.append(person)
            
            # 将persons列表赋值给消息
            msg.persons = persons_list
            
            self.tracked_persons_pub.publish(msg)
            
        except Exception as e:
            self.get_logger().error(f"发布跟踪结果错误: {e}")
            import traceback
            self.get_logger().error(f"详细错误信息: {traceback.format_exc()}")

    def publish_tracking_result(self, tracks, target_track, mode):
        """
        参数: tracks, target_track, mode - 跟踪结果和模式
        """
        try:
            msg = TrackingResult()
            msg.header.stamp = self.get_clock().now().to_msg()
            msg.header.frame_id = "camera"
            msg.mode = 'single_target' if mode in ['selecting', 'tracking'] else 'multi_target'
            msg.total_tracks = len(tracks)
            msg.target_detected = target_track is not None
            if target_track is not None:
                msg.target_id = target_track.track_id
                tlbr = target_track.tlbr
                msg.target_x = float((tlbr[0] + tlbr[2]) / 2)
                msg.target_y = float((tlbr[1] + tlbr[3]) / 2)
                msg.target_width = float(tlbr[2] - tlbr[0])
                msg.target_height = float(tlbr[3] - tlbr[1])
                msg.confidence = float(target_track.score)
                msg.distance = 0.0  # 距离测量已删除
            else:
                msg.target_id = -1
                msg.target_x = 0.0
                msg.target_y = 0.0
                msg.target_width = 0.0
                msg.target_height = 0.0
                msg.confidence = 0.0
                msg.distance = 0.0
            msg.track_ids = [track.track_id for track in tracks]
            msg.track_confidences = [float(track.score) for track in tracks]
            positions = []
            for track in tracks:
                tlbr = track.tlbr
                pos = Point()
                pos.x = float((tlbr[0] + tlbr[2]) / 2)
                pos.y = float((tlbr[1] + tlbr[3]) / 2)
                pos.z = 0.0
                positions.append(pos)
            msg.positions = positions
            if target_track is not None:
                msg.tracking_status = 'tracking'
            elif len(tracks) > 0:
                msg.tracking_status = 'idle'
            else:
                msg.tracking_status = 'searching'
            if hasattr(self, 'current_fps'):
                msg.fps = float(self.current_fps)
            else:
                msg.fps = 0.0
            msg.frame_count = self.frame_count
            self.tracking_result_pub.publish(msg)
        except Exception as e:
            self.get_logger().error(f"发布TrackingResult消息错误: {e}")

    def publish_detailed_tracking_data(self, tracks, target_track, mode):
        """
        参数: tracks, target_track, mode - 详细跟踪数据
        """
        try:
            def safe_int(value):
                if hasattr(value, 'item'):
                    return int(value.item())
                if hasattr(value, 'dtype') and 'int' in str(value.dtype):
                    return int(value)
                return int(value) if value is not None else 0
            
            def safe_float(value):
                if hasattr(value, 'item'):
                    return float(value.item())
                if hasattr(value, 'dtype') and 'float' in str(value.dtype):
                    return float(value)
                return float(value) if value is not None else 0.0
            
            def safe_list(value):
                if value is None:
                    return []
                if isinstance(value, np.ndarray):
                    return [safe_float(x) for x in value.tolist()]
                if isinstance(value, (list, tuple)):
                    return [safe_float(x) for x in value]
                return []

            detailed_data = {
                'timestamp': int(time.time() * 1000),
                'frame_id': safe_int(self.frame_count),
                'tracking_mode': str(mode),
                'target_detected': target_track is not None,
                'total_tracks': len(tracks),
                'tracks': [],
                'target_track': None,
                'statistics': {
                    'active_tracks': len([t for t in tracks if t.state == TrackState.TRACKED]),
                    'lost_tracks': len([t for t in tracks if t.state == TrackState.LOST]),
                    'new_tracks': len([t for t in tracks if t.state == TrackState.NEW]),
                }
            }

            for track in tracks:
                if track.state == TrackState.TRACKED:
                    tlbr = track.tlbr
                    center_x = safe_int(tlbr[0] + (tlbr[2] - tlbr[0]) / 2)
                    center_y = safe_int(tlbr[1] + (tlbr[3] - tlbr[1]) / 2)
                    track_data = {
                        'id': safe_int(track.track_id),
                        'status': 'tracking',
                        'position': {
                            'x': safe_float(center_x),
                            'y': safe_float(center_y),
                            'width': safe_float(tlbr[2] - tlbr[0]),
                            'height': safe_float(tlbr[3] - tlbr[1]),
                            'tlbr': [safe_float(tlbr[0]), safe_float(tlbr[1]), safe_float(tlbr[2]), safe_float(tlbr[3])]
                        },
                        'confidence': safe_float(track.score),
                        'age': safe_int(track.tracklet_len),
                        'time_since_update': safe_int(track.time_since_update),
                        'colors': {
                            'upper': safe_list(track.upper_color),
                            'lower': safe_list(track.lower_color)
                        },
                        'body_ratios': safe_list(track.body_ratios),
                        'distance': 0.0,  # 距离测量已删除
                        'is_target': track == target_track,
                        'tracking_quality': safe_float(self.calculate_tracking_quality(track))
                    }
                    detailed_data['tracks'].append(track_data)

            if target_track:
                tlbr = target_track.tlbr
                center_x = safe_int(tlbr[0] + (tlbr[2] - tlbr[0]) / 2)
                center_y = safe_int(tlbr[1] + (tlbr[3] - tlbr[1]) / 2)
                detailed_data['target_track'] = {
                    'id': safe_int(target_track.track_id),
                    'position': {
                        'x': safe_float(center_x),
                        'y': safe_float(center_y),
                        'width': safe_float(tlbr[2] - tlbr[0]),
                        'height': safe_float(tlbr[3] - tlbr[1]),
                        'tlbr': [safe_float(tlbr[0]), safe_float(tlbr[1]), safe_float(tlbr[2]), safe_float(tlbr[3])]
                    },
                    'confidence': safe_float(target_track.score),
                    'distance': 0.0,  # 距离测量已删除
                    'colors': {
                        'upper': safe_list(target_track.upper_color),
                        'lower': safe_list(target_track.lower_color)
                    },
                    'body_ratios': safe_list(target_track.body_ratios),
                    'tracking_quality': safe_float(self.calculate_tracking_quality(target_track)),
                    'velocity': {
                        'x': safe_float(target_track.mean[4]) if len(target_track.mean) > 4 else 0.0,
                        'y': safe_float(target_track.mean[5]) if len(target_track.mean) > 5 else 0.0
                    }
                }

            detailed_data['system_info'] = {
                'fps': safe_float(getattr(self, 'current_fps', 0.0)),
                'processing_time_ms': safe_float(getattr(self, 'processing_time', 0.0)),
                'memory_usage_mb': safe_float(self.get_memory_usage()),
                'features_received': len(self.current_detection_data)
            }

            json_msg = String()
            json_msg.data = json.dumps(detailed_data)
            self.detailed_tracking_pub.publish(json_msg)

        except Exception as e:
            self.get_logger().error(f'❌ 发布详细跟踪数据失败: {e}')
            import traceback
            self.get_logger().error(f'❌ 详细错误信息: {traceback.format_exc()}')

    def calculate_tracking_quality(self, track):
        """
        参数: track - 跟踪对象
        返回: float: 跟踪质量评分
        """
        try:
            quality_score = 0.0
            quality_score += min(40.0, track.score * 40)
            if track.tracklet_len > 10:
                stability = min(30.0, (track.tracklet_len / 50.0) * 30)
                quality_score += stability
            if track.time_since_update == 0:
                quality_score += 20.0
            elif track.time_since_update <= 3:
                quality_score += 15.0
            elif track.time_since_update <= 5:
                quality_score += 10.0
            if track.upper_color is not None and track.lower_color is not None:
                quality_score += 5.0
            if track.body_ratios is not None and len(track.body_ratios) > 0:
                quality_score += 5.0
            return min(100.0, quality_score)
        except Exception:
            return 50.0

    def get_memory_usage(self):
        """
        返回: float: 内存使用情况（MB）
        """
        try:
            import psutil
            process = psutil.Process()
            return process.memory_info().rss / 1024 / 1024
        except Exception:
            return 0.0

    def publish_status(self):
        """发布状态信息"""
        try:
            status_msg = String()
            status_info = {
                "mode": self.tracking_mode,
                "frame_count": self.frame_count,
                "tracked_objects": len(self.current_tracks),
                "target_id": self.current_target_track.track_id if self.current_target_track else None,
                "processing_time": self.processing_time
            }
            status_msg.data = str(status_info)
            self.status_pub.publish(status_msg)
        except Exception as e:
            self.get_logger().error(f"发布状态信息错误: {e}")

    def destroy_node(self):
        """节点销毁时的清理工作"""
        try:
            # 关闭OpenCV窗口
            cv2.destroyAllWindows()
            
            self.get_logger().info(f'📊 跟踪统计: 处理了 {self.frame_count} 帧')
            self.get_logger().info('✅ 简化版ByteTracker节点已关闭')
            super().destroy_node()
        except Exception as e:
            self.get_logger().error(f"节点销毁错误: {e}")

def main(args=None):
    """
    参数: args - 命令行参数
    """
    try:
        rclpy.init(args=args)
        node = ByteTrackerNode()
        try:
            rclpy.spin(node)
        except KeyboardInterrupt:
            pass
        finally:
            node.destroy_node()
            rclpy.shutdown()
    except Exception as e:
        print(f"主函数错误: {e}")
        traceback.print_exc()

if __name__ == '__main__':
    main()