
import logging
import cv2
import scipy.linalg
from collections import OrderedDict, deque
import time
import traceback
import openpyxl
from pathlib import Path
from rknn_colour_detect import detect_picture_with_confidence, Determine_the_position_of_the_entire_body
# from async_adapter import detect_picture_with_confidence
import numpy as np
from stero_vision import (
    StereoConfig, stereoCamera, getRectifyTransform,
    rectifyImage, stereoMatchSGBM_WLS, reprojectTo3D,
    measure_distance, cv2_put_chinese_text, preprocess,
    undistortion
)
from serial_utils import get_default_serial_port, detect_serial_ports,suggest_fix_permissions,check_serial_permissions
import platform
import serial
from threading import Thread, Lock

# 配置日志系统
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class TrackState:
    """跟踪状态枚举类"""
    NEW = 0
    TRACKED = 1
    LOST = 2
    REMOVED = 3


class BaseTrack:
    """基础跟踪类，定义了跟踪对象的基本属性和方法"""
    _count = 0

    def __init__(self):
        self.track_id = 0
        self.is_activated = False
        self.state = TrackState.NEW

        self.history = OrderedDict()
        self.features = []
        self.curr_feature = None
        self.score = 0
        self.start_frame = 0
        self.frame_id = 0
        self.time_since_update = 0

        # 颜色特征
        self.upper_color = None
        self.lower_color = None

        # 位置
        self.location = (np.inf, np.inf)

    @property
    def end_frame(self):
        return self.frame_id

    @staticmethod
    def next_id():
        BaseTrack._count += 1
        return BaseTrack._count

    def activate(self, *args):
        raise NotImplementedError("Need to be implemented in subclass")

    def predict(self):
        raise NotImplementedError("Need to be implemented in subclass")

    def update(self, *args, **kwargs):
        raise NotImplementedError("Need to be implemented in subclass")

    def mark_lost(self):
        self.state = TrackState.LOST
        logger.debug(f"Track {self.track_id} marked as lost")

    def mark_removed(self):
        self.state = TrackState.REMOVED
        logger.debug(f"Track {self.track_id} marked as removed")


class KalmanFilter:
    """卡尔曼滤波器实现，用于目标状态预测和更新"""

    def __init__(self):
        self.ndim = 4  # 状态维度：x, y, a, h (中心点x,y,宽高比,高度)
        self.dt = 1.0  # 时间步长

        # 创建卡尔曼滤波模型矩阵
        self._motion_mat = np.eye(2 * self.ndim, 2 * self.ndim)
        for i in range(self.ndim):
            self._motion_mat[i, self.ndim + i] = self.dt
        self._update_mat = np.eye(self.ndim, 2 * self.ndim)

        # 运动和观测不确定性权重
        self._std_weight_position = 1. / 20
        self._std_weight_velocity = 1. / 160

        # 卡方分布95%置信区间，用于门限计算
        self.chi2inv95 = {
            1: 3.8415,
            2: 5.9915,
            3: 7.8147,
            4: 9.4877,
            5: 11.070,
            6: 12.592,
            7: 14.067,
            8: 15.507,
            9: 16.919
        }

        logger.debug("Kalman filter initialized")

    def initiate(self, measurement):
        """从测量值初始化跟踪"""
        try:
            # 初始化状态均值
            mean_pos = measurement
            mean_vel = np.zeros_like(mean_pos)
            mean = np.r_[mean_pos, mean_vel]

            # 初始化协方差矩阵
            std = [
                2 * self._std_weight_position * measurement[3],
                2 * self._std_weight_position * measurement[3],
                1e-2,
                2 * self._std_weight_position * measurement[3],
                10 * self._std_weight_velocity * measurement[3],
                10 * self._std_weight_velocity * measurement[3],
                1e-5,
                10 * self._std_weight_velocity * measurement[3]
            ]
            covariance = np.diag(np.square(std))

            logger.debug(f"Track initiated: mean={mean[:4]}, cov_diag={np.diag(covariance)[:4]}")
            return mean, covariance
        except Exception as e:
            logger.error(f"Track initiation failed: {str(e)}")
            raise

    def predict(self, mean, covariance):
        """运行卡尔曼预测步骤"""
        try:
            # 计算过程噪声
            std_pos = [
                self._std_weight_position * mean[3],
                self._std_weight_position * mean[3],
                1e-2,
                self._std_weight_position * mean[3]
            ]
            std_vel = [
                self._std_weight_velocity * mean[3],
                self._std_weight_velocity * mean[3],
                1e-5,
                self._std_weight_velocity * mean[3]
            ]
            motion_cov = np.diag(np.square(np.r_[std_pos, std_vel]))

            # 预测新的状态
            mean = np.dot(mean, self._motion_mat.T)
            covariance = np.linalg.multi_dot((
                self._motion_mat, covariance, self._motion_mat.T)) + motion_cov

            logger.debug(f"State prediction: new_mean={mean[:4]}")
            return mean, covariance
        except Exception as e:
            logger.error(f"State prediction failed: {str(e)}")
            raise

    def multi_predict(self, means, covariances):
        """批量预测多个目标的状态"""
        try:
            # 验证输入
            if len(means) == 0:
                return means, covariances

            # 计算噪声协方差
            std_pos = [
                self._std_weight_position * means[:, 3],
                self._std_weight_position * means[:, 3],
                1e-2 * np.ones_like(means[:, 3]),
                self._std_weight_position * means[:, 3]
            ]
            std_vel = [
                self._std_weight_velocity * means[:, 3],
                self._std_weight_velocity * means[:, 3],
                1e-5 * np.ones_like(means[:, 3]),
                self._std_weight_velocity * means[:, 3]
            ]
            sqr = np.square(np.r_[std_pos, std_vel]).T

            # 构建协方差矩阵列表
            motion_covs = []
            for i in range(len(means)):
                motion_covs.append(np.diag(sqr[i]))
            motion_covs = np.asarray(motion_covs)

            # 预测新的状态
            new_means = np.dot(means, self._motion_mat.T)

            # 更新协方差
            new_covariances = []
            for i, (mean, cov) in enumerate(zip(means, covariances)):
                new_cov = np.linalg.multi_dot((
                    self._motion_mat, cov, self._motion_mat.T)) + motion_covs[i]
                new_covariances.append(new_cov)

            logger.debug(f"Batch prediction completed: predicted {len(means)} objects")
            return new_means, np.asarray(new_covariances)
        except Exception as e:
            logger.error(f"Batch prediction failed: {str(e)}")
            raise

    def update(self, mean, covariance, measurement):
        """卡尔曼更新步骤"""
        try:
            # 投影状态分布到测量空间
            projected_mean, projected_cov = self.project(mean, covariance)

            # 计算卡尔曼增益
            chol_factor, lower = scipy.linalg.cho_factor(
                projected_cov, lower=True, check_finite=False)
            kalman_gain = scipy.linalg.cho_solve(
                (chol_factor, lower), np.dot(covariance, self._update_mat.T).T,
                check_finite=False).T

            # 计算新状态
            innovation = measurement - projected_mean
            new_mean = mean + np.dot(innovation, kalman_gain.T)
            new_covariance = covariance - np.linalg.multi_dot((
                kalman_gain, projected_cov, kalman_gain.T))

            logger.debug(f"State update: measurement={measurement}, new_mean={new_mean[:4]}")
            return new_mean, new_covariance
        except Exception as e:
            logger.error(f"State update failed: {str(e)}")
            raise

    def project(self, mean, covariance):
        """将状态投影到测量空间"""
        try:
            # 计算观测噪声
            std = [
                self._std_weight_position * mean[3],
                self._std_weight_position * mean[3],
                1e-1,
                self._std_weight_position * mean[3]
            ]
            innovation_cov = np.diag(np.square(std))

            # 投影
            mean = np.dot(self._update_mat, mean)
            covariance = np.linalg.multi_dot((
                self._update_mat, covariance, self._update_mat.T))

            return mean, covariance + innovation_cov
        except Exception as e:
            logger.error(f"State projection failed: {str(e)}")
            raise

    def gating_distance(self, mean, covariance, measurements, only_position=False, metric='maha'):
        """计算状态分布与测量之间的门控距离"""
        mean, covariance = self.project(mean, covariance)
        if only_position:
            mean, covariance = mean[:2], covariance[:2, :2]
            measurements = measurements[:, :2]

        d = measurements - mean
        if metric == 'gaussian':
            return np.sum(d * d, axis=1)
        elif metric == 'maha':
            cholesky_factor = np.linalg.cholesky(covariance)
            z = scipy.linalg.solve_triangular(
                cholesky_factor, d.T, lower=True, check_finite=False,
                overwrite_b=True)
            squared_maha = np.sum(z * z, axis=0)
            return squared_maha
        else:
            raise ValueError('Invalid distance metric')


class STrack(BaseTrack):
    """单个目标跟踪器实现"""

    # 共享卡尔曼滤波器
    shared_kalman = KalmanFilter()

    def __init__(self, tlwh, score, temp_feat=None, upper_color=None, lower_color=None, body_ratios=None):
        """初始化跟踪对象

        Args:
            tlwh (array): [x,y,w,h] 格式的边界框
            score (float): 检测置信度
            temp_feat (array, optional): 临时特征向量
            upper_color (tuple, optional): 上衣颜色 (B,G,R)
            lower_color (tuple, optional): 下装颜色 (B,G,R)
            body_ratios (list, optional): 身体比例特征，16个浮点数
        """
        super(STrack, self).__init__()

        # 等待激活
        self._tlwh = np.asarray(tlwh, dtype=np.float64)
        self.kalman_filter = None
        self.mean, self.covariance = None, None

        # 添加预测状态存储
        self.predicted_mean, self.predicted_covariance = None, None
        self.is_activated = False

        self.score = score
        self.tracklet_len = 0

        # 添加颜色特征
        self.upper_color = upper_color
        self.lower_color = lower_color

        # 添加身体比例特征
        self.body_ratios = body_ratios if body_ratios is not None else [0.0] * 16

        logger.debug(
            f"Created new track: tlwh={tlwh}, score={score}, upper_color={upper_color}, lower_color={lower_color}, body_ratios={self.body_ratios[:3]}...")

    def predict(self):
        """预测下一个状态"""
        if self.state != TrackState.TRACKED:
            logger.debug(f"Track {self.track_id} not in tracked state, setting velocity to 0")
            mean_state = self.mean.copy()
            mean_state[7] = 0  # 将垂直速度置为0
            self.mean, self.covariance = self.kalman_filter.predict(mean_state, self.covariance)
        else:
            logger.debug(f"Track {self.track_id} normal prediction")
            self.mean, self.covariance = self.kalman_filter.predict(self.mean, self.covariance)

            # 保存预测状态
            self.predicted_mean = self.mean.copy()
            self.predicted_covariance = self.covariance.copy()

    @staticmethod
    def multi_predict(stracks):
        """批量预测多个目标的下一状态"""
        if len(stracks) > 0:
            multi_mean = np.asarray([st.mean.copy() for st in stracks])
            multi_covariance = np.asarray([st.covariance for st in stracks])

            # 将非跟踪状态的目标速度置为0
            for i, st in enumerate(stracks):
                if st.state != TrackState.TRACKED:
                    multi_mean[i][7] = 0

            # 批量预测
            multi_mean, multi_covariance = STrack.shared_kalman.multi_predict(multi_mean, multi_covariance)

            # 更新每个跟踪对象的状态
            for i, (mean, cov) in enumerate(zip(multi_mean, multi_covariance)):
                stracks[i].mean = mean
                stracks[i].covariance = cov
                # 保存预测状态
                stracks[i].predicted_mean = mean.copy()
                stracks[i].predicted_covariance = cov.copy()

            logger.debug(f"Batch predicted {len(stracks)} objects")

    def activate(self, kalman_filter, frame_id):
        """激活新的轨迹"""
        self.kalman_filter = kalman_filter
        self.track_id = self.next_id()

        # 初始化卡尔曼状态
        self.mean, self.covariance = self.kalman_filter.initiate(self.tlwh_to_xyah(self._tlwh))

        # 更新状态
        self.tracklet_len = 0
        self.state = TrackState.TRACKED

        # 第一帧直接激活，否则等待确认
        if frame_id == 1:
            self.is_activated = True

        self.frame_id = frame_id
        self.start_frame = frame_id

        logger.info(f"Activated new track: ID={self.track_id}, frame={frame_id}, position={self._tlwh}")

    def re_activate(self, new_track, frame_id, new_id=False):
        """重新激活丢失的轨迹"""
        # 更新卡尔曼状态
        self.mean, self.covariance = self.kalman_filter.update(
            self.mean, self.covariance, self.tlwh_to_xyah(new_track.tlwh)
        )

        # 更新轨迹状态
        self.tracklet_len = 0
        self.state = TrackState.TRACKED
        self.is_activated = True
        self.frame_id = frame_id

        # 更新颜色特征
        if new_track.upper_color is not None and all(c >= 0 for c in new_track.upper_color):
            self.upper_color = new_track.upper_color
        if new_track.lower_color is not None and all(c >= 0 for c in new_track.lower_color):
            self.lower_color = new_track.lower_color

        # 更新身体比例特征
        if hasattr(new_track, 'body_ratios') and new_track.body_ratios:
            if len(new_track.body_ratios) == len(self.body_ratios):
                self.body_ratios = new_track.body_ratios

        # 可选择分配新ID
        if new_id:
            self.track_id = self.next_id()

        self.score = new_track.score

        logger.info(f"Reactivated track: ID={self.track_id}, frame={frame_id}, new_position={new_track.tlwh}")

    def update(self, new_track, frame_id):
        """更新匹配的轨迹"""
        self.frame_id = frame_id
        self.tracklet_len += 1

        # 更新卡尔曼状态
        new_tlwh = new_track.tlwh
        self.mean, self.covariance = self.kalman_filter.update(
            self.mean, self.covariance, self.tlwh_to_xyah(new_tlwh)
        )

        # 更新状态
        self.state = TrackState.TRACKED
        self.is_activated = True
        self.score = new_track.score

        # 更新颜色特征(如果新的颜色有效)
        if new_track.upper_color is not None and all(c >= 0 for c in new_track.upper_color):
            # 如果已有颜色，采用加权平均
            if self.upper_color is not None:
                weight = min(0.3, 1.0 / self.tracklet_len)  # 随着轨迹长度增加权重减小
                self.upper_color = tuple(int((1 - weight) * c1 + weight * c2)
                                         for c1, c2 in zip(self.upper_color, new_track.upper_color))
            else:
                self.upper_color = new_track.upper_color

        if new_track.lower_color is not None and all(c >= 0 for c in new_track.lower_color):
            # 如果已有颜色，采用加权平均
            if self.lower_color is not None:
                weight = min(0.3, 1.0 / self.tracklet_len)
                self.lower_color = tuple(int((1 - weight) * c1 + weight * c2)
                                         for c1, c2 in zip(self.lower_color, new_track.lower_color))
            else:
                self.lower_color = new_track.lower_color

        # 更新身体比例特征（如果有新的数据）
        if hasattr(new_track, 'body_ratios') and new_track.body_ratios:
            if len(new_track.body_ratios) == len(self.body_ratios):
                # 也采用加权平均更新，但权重更低
                weight = min(0.2, 1.0 / self.tracklet_len)
                for i in range(len(self.body_ratios)):
                    if new_track.body_ratios[i] > 0:  # 只使用有效数据更新
                        self.body_ratios[i] = (1 - weight) * self.body_ratios[i] + weight * new_track.body_ratios[i]

        logger.debug(f"Updated track: ID={self.track_id}, frame={frame_id}, tracklet_len={self.tracklet_len}")

    @property
    def tlwh(self):
        """获取边界框 [x,y,w,h] 格式"""
        if self.mean is None:
            return self._tlwh.copy()

        ret = self.mean[:4].copy()
        ret[2] *= ret[3]  # 宽度 = 宽高比 * 高度
        ret[:2] -= ret[2:] / 2  # 中心点坐标转左上角坐标

        return ret

    @property
    def predicted_tlwh(self):
        """获取预测的边界框 [x,y,w,h] 格式"""
        if self.predicted_mean is None:
            return self.tlwh.copy()  # 如果没有预测状态，返回当前状态

        ret = self.predicted_mean[:4].copy()
        ret[2] *= ret[3]  # 宽度 = 宽高比 * 高度
        ret[:2] -= ret[2:] / 2  # 中心点坐标转左上角坐标

        return ret

    @property
    def predicted_tlbr(self):
        """获取预测的边界框 [x1,y1,x2,y2] 格式"""
        ret = self.predicted_tlwh.copy()
        ret[2:] += ret[:2]  # [x,y,w,h] -> [x1,y1,x2,y2]
        return ret

    @property
    def tlbr(self):
        """获取边界框 [x1,y1,x2,y2] 格式"""
        ret = self.tlwh.copy()
        ret[2:] += ret[:2]  # [x,y,w,h] -> [x1,y1,x2,y2]
        return ret

    @staticmethod
    def tlwh_to_xyah(tlwh):
        """[x,y,w,h] -> [cx,cy,aspect,h] 转换"""
        ret = np.asarray(tlwh).copy()
        ret[:2] += ret[2:] / 2  # 左上角坐标转中心点坐标
        ret[2] /= ret[3]  # 宽高比 = 宽度 / 高度
        return ret

    def to_xyah(self):
        return self.tlwh_to_xyah(self.tlwh)

    @staticmethod
    def tlbr_to_tlwh(tlbr):
        """[x1,y1,x2,y2] -> [x,y,w,h] 转换"""
        ret = np.asarray(tlbr).copy()
        ret[2:] -= ret[:2]  # [x1,y1,x2,y2] -> [x,y,w,h]
        return ret

    @staticmethod
    def tlwh_to_tlbr(tlwh):
        """[x,y,w,h] -> [x1,y1,x2,y2] 转换"""
        ret = np.asarray(tlwh).copy()
        ret[2:] += ret[:2]  # [x,y,w,h] -> [x1,y1,x2,y2]
        return ret

    def __repr__(self):
        return f'OT_{self.track_id}({self.start_frame}-{self.end_frame})'


def iou_distance(atracks, btracks):
    """计算IOU距离"""
    if (len(atracks) > 0 and isinstance(atracks[0], np.ndarray)) or \
            (len(btracks) > 0 and isinstance(btracks[0], np.ndarray)):
        atlbrs = atracks
        btlbrs = btracks
    else:
        atlbrs = [track.tlbr for track in atracks]
        btlbrs = [track.tlbr for track in btracks]

    # 计算IoU矩阵
    ious = optimized_bbox_ious(np.array(atlbrs), np.array(btlbrs))
    cost_matrix = 1 - ious  # 距离 = 1 - IoU

    logger.debug(f"Calculated IOU distance matrix: shape={cost_matrix.shape}")
    return cost_matrix


def color_distance(atracks, btracks, alpha=0.5):
    """计算颜色距离，结合上衣和下装颜色"""
    cost_matrix = np.zeros((len(atracks), len(btracks)), dtype=np.float64)

    for i, atrack in enumerate(atracks):
        for j, btrack in enumerate(btracks):
            # 上衣颜色距离
            upper_dist = 1.0  # 默认为最大距离
            if hasattr(atrack, 'upper_color') and hasattr(btrack, 'upper_color') and \
                    atrack.upper_color is not None and btrack.upper_color is not None:
                # 计算颜色欧氏距离并归一化
                upper_dist = np.sqrt(np.sum([(a - b) ** 2 for a, b in zip(atrack.upper_color, btrack.upper_color)]))
                # 颜色距离最大可达到 sqrt(255^2 * 3) = 441.7
                upper_dist = min(upper_dist / 442.0, 1.0)

            # 下装颜色距离
            lower_dist = 1.0  # 默认为最大距离
            if hasattr(atrack, 'lower_color') and hasattr(btrack, 'lower_color') and \
                    atrack.lower_color is not None and btrack.lower_color is not None:
                # 计算颜色欧氏距离并归一化
                lower_dist = np.sqrt(np.sum([(a - b) ** 2 for a, b in zip(atrack.lower_color, btrack.lower_color)]))
                lower_dist = min(lower_dist / 442.0, 1.0)

            # 组合上衣和下装距离
            if upper_dist < 1.0 and lower_dist < 1.0:
                # 两者都有颜色信息，加权组合
                cost_matrix[i, j] = alpha * upper_dist + (1 - alpha) * lower_dist
            elif upper_dist < 1.0:
                # 只有上衣颜色信息
                cost_matrix[i, j] = upper_dist
            elif lower_dist < 1.0:
                # 只有下装颜色信息
                cost_matrix[i, j] = lower_dist
            else:
                # 都没有颜色信息
                cost_matrix[i, j] = 1.0

    logger.debug(f"Calculated color distance matrix: shape={cost_matrix.shape}")
    return cost_matrix


def fuse_motion(kf, cost_matrix, tracks, detections, lambda_=0.98):
    """融合运动信息到代价矩阵"""
    if cost_matrix.size == 0:
        return cost_matrix

    gating_threshold = kf.chi2inv95[4]  # 使用4自由度
    measurements = np.asarray([det.tlwh_to_xyah(det.tlwh) for det in detections])

    for row, track in enumerate(tracks):
        gating_distance = kf.gating_distance(
            track.mean, track.covariance, measurements)

        # 应用门限
        cost_matrix[row, gating_distance > gating_threshold] = np.inf
        cost_matrix[row] = lambda_ * cost_matrix[row] + (1 - lambda_) * gating_distance

    return cost_matrix


def fuse_iou_with_color(iou_cost, color_cost, detections, w_iou=0.7, w_color=0.3):
    """融合IOU和颜色代价"""
    if iou_cost.size == 0:
        return iou_cost

    # 检查颜色代价矩阵是否有效
    if color_cost.size == 0 or color_cost.shape != iou_cost.shape:
        logger.warning("Invalid color cost matrix, using only IOU cost")
        w_iou, w_color = 1.0, 0.0

    # IOU相似度和颜色相似度
    iou_sim = 1 - iou_cost
    color_sim = 1 - color_cost

    # 获取检测置信度
    det_scores = np.array([det.score for det in detections])
    det_scores = np.expand_dims(det_scores, axis=0).repeat(iou_cost.shape[0], axis=0)

    # 融合相似度 (IOU + 颜色 + 置信度)
    if w_color > 0:
        # 使用IOU、颜色和置信度
        fuse_sim = w_iou * iou_sim + w_color * color_sim
        fuse_sim = fuse_sim * (1 + det_scores) / 2
    else:
        # 只使用IOU和置信度
        fuse_sim = iou_sim * (1 + det_scores) / 2

    # 转回代价形式
    fuse_cost = 1 - fuse_sim

    logger.debug(f"Fused cost matrix: w_iou={w_iou}, w_color={w_color}, shape={fuse_cost.shape}")
    return fuse_cost


def linear_assignment(cost_matrix, thresh=0.7):
    """线性分配算法(匈牙利算法)"""
    try:
        import lap
        if cost_matrix.size == 0:
            return [], list(range(cost_matrix.shape[0])), list(range(cost_matrix.shape[1]))

        matches, unmatched_a, unmatched_b = [], [], []
        cost, x, y = lap.lapjv(cost_matrix, extend_cost=True, cost_limit=thresh)

        for ix, mx in enumerate(x):
            if mx >= 0:
                matches.append([ix, mx])

        unmatched_a = np.where(x < 0)[0].tolist()
        unmatched_b = np.where(y < 0)[0].tolist()

        logger.debug(
            f"Linear assignment results: matches={len(matches)}, unmatched_a={len(unmatched_a)}, unmatched_b={len(unmatched_b)}")

        return matches, unmatched_a, unmatched_b
    except ImportError:
        logger.warning("lap library not found, fallback to scipy")
        # 降级使用scipy实现
        from scipy.optimize import linear_sum_assignment

        # 处理无限值
        cost_matrix = np.copy(cost_matrix)
        cost_matrix[cost_matrix > thresh] = thresh + 1e-4

        row_ind, col_ind = linear_sum_assignment(cost_matrix)
        matches = []
        unmatched_a = list(range(cost_matrix.shape[0]))
        unmatched_b = list(range(cost_matrix.shape[1]))

        for r, c in zip(row_ind, col_ind):
            if cost_matrix[r, c] <= thresh:
                matches.append([r, c])
                unmatched_a.remove(r)
                unmatched_b.remove(c)

        return matches, unmatched_a, unmatched_b


def optimized_bbox_ious(atlbrs, btlbrs):
    """向量化计算IOU，比循环快10-20倍"""
    # 转换为numpy数组
    atlbrs = np.asarray(atlbrs)
    btlbrs = np.asarray(btlbrs)

    # 获取数组形状
    alen, blen = len(atlbrs), len(btlbrs)
    if alen == 0 or blen == 0:
        return np.zeros((alen, blen), dtype=np.float64)

    # 计算交集区域
    xx1 = np.maximum(atlbrs[:, 0].reshape(alen, 1), btlbrs[:, 0].reshape(1, blen))
    yy1 = np.maximum(atlbrs[:, 1].reshape(alen, 1), btlbrs[:, 1].reshape(1, blen))
    xx2 = np.minimum(atlbrs[:, 2].reshape(alen, 1), btlbrs[:, 2].reshape(1, blen))
    yy2 = np.minimum(atlbrs[:, 3].reshape(alen, 1), btlbrs[:, 3].reshape(1, blen))

    w = np.maximum(0, xx2 - xx1)
    h = np.maximum(0, yy2 - yy1)
    inter = w * h

    # 计算各框面积
    area_a = ((atlbrs[:, 2] - atlbrs[:, 0]) * (atlbrs[:, 3] - atlbrs[:, 1])).reshape(alen, 1)
    area_b = ((btlbrs[:, 2] - btlbrs[:, 0]) * (btlbrs[:, 3] - btlbrs[:, 1])).reshape(1, blen)
    union = area_a + area_b - inter

    # 计算IoU
    ious = np.where(union > 0, inter / union, 0)
    return ious


def bbox_ious(atlbrs, btlbrs):
    """计算两组边界框之间的IoU"""
    ious = np.zeros((len(atlbrs), len(btlbrs)), dtype=np.float64)
    if ious.size == 0:
        return ious

    # 计算IoU
    for i, atlbr in enumerate(atlbrs):
        for j, btlbr in enumerate(btlbrs):
            # 计算交集区域
            xx1 = max(atlbr[0], btlbr[0])
            yy1 = max(atlbr[1], btlbr[1])
            xx2 = min(atlbr[2], btlbr[2])
            yy2 = min(atlbr[3], btlbr[3])

            w = max(0, xx2 - xx1)
            h = max(0, yy2 - yy1)
            inter = w * h

            # 计算并集区域
            area_a = (atlbr[2] - atlbr[0]) * (atlbr[3] - atlbr[1])
            area_b = (btlbr[2] - btlbr[0]) * (btlbr[3] - btlbr[1])
            union = area_a + area_b - inter

            if union > 0:
                ious[i, j] = inter / union

    return ious


def joint_stracks(tlista, tlistb):
    """合并两个轨迹列表，确保没有重复ID"""
    exists = {}
    res = []
    for t in tlista:
        exists[t.track_id] = 1
        res.append(t)

    for t in tlistb:
        tid = t.track_id
        if not exists.get(tid, 0):
            exists[tid] = 1
            res.append(t)

    return res


def sub_stracks(tlista, tlistb):
    """从列表A中移除在列表B中出现的轨迹"""
    stracks = {}
    for t in tlista:
        stracks[t.track_id] = t

    for t in tlistb:
        tid = t.track_id
        if stracks.get(tid, 0):
            del stracks[tid]

    return list(stracks.values())


def remove_duplicate_stracks(stracksa, stracksb):
    """移除重复轨迹，保留跟踪时间较长的"""
    pdist = iou_distance(stracksa, stracksb)
    pairs = np.where(pdist < 0.15)
    dupa, dupb = list(), list()

    for p, q in zip(*pairs):
        timep = stracksa[p].frame_id - stracksa[p].start_frame
        timeq = stracksb[q].frame_id - stracksb[q].start_frame

        if timep > timeq:
            dupb.append(q)
        else:
            dupa.append(p)

    resa = [t for i, t in enumerate(stracksa) if i not in dupa]
    resb = [t for i, t in enumerate(stracksb) if i not in dupb]

    return resa, resb


class ColorByteTracker:
    """融合颜色特征的ByteTracker算法实现"""

    def __init__(self, args, frame_rate=30):
        """初始化跟踪器

        Args:
            args: 包含配置项的对象
            frame_rate: 视频帧率
        """
        # 轨迹列表
        self.tracked_stracks = []  # 当前正在跟踪的轨迹
        self.lost_stracks = []  # 暂时丢失的轨迹
        self.removed_stracks = []  # 已移除的轨迹

        # 参数设置
        self.frame_id = 0
        self.args = args
        self.det_thresh = args.get('track_thresh', 0.5) + 0.1  # 稍高的阈值用于激活新轨迹

        # 缓冲区和最大丢失时间设置
        self.buffer_size = int(frame_rate / 30.0 * args.get('track_buffer', 30))
        self.max_time_lost = self.buffer_size

        # 初始化卡尔曼滤波器
        self.kalman_filter = KalmanFilter()

        # 颜色权重
        self.color_weight = args.get('color_weight', 0.3)

        logger.info(
            f"ColorByteTracker initialized: det_thresh={self.det_thresh}, buffer_size={self.buffer_size}, color_weight={self.color_weight}")

    def update(self, detection_results, img=None):
        """更新跟踪状态

        Args:
            detection_results: 检测结果列表，每项为(上衣坐标, 下装坐标, 上衣颜色, 下装颜色, 置信度, 身体比例)
            img: 当前帧图像，用于颜色提取

        Returns:
            list: 当前活跃的跟踪轨迹
        """
        self.frame_id += 1
        logger.debug(f"Processing frame {self.frame_id}, detected {len(detection_results)} objects")

        # 结果容器
        activated_starcks = []
        refind_stracks = []
        lost_stracks = []
        removed_stracks = []

        # 如果没有检测结果，就将所有轨迹标记为丢失
        if not detection_results:
            for track in self.tracked_stracks:
                if track.state == TrackState.TRACKED:
                    track.mark_lost()
                    lost_stracks.append(track)

            logger.debug(f"No detections, {len(lost_stracks)} tracks marked as lost")

            # 更新状态并返回
            self._update_status(activated_starcks, refind_stracks, lost_stracks, removed_stracks)
            return [track for track in self.tracked_stracks if track.is_activated]

        # 处理检测结果，创建 STrack 对象
        detections = []
        low_detections = []

        for result in detection_results:
            # 解析检测结果
            try:
                # 处理不同长度的结果
                if len(result) >= 7:  # 结果包含身体比例
                    upper_bbox, lower_bbox, upper_color, lower_color, upper_confidence, lower_confidence,body_ratios = result
                else:  # 结果不包含身体比例
                    upper_bbox, lower_bbox, upper_color, lower_color,  upper_confidence, lower_confidence = result
                    body_ratios = None

                confidence = upper_confidence + lower_confidence

                # 从上衣和下装位置估计整体位置
                if len(upper_bbox) == 4 and len(lower_bbox) == 4:
                    # 使用上衣和下装共同确定位置
                    full_bbox = self._integrate_position(upper_bbox, lower_bbox)
                elif len(upper_bbox) == 4:
                    # 只有上衣
                    full_bbox = self._expand_position(upper_bbox, is_upper=True)
                elif len(lower_bbox) == 4:
                    # 只有下装
                    full_bbox = self._expand_position(lower_bbox, is_upper=False)
                else:
                    logger.warning(f"Invalid detection result: {result}")
                    continue

                # 构建跟踪对象
                strack = STrack(
                    STrack.tlbr_to_tlwh(np.array(full_bbox)),
                    confidence,
                    upper_color=upper_color,
                    lower_color=lower_color,
                    body_ratios=body_ratios
                )

                # 根据置信度分类
                if confidence >= self.args.get('track_thresh', 0.5):
                    detections.append(strack)
                elif confidence >= 0.1:  # 低置信度阈值
                    low_detections.append(strack)

            except Exception as e:
                logger.error(f"Error processing detection result: {str(e)}")
                continue

        logger.debug(f"Created track objects: high_conf={len(detections)}, low_conf={len(low_detections)}")

        # 将现有轨迹分为已确认和未确认
        unconfirmed = []
        tracked_stracks = []

        for track in self.tracked_stracks:
            if not track.is_activated:
                unconfirmed.append(track)
            else:
                tracked_stracks.append(track)

        # 合并确认轨迹和丢失轨迹作为匹配池
        strack_pool = joint_stracks(tracked_stracks, self.lost_stracks)

        # 预测所有轨迹的新位置
        STrack.multi_predict(strack_pool)

        # 第一阶段关联 - 与高置信度检测关联
        # 计算IOU距离
        iou_dists = iou_distance(strack_pool, detections)

        # 计算颜色距离并融合
        if self.color_weight > 0:
            color_dists = color_distance(strack_pool, detections)
            dists = fuse_iou_with_color(
                iou_dists, color_dists, detections,
                w_iou=1 - self.color_weight, w_color=self.color_weight
            )
        else:
            dists = iou_dists

        # 使用匈牙利算法分配
        matches, u_track, u_detection = linear_assignment(
            dists, thresh=self.args.get('match_thresh', 0.8)
        )

        # 处理匹配结果
        for i_track, i_det in matches:
            track = strack_pool[i_track]
            det = detections[i_det]

            if track.state == TrackState.TRACKED:
                # 已跟踪轨迹，更新
                track.update(det, self.frame_id)
                activated_starcks.append(track)
            else:
                # 丢失轨迹，重新激活
                track.re_activate(det, self.frame_id, new_id=False)
                refind_stracks.append(track)

        logger.debug(
            f"First association: matches={len(matches)}, unmatched_tracks={len(u_track)}, unmatched_detections={len(u_detection)}")

        # 第二阶段关联 - 与低置信度检测关联
        if len(low_detections) > 0:
            # 仅使用仍在跟踪中的轨迹
            r_tracked_stracks = [strack_pool[i] for i in u_track if strack_pool[i].state == TrackState.TRACKED]

            # 计算IOU距离
            iou_dists = iou_distance(r_tracked_stracks, low_detections)

            # 颜色距离和融合
            if self.color_weight > 0:
                color_dists = color_distance(r_tracked_stracks, low_detections)
                dists = fuse_iou_with_color(
                    iou_dists, color_dists, low_detections,
                    w_iou=1 - self.color_weight, w_color=self.color_weight
                )
            else:
                dists = iou_dists

            # 使用更宽松的阈值
            matches, u_tracks, u_low_detection = linear_assignment(dists, thresh=0.5)

            # 处理匹配结果
            for i_track, i_det in matches:
                track = r_tracked_stracks[i_track]
                det = low_detections[i_det]

                if track.state == TrackState.TRACKED:
                    track.update(det, self.frame_id)
                    activated_starcks.append(track)
                else:
                    track.re_activate(det, self.frame_id, new_id=False)
                    refind_stracks.append(track)

            # 处理未匹配轨迹
            for i in u_tracks:
                track = r_tracked_stracks[i]
                if track.state != TrackState.LOST:
                    track.mark_lost()
                    lost_stracks.append(track)

            logger.debug(f"Second association: matches={len(matches)}, unmatched_tracks={len(u_tracks)}")
        else:
            # 没有低置信度检测，将所有未匹配轨迹标记为丢失
            for i in u_track:
                track = strack_pool[i]
                if track.state != TrackState.LOST:
                    track.mark_lost()
                    lost_stracks.append(track)

        # 处理未确认轨迹
        u_detection_track = [detections[i] for i in u_detection]
        iou_dists = iou_distance(unconfirmed, u_detection_track)

        if self.color_weight > 0:
            color_dists = color_distance(unconfirmed, u_detection_track)
            dists = fuse_iou_with_color(
                iou_dists, color_dists, u_detection_track,
                w_iou=1 - self.color_weight, w_color=self.color_weight
            )
        else:
            dists = iou_dists

        matches, u_unconfirmed, u_detection = linear_assignment(dists, thresh=0.7)

        for i_track, i_det in matches:
            unconfirmed[i_track].update(u_detection_track[i_det], self.frame_id)
            activated_starcks.append(unconfirmed[i_track])

        # 移除未匹配的未确认轨迹
        for i in u_unconfirmed:
            track = unconfirmed[i]
            track.mark_removed()
            removed_stracks.append(track)

        logger.debug(f"Process unconfirmed tracks: matches={len(matches)}, removed={len(u_unconfirmed)}")

        # 创建新轨迹
        for i in u_detection:
            track = u_detection_track[i]
            if track.score < self.det_thresh:
                continue

            # 激活新轨迹
            track.activate(self.kalman_filter, self.frame_id)
            activated_starcks.append(track)

        logger.debug(
            f"New tracks: {len([t for t in activated_starcks if t.track_id > BaseTrack._count - len(u_detection)])}")

        # 处理丢失时间过长的轨迹
        for track in self.lost_stracks:
            if self.frame_id - track.end_frame > self.max_time_lost:
                track.mark_removed()
                removed_stracks.append(track)

        logger.debug(f"Remove lost tracks: {len(removed_stracks)}")

        # 更新跟踪器状态
        self._update_status(activated_starcks, refind_stracks, lost_stracks, removed_stracks)

        # 返回激活的轨迹
        return [track for track in self.tracked_stracks if track.is_activated]

    def _update_status(self, activated_starcks, refind_stracks, lost_stracks, removed_stracks):
        """更新跟踪器内部状态"""
        # 更新跟踪中轨迹
        self.tracked_stracks = [t for t in self.tracked_stracks if t.state == TrackState.TRACKED]
        self.tracked_stracks = joint_stracks(self.tracked_stracks, activated_starcks)
        self.tracked_stracks = joint_stracks(self.tracked_stracks, refind_stracks)

        # 更新丢失轨迹
        self.lost_stracks = sub_stracks(self.lost_stracks, self.tracked_stracks)
        self.lost_stracks.extend(lost_stracks)
        self.lost_stracks = sub_stracks(self.lost_stracks, self.removed_stracks)

        # 更新移除轨迹
        self.removed_stracks.extend(removed_stracks)

        # 移除重复轨迹
        self.tracked_stracks, self.lost_stracks = remove_duplicate_stracks(
            self.tracked_stracks, self.lost_stracks)

        logger.debug(
            f"Tracker status: tracked={len(self.tracked_stracks)}, lost={len(self.lost_stracks)}, removed={len(self.removed_stracks)}")

    def _integrate_position(self, upper_pos, lower_pos):
        """整合上衣和下装位置得到整体位置"""
        x_min = min(upper_pos[0], lower_pos[0])
        y_min = min(upper_pos[1], lower_pos[1])
        x_max = max(upper_pos[2], lower_pos[2])
        y_max = max(upper_pos[3], lower_pos[3])

        return [x_min, y_min, x_max, y_max]

    def _expand_position(self, part_pos, is_upper=True):
        """根据部分位置(上衣或下装)估计整体位置"""
        x_min, y_min, x_max, y_max = part_pos
        width = x_max - x_min
        height = y_max - y_min

        if is_upper:
            # 从上衣估计整体位置
            return [
                max(0, x_min - int(width * 0.1)),
                max(0, y_min - int(height * 0.1)),
                x_max + int(width * 0.1),
                y_max + int(height * 1.5)  # 向下扩展更多
            ]
        else:
            # 从下装估计整体位置
            return [
                max(0, x_min - int(width * 0.1)),
                max(0, y_min - int(height * 1.2)),  # 向上扩展更多
                x_max + int(width * 0.1),
                y_max + int(height * 0.1)
            ]


def validate_color(color):
    """验证并修复颜色值，确保其为有效的BGR格式"""
    if color is None:
        return (0, 0, 255)  # 默认红色

    try:
        # 尝试将color转换为列表
        color_list = []

        # 处理各种可能的输入类型
        if hasattr(color, '__iter__'):  # 可迭代对象
            for c in color:
                try:
                    # 尝试转换为数字
                    val = float(c)
                    # 限制范围到0-255
                    val = max(0, min(255, val))
                    color_list.append(int(val))
                except:
                    color_list.append(0)
        else:
            # 如果不是可迭代对象，返回默认颜色
            return (0, 0, 255)

        # 确保有3个BGR值
        while len(color_list) < 3:
            color_list.append(0)

        # 只保留前3个值
        return tuple(color_list[:3])
    except Exception as e:
        print(f"颜色验证错误: {e}, 颜色值: {color}, 类型: {type(color)}")
        return (0, 0, 255)  # 出错返回红色


def track_with_color_features(video_path=None, output_path=None):
    """使用融合颜色的ByteTracker跟踪行人"""
    import cv2 as cv
    import time

    # 尝试导入姿态估计模块，如果可用的话
    try:
        from obtain_features import get_keypoints, calculate_body_ratios
        pose_estimation_available = True
        logger.info("姿态估计模块已加载，将用于获取身体比例特征")
    except ImportError:
        pose_estimation_available = False
        logger.warning("姿态估计模块未加载，将不使用身体比例特征")

    # 跟踪参数
    args = {
        'track_thresh': 0.5,  # 跟踪阈值
        'track_buffer': 100,  # 轨迹缓冲
        'match_thresh': 0.8,  # 匹配阈值
        'color_weight': 0.5  # 颜色权重
    }

    # 初始化跟踪器
    tracker = ColorByteTracker(args)
    detection_results = None

    # 打开视频
    if video_path:
        cap = cv.VideoCapture(video_path)
    else:
        cap = cv.VideoCapture(0)  # 使用摄像头

    # 获取视频原始帧率
    original_fps = cap.get(cv.CAP_PROP_FPS)
    logger.info(f"原始视频帧率: {original_fps} FPS")

    # 计算每帧应该的时间间隔（毫秒）
    frame_interval_ms = int(1000 / original_fps)
    logger.info(f"每帧间隔: {frame_interval_ms} 毫秒")

    # 准备输出视频
    writer = None
    if output_path:
        width = int(cap.get(cv.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv.CAP_PROP_FRAME_HEIGHT))
        writer = cv.VideoWriter(
            output_path,
            cv.VideoWriter_fourcc(*'mp4v'),
            original_fps,  # 使用原始帧率
            (width, height)
        )

    # 轨迹颜色映射
    track_colors = {}
    color_palette = [
        (255, 0, 0), (0, 255, 0), (0, 0, 255), (255, 255, 0),
        (0, 255, 255), (255, 0, 255), (128, 128, 0), (0, 128, 128),
        (128, 0, 128), (255, 128, 0), (128, 255, 0), (0, 255, 128)
    ]

    frame_count = 0
    prev_frame_time = 0  # 上一帧处理的时间点

    try:
        while True:
            # 帧率控制 - 计算从上一帧到现在经过的时间
            curr_time = time.time()
            if prev_frame_time > 0:  # 不是第一帧
                elapsed_time = (curr_time - prev_frame_time) * 1000  # 毫秒
                sleep_time = max(1, frame_interval_ms - int(elapsed_time))
                if sleep_time > 0:
                    # 等待足够的时间来匹配原始帧率
                    if cv.waitKey(sleep_time) == 27:  # ESC退出
                        break

            # 读取下一帧
            ret, frame = cap.read()
            if not ret:
                break

            frame = cv.resize(frame, (640, 640))

            # 记录这一帧开始处理的时间
            prev_frame_time = time.time()

            frame_count += 1
            logger.info(f"Processing frame {frame_count}")

            # 准备跟踪数据
            tracking_inputs = []

            # 检测目标
            start_time = time.time()
            detection_results = detect_picture_with_confidence(frame)
            detect_time = time.time() - start_time

            # 从检测结果中获取身体比例
            start_time = time.time()
            tracking_inputs = extract_body_ratios_from_detections(frame, detection_results)
            ratio_time = time.time() - start_time
            logger.debug(f"身体比例提取耗时: {ratio_time:.3f}s")

            # 更新跟踪器
            start_time = time.time()
            tracking_results = tracker.update(tracking_inputs, frame)
            track_time = time.time() - start_time

            # 可视化跟踪结果
            for track in tracking_results:
                if track.track_id not in track_colors:
                    track_colors[track.track_id] = color_palette[track.track_id % len(color_palette)]

                color = track_colors[track.track_id]
                tlbr = track.tlbr

                cv.rectangle(
                    frame,
                    (int(tlbr[0]), int(tlbr[1])),
                    (int(tlbr[2]), int(tlbr[3])),
                    color, 2
                )

                # 显示ID和颜色信息
                upper_color_text = f"U:{track.upper_color}" if track.upper_color else "No Upper"
                lower_color_text = f"L:{track.lower_color}" if track.lower_color else "No Lower"
                cv.putText(
                    frame,
                    f"ID:{track.track_id}",
                    (int(tlbr[0]), int(tlbr[1] - 10)),
                    cv.FONT_HERSHEY_SIMPLEX, 0.5, color, 2
                )

                # 在图像上标记上衣和下装颜色样本
                if track.upper_color:
                    upper_color = validate_color(track.upper_color)
                    cv.rectangle(
                        frame,
                        (int(tlbr[0]), int(tlbr[1])),
                        (int(tlbr[0] + 15), int(tlbr[1] + 15)),
                        upper_color, -1
                    )

                if track.lower_color:
                    lower_color = validate_color(track.lower_color)
                    cv.rectangle(
                        frame,
                        (int(tlbr[0] + 20), int(tlbr[1])),
                        (int(tlbr[0] + 35), int(tlbr[1] + 15)),
                        lower_color, -1
                    )

            # 显示性能信息
            cv.putText(
                frame,
                f"Detection time: {detect_time:.3f}s, Tracking time: {track_time:.3f}s",
                (10, 30),
                cv.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2
            )

            # 显示跟踪统计和当前帧率
            actual_fps = 1.0 / (time.time() - prev_frame_time) if prev_frame_time > 0 else 0
            cv.putText(
                frame,
                f"Tracking objects: {len(tracking_results)}, FPS: {actual_fps:.1f}/{original_fps:.1f}",
                (10, 60),
                cv.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2
            )

            # 显示结果
            cv.imshow("Color-enhanced ByteTracker", frame)

            # 保存视频
            if writer:
                writer.write(frame)

    except Exception as e:
        logger.error(f"Error during tracking: {str(e)}")
        traceback.print_exc()

    finally:
        # 释放资源
        cap.release()
        if writer:
            writer.release()
        cv.destroyAllWindows()


# 添加单目标跟踪功能
class SingleTargetTracker:
    """单目标跟踪器，可以根据目标特征锁定并跟踪特定人物"""

    # 跟踪模式
    MODE_SELECTING = 0  # 目标选择模式
    MODE_TRACKING = 1  # 目标跟踪模式

    def __init__(self, tracker_args, target_features, max_lost_time=30, control_car=False, car_serial_port=None):
        """初始化单目标跟踪器

        参数:
            tracker_args: 传递给ByteTracker的参数
            target_features: 目标特征，格式为(body_ratios, shirt_color, pants_color)
            max_lost_time: 最大丢失时间，超过此时间将切换回选择模式
            control_car: 是否启用小车控制
            car_serial_port: 小车串口端口，None表示自动检测
        """
        # 初始化ByteTracker
        self.base_tracker = ColorByteTracker(tracker_args)

        # 保存目标特征
        self.target_body_ratios, self.target_shirt_color, self.target_pants_color = target_features

        # 初始化状态
        self.mode = self.MODE_SELECTING  # 初始为选择模式
        self.target_id = None  # 目标ID，None表示尚未锁定
        self.lost_frames = 0  # 目标丢失帧数
        self.max_lost_frames = max_lost_time  # 最大丢失帧数

        self.frame_count = 0
        self.target_score_history = []  # 记录目标得分历史
        self.confirmed_target_position = None  # 确认的目标位置
        self.target_trajectory = deque(maxlen=30)  # 记录目标轨迹

        # 小车控制相关
        self.control_car = control_car
        self.ser_car = None
        self.my_instruction1 = ''  # 前进/后退指令
        self.my_instruction2 = ''  # 左转/右转指令
        self.stereo_distance = None  # 双目立体视觉计算的距离
        self.stereo_distance_instruction = ''  # 基于距离的指令

        if control_car:
            # 使用辅助函数初始化串口连接
            self.init_serial_connection(car_serial_port)

        logger.info("单目标跟踪器已初始化，当前为目标选择模式")
        if control_car and self.ser_car:
            logger.info("小车控制功能已启用")

    def init_serial_connection(self, car_serial_port):
        """
        初始化串口连接

        Args:
            car_serial_port: 串口设备路径
        """
        if not self.control_car:
            return

        # 检测可用串口
        available_ports = detect_serial_ports()
        if not available_ports:
            logger.warning("系统未检测到可用串口设备")
            if platform.system() == 'Linux':
                suggest_fix_permissions()
            self.control_car = False
            return

        # 提示可用端口
        logger.info(f"系统可用串口: {available_ports}")

        # 检查指定端口是否存在
        if car_serial_port not in available_ports:
            logger.warning(f"指定的串口 {car_serial_port} 不在可用列表中")
            if available_ports:
                logger.info(f"将尝试使用第一个可用串口: {available_ports[0]}")
                car_serial_port = available_ports[0]

        # 检查权限
        if platform.system() == 'Linux' and not check_serial_permissions(car_serial_port):
            logger.error(f"没有串口 {car_serial_port} 的访问权限")
            suggest_fix_permissions()
            self.control_car = False
            return

        # 尝试连接串口
        try:
            import serial
            self.ser_car = serial.Serial(car_serial_port, 115200, timeout=0.5)
            logger.info(f"成功连接小车串口: {car_serial_port}")
        except Exception as e:
            logger.error(f"警告：无法连接小车串口({car_serial_port}): {e}")
            logger.error("小车控制功能将不可用")
            self.control_car = False

    def body_ratio_similarity(self, detected_ratios):
        """计算身体比例相似度

        参数:
            detected_ratios: 检测到的身体比例

        返回值:
            float: 相似度，0-1之间，越高越相似
        """
        if not detected_ratios or len(detected_ratios) != len(self.target_body_ratios):
            return 0.0

        # 使用余弦相似度
        try:
            target = np.array(self.target_body_ratios)
            detected = np.array(detected_ratios)

            # 处理零向量情况
            if np.all(target == 0) or np.all(detected == 0):
                return 0.0

            # 计算余弦相似度
            dot_product = np.dot(target, detected)
            norm_product = np.linalg.norm(target) * np.linalg.norm(detected)

            if norm_product == 0:
                return 0.0

            similarity = dot_product / norm_product
            # 将相似度范围从[-1,1]调整到[0,1]
            similarity = (similarity + 1) / 2
            return similarity

        except Exception as e:
            logger.error(f"计算身体比例相似度出错: {str(e)}")
            return 0.0

    def calculate_stereo_distance_instruction(self, distance):
        """
        基于双目立体视觉测量的距离计算控制指令

        参数:
            distance: 双目测量的距离（米）
        """
        if distance is None or distance <= 0:
            self.stereo_distance_instruction = '无效数据'
            return

        # 定义基于实际距离的控制阈值（单位：米）
        if 1.3 <= distance <= 1.7:
            self.my_instruction1 = '停止'
            instruction = '距离适中'
        elif distance < 1.3:
            self.my_instruction1 = '慢速后退'
            instruction = '距离较近'
        elif 1.7 < distance < 2.5:
            self.my_instruction1 = '慢速前进'
            instruction = '距离稍远'
        elif 2.5 <= distance < 3:
            self.my_instruction1 = '中速前进'
            instruction = '距离较远'
        elif distance >= 3:
            self.my_instruction1 = '快速前进'
            instruction = '距离较远'
        else:
            self.my_instruction1 = ''
            instruction = '无效数据'

        self.stereo_distance_instruction = instruction
        logger.debug(f"双目距离:{distance:.2f}米, 指令:{instruction}")

    def check_position_right_left(self, x_second, y_second):
        """判断人体在画面中的左右位置，决定是否需要转向"""
        target_x = 320  # 图像中心x坐标
        target_y = 240  # 图像中心y坐标
        x_threshold = 120  # 左右判断阈值
        y_threshold = 120  # 上下判断阈值

        if x_second == 0 and y_second == 0:
            comparison_result = 'stop'
            self.my_instruction2 = ''
        else:
            if x_second > target_x + x_threshold and y_second < target_y - y_threshold:
                comparison_result = '右上'
                self.my_instruction2 = '右转'
            elif x_second > target_x + x_threshold and y_second > target_y + y_threshold:
                comparison_result = '右下'
                self.my_instruction2 = '右转'
            elif x_second > target_x + x_threshold and target_y - y_threshold <= y_second <= target_y + y_threshold:
                comparison_result = '右中'
                self.my_instruction2 = '右转'
            elif x_second < target_x - x_threshold and y_second < target_y - y_threshold:
                comparison_result = '左上'
                self.my_instruction2 = '左转'
            elif x_second < target_x - x_threshold and y_second > target_y + y_threshold:
                comparison_result = '左下'
                self.my_instruction2 = '左转'
            elif x_second < target_x - x_threshold and target_y - y_threshold <= y_second <= target_y + y_threshold:
                comparison_result = '左中'
                self.my_instruction2 = '左转'
            elif target_x - x_threshold <= x_second <= target_x + x_threshold and y_second < target_y - y_threshold:
                comparison_result = '中上'
                self.my_instruction2 = ''
            elif target_x - x_threshold <= x_second <= target_x + x_threshold and y_second > target_y + y_threshold:
                comparison_result = '中下'
                self.my_instruction2 = ''
            else:
                comparison_result = 'stop'
                self.my_instruction2 = ''

        return comparison_result

    def send_car_command(self):
        """向小车发送控制命令"""
        if not self.ser_car:
            logger.warning("警告：小车串口未连接，无法发送命令")
            return

        send_data = ''
        # 根据指令生成命令
        if self.my_instruction1 == '慢速前进' and self.my_instruction2 == '':
            logger.info('只慢速前进')
            send_data = '7B 00 00 01 90 00 00 00 00 EA 7D'
        elif self.my_instruction1 == '中速前进' and self.my_instruction2 == '':
            logger.info('只中速前进')
            send_data = '7B 00 00 02 8A 00 00 00 00 F3 7D'
        elif self.my_instruction1 == '快速前进' and self.my_instruction2 == '':
            logger.info('只快速前进')
            send_data = '7B 00 00 03 20 00 00 00 00 58 7D'
        elif self.my_instruction1 == '慢速后退' and self.my_instruction2 == '':
            logger.info('只慢速后退')
            send_data = '7B 00 00 FE 70 00 00 00 00 F5 7D'
        elif self.my_instruction1 == '中速前进' and self.my_instruction2 == '左转':
            logger.info('左前方')
            send_data = '7B 00 00 02 8A 00 00 01 5E AC 7D'
        elif self.my_instruction1 == '慢速前进' and self.my_instruction2 == '左转':
            logger.info('左前方')
            send_data = '7B 00 00 01 90 00 00 00 AF 45 7D'
        elif self.my_instruction1 == '慢速后退' and self.my_instruction2 == '左转':
            logger.info('左后方')
            send_data = '7B 00 00 FE 70 00 00 01 5E AA 7D'
        elif self.my_instruction1 == '慢速前进' and self.my_instruction2 == '右转':
            logger.info('右前方')
            send_data = '7B 00 00 01 90 00 00 FF 51 44 7D'
        elif self.my_instruction1 == '中速前进' and self.my_instruction2 == '右转':
            logger.info('右前方')
            send_data = '7B 00 00 02 8A 00 00 FE A2 AF 7D'
        elif self.my_instruction1 == '慢速后退' and self.my_instruction2 == '  右转':
            logger.info('右后方')
            send_data = '7B 00 00 FE 70 00 00 FE A2 A9 7D'
        elif self.my_instruction1 == '快速前进' and self.my_instruction2 == '右转':
            logger.info('只快速前进')
            send_data = '7B 00 00 03 20 00 00 FF 51 F6 7D'
        elif self.my_instruction1 == '快速前进' and self.my_instruction2 == '左转':
            logger.info('只快速前进')
            send_data = '7B 00 00 03 20 00 00 00 AF F7 7D'
        elif self.my_instruction1 == '停止':
            logger.info('立即停止')
            send_data = '7B 00 01 00 00 00 00 00 00 7A 7D'
        else:
            logger.debug('无操作指令')
            send_data = '7B 00 01 00 00 00 00 00 00 7A 7D'


        if send_data:
            try:
                send_data_hex = bytes.fromhex(send_data)
                self.ser_car.write(send_data_hex)
                logger.info(f"发送成功: {send_data}")
            except Exception as e:
                logger.error(f"发送失败: {e}")

        self.my_instruction1 = ''
        self.my_instruction2 = ''

    def color_similarity(self, color1, color2):
        """计算颜色相似度

        参数:
            color1, color2: BGR颜色元组

        返回值:
            float: 相似度，0-1之间，越高越相似
        """
        if not color1 or not color2:
            return 0.0

        try:
            # 转换为numpy数组
            c1 = np.array(color1)
            c2 = np.array(color2)

            # 计算欧氏距离
            distance = np.sqrt(np.sum((c1 - c2) ** 2))
            # 颜色距离最大可能值为 sqrt(255^2 * 3) = 441.67
            max_distance = 442.0

            # 转换为相似度
            similarity = 1.0 - min(distance / max_distance, 1.0)
            return similarity

        except Exception as e:
            logger.error(f"计算颜色相似度出错: {str(e)}")
            return 0.0

    def calculate_target_score(self, track):
        """计算一个轨迹与目标的匹配得分

        参数:
            track: 跟踪对象

        返回值:
            float: 得分，0-1之间，越高越匹配
        """
        # 获取检测到的颜色信息
        shirt_color = track.upper_color if hasattr(track, 'upper_color') and track.upper_color else None
        pants_color = track.lower_color if hasattr(track, 'lower_color') and track.lower_color else None

        # 计算颜色相似度
        shirt_sim = self.color_similarity(self.target_shirt_color, shirt_color) if shirt_color else 0.0
        pants_sim = self.color_similarity(self.target_pants_color, pants_color) if pants_color else 0.0

        # 计算身体比例相似度
        body_ratio_sim = 0.0
        has_body_ratios = False

        if hasattr(track, 'body_ratios') and track.body_ratios:
            # 检查是否有有效的身体比例数据
            valid_ratios = sum(1 for r in track.body_ratios if r > 0)
            if valid_ratios > 4:  # 至少需要4个有效比例
                body_ratio_sim = self.body_ratio_similarity(track.body_ratios)
                has_body_ratios = True
                logger.debug(f"轨迹 ID{track.track_id} 身体比例相似度: {body_ratio_sim:.3f} (有效比例: {valid_ratios})")

        # 根据可用信息设置权重和计算总分
        feature_count = 0
        weighted_sum = 0.0

        # 上衣颜色 (权重 0.4)
        if shirt_color:
            weighted_sum += 0.4 * shirt_sim
            feature_count += 1

        # 下装颜色 (权重 0.3)
        if pants_color:
            weighted_sum += 0.3 * pants_sim
            feature_count += 1

        # 身体比例 (权重 0.3)
        if has_body_ratios:
            weighted_sum += 0.3 * body_ratio_sim
            feature_count += 1

        # 计算最终得分
        if feature_count > 0:
            # 根据可用特征数量调整得分
            base_score = weighted_sum / feature_count * (0.8 + 0.2 * feature_count)

            # 如果有所有特征，增加一点额外奖励
            if feature_count == 3:
                base_score *= 1.05

            # 如果只有一种特征，降低可信度
            if feature_count == 1:
                base_score *= 0.9
        else:
            # 没有任何特征信息
            base_score = 0.0

        # 如果已经在跟踪模式，考虑历史信息
        if self.mode == self.MODE_TRACKING and self.target_id == track.track_id:
            base_score = base_score * 1.2  # 提高已锁定目标的得分

        logger.info(
            f"轨迹 ID{track.track_id} 得分组成: 上衣={shirt_sim:.2f}, 下装={pants_sim:.2f}, 身体比例={body_ratio_sim:.2f}, 总分={base_score:.2f}")

        return min(base_score, 1.0)  # 确保得分不超过1.0

    def select_target(self, tracks):
        """从当前轨迹中选择最匹配的目标

        参数:
            tracks: 当前跟踪的轨迹列表

        返回值:
            best_track: 最匹配的轨迹，没有合适的则返回None
            score: 最高的匹配得分
        """
        if not tracks:
            return None, 0.0

        best_track = None
        best_score = 0.0

        for track in tracks:
            score = self.calculate_target_score(track)

            # 记录得分用于调试
            logger.debug(f"轨迹 ID{track.track_id}: 得分 {score:.3f}")

            if score > best_score:
                best_score = score
                best_track = track

        # 只有得分超过阈值才认为找到了目标
        min_score_threshold = 0.4  # 最低匹配阈值
        if best_score >= min_score_threshold:
            logger.info(f"已选择目标: ID{best_track.track_id}, 得分: {best_score:.3f}")
            return best_track, best_score
        else:
            logger.info(f"未找到合适的目标，最高得分: {best_score:.3f}")
            return None, best_score

    def update(self, detection_results, image=None):
        """更新跟踪器状态

        参数:
            detection_results: 检测结果列表
            image: 当前帧图像

        返回值:
            tuple: (track_results, target_track, mode)
                track_results: 所有跟踪结果
                target_track: 目标轨迹 (如果在跟踪模式下)
                mode: 当前模式
        """
        self.frame_count += 1

        # 使用ByteTracker更新所有轨迹
        all_tracks = self.base_tracker.update(detection_results, image)

        if not all_tracks:
            if self.mode == self.MODE_TRACKING:
                self.lost_frames += 1
                logger.debug(f"目标轨迹丢失，当前丢失帧数: {self.lost_frames}")

                # 检查是否需要切换回选择模式
                if self.lost_frames > self.max_lost_frames:
                    logger.info(f"目标丢失超过{self.max_lost_frames}帧，切换回选择模式")
                    self.mode = self.MODE_SELECTING
                    self.target_id = None
                    self.target_trajectory.clear()

            return all_tracks, None, self.mode

        # 根据当前模式处理
        if self.mode == self.MODE_SELECTING:
            # 选择模式：尝试找到匹配目标
            best_track, score = self.select_target(all_tracks)

            # 记录得分历史，用于稳定目标选择
            self.target_score_history.append(score)
            if len(self.target_score_history) > 10:
                self.target_score_history.pop(0)

            # 如果连续多帧得分都很高，则锁定目标
            if best_track and len(self.target_score_history) >= 3:
                recent_scores = self.target_score_history[-3:]
                if all(s >= 0.4 for s in recent_scores):
                    self.target_id = best_track.track_id
                    self.mode = self.MODE_TRACKING
                    self.lost_frames = 0
                    self.target_trajectory.clear()

                    # 记录位置用于轨迹显示
                    center_x = (best_track.tlbr[0] + best_track.tlbr[2]) / 2
                    center_y = (best_track.tlbr[1] + best_track.tlbr[3]) / 2
                    self.target_trajectory.append((int(center_x), int(center_y)))

                    logger.info(f"目标锁定: ID{self.target_id}，切换到跟踪模式")

            return all_tracks, best_track, self.mode

        else:  # 跟踪模式
            # 在当前轨迹中查找目标ID
            target_track = None
            for track in all_tracks:
                if track.track_id == self.target_id:
                    target_track = track
                    self.lost_frames = 0  # 重置丢失计数

                    # 更新轨迹
                    center_x = (track.tlbr[0] + track.tlbr[2]) / 2
                    center_y = (track.tlbr[1] + track.tlbr[3]) / 2
                    self.target_trajectory.append((int(center_x), int(center_y)))

                    break

            # 如果目标不在当前帧，增加丢失计数
            if target_track is None:
                self.lost_frames += 1
                logger.debug(f"目标轨迹丢失，当前丢失帧数: {self.lost_frames}")

                # 检查是否需要切换回选择模式
                if self.lost_frames > self.max_lost_frames:
                    logger.info(f"目标丢失超过{self.max_lost_frames}帧，切换回选择模式")
                    self.mode = self.MODE_SELECTING
                    self.target_id = None
                    # 清空轨迹
                    self.target_trajectory.clear()

            return all_tracks, target_track, self.mode


def read_target_features(xlsx_path):
    """从Excel文件中读取目标特征信息

    参数:
        xlsx_path: Excel文件路径

    返回值:
        tuple: (body_ratios, shirt_color, pants_color)
    """
    try:
        # 打开Excel文件
        wb = openpyxl.load_workbook(xlsx_path)
        sheet = wb.active

        # 读取身体比例数据 (前16行)
        body_ratios = []
        for i in range(1, 17):
            value = sheet.cell(row=i, column=1).value
            try:
                # 尝试转换为浮点数
                value = float(value) if value is not None else 0.0
            except (ValueError, TypeError):
                value = 0.0
            body_ratios.append(value)

        # 读取颜色数据 (第17和18行)
        shirt_color_str = sheet.cell(row=17, column=1).value
        pants_color_str = sheet.cell(row=18, column=1).value

        # 解析颜色字符串 "(R,G,B)"
        def parse_color(color_str):
            if not color_str:
                return (0, 0, 0)  # 默认黑色

            try:
                if isinstance(color_str, str):
                    # 处理字符串形式 "(R,G,B)"
                    color_str = color_str.strip()
                    if color_str.startswith('(') and color_str.endswith(')'):
                        # 使用eval解析元组字符串
                        return eval(color_str)
                    else:
                        # 尝试其他解析方法
                        color_str = color_str.replace('(', '').replace(')', '')
                        values = [int(x.strip()) for x in color_str.split(',')]
                        return tuple(values[:3])
                elif isinstance(color_str, tuple):
                    # 已经是元组形式
                    return color_str
                else:
                    return (0, 0, 0)  # 默认黑色
            except:
                return (0, 0, 0)  # 解析失败返回黑色

        shirt_color = parse_color(shirt_color_str)
        pants_color = parse_color(pants_color_str)

        logger.info(f"已读取目标特征: 身体比例数 {len(body_ratios)}, 上衣颜色 {shirt_color}, 下装颜色 {pants_color}")
        return body_ratios, shirt_color, pants_color

    except Exception as e:
        logger.error(f"读取目标特征失败: {str(e)}")
        # 返回默认值
        return [0.0] * 16, (0, 0, 0), (0, 0, 0)


def track_single_target(video_path=None, target_features_xlsx=None, output_path=None, control_car=False,
                        car_serial_port='COM7', callback=None, stop_event=None):
    """单目标跟踪主函数

    参数:
        video_path: 视频路径，None表示使用摄像头
        target_features_xlsx: 目标特征Excel文件路径
        output_path: 输出视频路径
        control_car: 是否启用小车控制
        car_serial_port: 小车串口端口
        callback: 回调函数，接收参数(frame, target_track, distance, command, mode)
        stop_event: 停止事件，用于外部终止跟踪
    """
    # 尝试导入姿态估计模块
    try:
        from obtain_features import get_keypoints, calculate_body_ratios
        pose_estimation_available = True
        logger.info("姿态估计模块已加载，将用于获取身体比例特征")
    except ImportError:
        pose_estimation_available = False
        logger.warning("姿态估计模块未加载，将不使用身体比例特征")
    # 读取目标特征
    if not target_features_xlsx:
        logger.error("未提供目标特征文件")
        return

    target_features = read_target_features(target_features_xlsx)
    if not target_features[0]:
        logger.error("读取目标特征失败")
        return

    # 跟踪参数
    args = {
        'track_thresh': 0.4,  # 跟踪阈值
        'track_buffer': 60,  # 轨迹缓冲
        'match_thresh': 0.8,  # 匹配阈值
        'color_weight': 0.6  # 颜色权重
    }

    my_instruction1 = ""
    my_instruction2 = ""

    # 初始化单目标跟踪器
    tracker = SingleTargetTracker(args, target_features, max_lost_time=60, control_car=control_car,
                                  car_serial_port=car_serial_port)

    # 打开视频
    if video_path:
        cap = cv2.VideoCapture(video_path)
    else:
        # 使用双目相机，设置分辨率为宽屏模式以获取左右图像
        cap = cv2.VideoCapture(1)
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1280)  # 双目宽屏模式
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

    # 获取视频原始帧率
    original_fps = cap.get(cv2.CAP_PROP_FPS)
    logger.info(f"原始视频帧率: {original_fps} FPS")

    # 准备输出视频
    writer = None
    if output_path:
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        writer = cv2.VideoWriter(
            output_path,
            cv2.VideoWriter_fourcc(*'mp4v'),
            original_fps,  # 使用原始帧率
            (width, height)
        )

    # 轨迹颜色映射
    track_colors = {}
    color_palette = [
        (255, 0, 0), (0, 255, 0), (0, 0, 255), (255, 255, 0),
        (0, 255, 255), (255, 0, 255), (128, 128, 0), (0, 128, 128),
        (128, 0, 128), (255, 128, 0), (128, 255, 0), (0, 255, 128)
    ]

    # 目标颜色（始终使用红色标记目标）
    TARGET_COLOR = (0, 0, 255)  # 红色，BGR格式

    frame_count = 0
    prev_frame_time = 0  # 上一帧处理的时间点

    # 初始化双目立体视觉系统
    stereo_config = StereoConfig()
    camera_config = stereoCamera()

    # 计算校正变换矩阵
    height, width = 480, 640  # 根据实际相机调整
    map1x, map1y, map2x, map2y, Q = getRectifyTransform(height, width, camera_config)

    try:
        while True:
            # 检查是否请求停止
            if stop_event is not None and stop_event.is_set():
                logger.info("收到停止请求，终止跟踪")
                break

            # 读取下一帧
            ret, frame = cap.read()
            if not ret:
                break

            left_img = None
            right_img = None
            if video_path is None:
                # 分割左右图像
                frame_width = frame.shape[1]
                mid_x = frame_width // 2
                left_img = frame[:, :mid_x]
                right_img = frame[:, mid_x:]
                frame = left_img

            # 记录这一帧开始处理的时间
            prev_frame_time = time.time()

            frame_count += 1
            logger.info(f"正在处理第 {frame_count} 帧")

            # 准备跟踪数据
            tracking_inputs = []

            # 检测目标
            start_time = time.time()
            detection_results = detect_picture_with_confidence(frame)
            detect_time = time.time() - start_time

            # 从检测结果中获取身体比例
            start_time = time.time()
            calculate_ratios = (tracker.mode == SingleTargetTracker.MODE_SELECTING)
            if calculate_ratios:
                logger.debug("目标选择模式：计算身体比例")
                tracking_inputs = extract_body_ratios_from_detections(frame, detection_results, calculate_ratios=True)
            else:
                logger.debug("目标跟踪模式：跳过身体比例计算")
                tracking_inputs = extract_body_ratios_from_detections(frame, detection_results, calculate_ratios=False)
            ratio_time = time.time() - start_time

            # 更新跟踪器
            start_time = time.time()
            all_tracks, target_track, mode = tracker.update(tracking_inputs, frame)
            track_time = time.time() - start_time

            # 创建可视化帧
            viz_frame = frame.copy()

            # 确定当前模式文本
            mode_text = "target_select" if mode == SingleTargetTracker.MODE_SELECTING else "target_tracker"

            # 如果检测到目标人物，计算距离
            distance = None
            if target_track and mode == SingleTargetTracker.MODE_TRACKING and video_path is None and left_img is not None and right_img is not None:
                # 进行立体视觉处理
                # 畸变校正
                iml = undistortion(left_img, camera_config.cam_matrix_left, camera_config.distortion_l)
                imr = undistortion(right_img, camera_config.cam_matrix_right, camera_config.distortion_r)

                # 预处理
                iml_, imr_ = preprocess(iml, imr)

                # 立体校正
                rectified_left, rectified_right = rectifyImage(iml_, imr_, map1x, map1y, map2x, map2y)

                # 计算视差图
                disparity, _, _ = stereoMatchSGBM_WLS(rectified_left, rectified_right, stereo_config)

                # 计算3D点云
                points_3d = reprojectTo3D(disparity, Q)

                # 获取目标中心点
                target_tlbr = target_track.tlbr
                center_x = int((target_tlbr[0] + target_tlbr[2]) / 2)
                center_y = int((target_tlbr[1] + target_tlbr[3]) / 2)

                # 测量距离
                distance = measure_distance(points_3d, center_x, center_y)

                cv2.circle(viz_frame, (center_x, center_y), 6, (0, 0, 255), -1)  # 实心红色圆点

            if distance is not None:
                distance_text = f"distance: {distance:.2f}meter"
                viz_frame = cv2.putText(
                    viz_frame,
                    distance_text,
                    (10, 180),
                    cv2.FONT_HERSHEY_SIMPLEX,
                    0.7,
                    (0, 255, 0),
                    2
                )
                # 保存双目距离到跟踪器并计算指令
                tracker.stereo_distance = distance
                tracker.calculate_stereo_distance_instruction(distance)

                # 根据目标位置更新方向
                if target_track:
                    center_x = (target_track.tlbr[0] + target_track.tlbr[2]) / 2
                    center_y = (target_track.tlbr[1] + target_track.tlbr[3]) / 2
                    tracker.check_position_right_left(int(center_x), int(center_y))

                # 在显示结果前添加小车控制命令发送:
                if tracker.control_car:
                    my_instruction1 = tracker.my_instruction1
                    my_instruction2 = tracker.my_instruction2
                    tracker.send_car_command()

            # 可视化所有轨迹
            for track in all_tracks:
                if track.track_id not in track_colors:
                    track_colors[track.track_id] = color_palette[track.track_id % len(color_palette)]

                # 确定颜色 - 目标使用特殊颜色，其他使用轨迹ID颜色
                is_target = (target_track is not None and track.track_id == target_track.track_id)
                color = TARGET_COLOR if is_target else track_colors[track.track_id]

                # 线宽 - 目标轨迹使用粗线条
                thickness = 3 if is_target else 2

                # 绘制边界框
                tlbr = track.tlbr
                cv2.rectangle(
                    viz_frame,
                    (int(tlbr[0]), int(tlbr[1])),
                    (int(tlbr[2]), int(tlbr[3])),
                    color, thickness
                )

                # 显示ID
                id_text = f"target ID:{track.track_id}" if is_target else f"ID:{track.track_id}"
                cv2.putText(
                    viz_frame,
                    id_text,
                    (int(tlbr[0]), int(tlbr[1] - 10)),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, thickness
                )

                # 绘制预测边界框（虚线）
                if hasattr(track, 'predicted_mean') and track.predicted_mean is not None:
                    pred_tlbr = track.predicted_tlbr
                    # 使用相同颜色但更浅的色调
                    pred_color = tuple(min(255, c + 70) for c in color)
                    # 绘制虚线矩形
                    draw_dashed_rectangle(
                        viz_frame,
                        (int(pred_tlbr[0]), int(pred_tlbr[1])),
                        (int(pred_tlbr[2]), int(pred_tlbr[3])),
                        pred_color, 1
                    )

                    # 在预测框左上角标注"预测"
                    cv2.putText(
                        viz_frame,
                        "predict",
                        (int(pred_tlbr[0]), int(pred_tlbr[1] - 5)),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, pred_color, 1
                    )

                # 显示颜色样本
                if hasattr(track, 'upper_color') and track.upper_color:
                    upper_color = validate_color(track.upper_color)
                    cv2.rectangle(
                        viz_frame,
                        (int(tlbr[0]), int(tlbr[1])),
                        (int(tlbr[0] + 15), int(tlbr[1] + 15)),
                        upper_color, -1
                    )

                if hasattr(track, 'lower_color') and track.lower_color:
                    lower_color = validate_color(track.lower_color)
                    cv2.rectangle(
                        viz_frame,
                        (int(tlbr[0] + 20), int(tlbr[1])),
                        (int(tlbr[0] + 35), int(tlbr[1] + 15)),
                        lower_color, -1
                    )

            # 绘制目标轨迹
            # if mode == SingleTargetTracker.MODE_TRACKING and len(tracker.target_trajectory) > 1:
            #     for i in range(1, len(tracker.target_trajectory)):
            #         cv2.line(
            #             viz_frame,
            #             tracker.target_trajectory[i - 1],
            #             tracker.target_trajectory[i],
            #             TARGET_COLOR,
            #             2
            #         )

            # 显示当前模式
            cv2.putText(
                viz_frame,
                f"mode: {mode_text}",
                (10, 30),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2
            )

            actual_fps = 1.0 / (time.time() - prev_frame_time) if prev_frame_time > 0 else 0
            cv2.putText(
                viz_frame,
                f"FPS: {actual_fps:.1f}/{original_fps:.1f}",
                (10, 210),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2
            )

            # 显示目标信息
            if target_track:
                cv2.putText(
                    viz_frame,
                    f"target ID: {target_track.track_id}",
                    (10, 60),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, TARGET_COLOR, 2
                )
            elif mode == SingleTargetTracker.MODE_SELECTING:
                cv2.putText(
                    viz_frame,
                    "finding target...",
                    (10, 60),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2
                )
            else:
                cv2.putText(
                    viz_frame,
                    f"target lost，{tracker.lost_frames}/{tracker.max_lost_frames}",
                    (10, 60),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2
                )

            # 显示目标特征信息
            cv2.putText(
                viz_frame,
                f"target_upper: {tracker.target_shirt_color}",
                (10, 90),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, tracker.target_shirt_color, 2
            )

            cv2.putText(
                viz_frame,
                f"target_lower: {tracker.target_pants_color}",
                (10, 120),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, tracker.target_pants_color, 2
            )

            # 显示处理时间
            cv2.putText(
                viz_frame,
                f"detect: {detect_time:.3f}s, track: {track_time:.3f}s",
                (10, 150),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2
            )

            # 在这里调用回调函数，传递结果
            if callback:
                # 获取当前控制命令
                command = f"{my_instruction1} {my_instruction2}".strip()
                # 调用回调函数，传递可视化图像、目标轨迹、距离、控制命令和当前模式
                callback(viz_frame, target_track, distance, command, mode)
                my_instruction1 = ""
                my_instruction2 = ""

            # 显示结果
            cv2.imshow("single_tracker", viz_frame)

            # 保存视频
            if writer:
                writer.write(viz_frame)

            # 按ESC退出
            if cv2.waitKey(1) == 27:
                break

    except Exception as e:
        logger.error(f"跟踪过程中出错: {str(e)}")
        traceback.print_exc()

    finally:
        # 释放资源
        cap.release()
        if writer:
            writer.release()
        cv2.destroyAllWindows()


def draw_dashed_rectangle(img, pt1, pt2, color, thickness=1, dash_length=8, gap_length=5):
    """绘制虚线矩形

    参数:
        img: 图像
        pt1: 左上角坐标 (x1, y1)
        pt2: 右下角坐标 (x2, y2)
        color: 颜色 (B, G, R)
        thickness: 线宽
        dash_length: 虚线段长度
        gap_length: 虚线间隔长度
    """
    x1, y1 = pt1
    x2, y2 = pt2

    # 绘制水平线段
    for x in range(x1, x2, dash_length + gap_length):
        x_end = min(x + dash_length, x2)
        cv2.line(img, (x, y1), (x_end, y1), color, thickness)
        cv2.line(img, (x, y2), (x_end, y2), color, thickness)

    # 绘制垂直线段
    for y in range(y1, y2, dash_length + gap_length):
        y_end = min(y + dash_length, y2)
        cv2.line(img, (x1, y), (x1, y_end), color, thickness)
        cv2.line(img, (x2, y), (x2, y_end), color, thickness)

def extract_body_ratios_from_detections(frame, detection_results, calculate_ratios=True):
    """
    通过裁剪检测区域进行姿态估计，确保身体比例与服装检测结果一一对应

    参数:
        frame: 原始图像
        detection_results: 服装检测结果列表，每项为(上衣坐标, 下装坐标, 上衣颜色, 下装颜色, 上衣置信度, 下装置信度)

    返回值:
        list: 匹配后的结果列表，每项为(上衣坐标, 下装坐标, 上衣颜色, 下装颜色, 上衣置信度, 下装置信度, 身体比例)
    """
    if not calculate_ratios:
        return [(r[0], r[1], r[2], r[3], r[4], r[5], None) for r in detection_results]
    # 尝试导入姿态估计模块
    try:
        from obtain_features import get_keypoints, calculate_body_ratios
        from yolov8_pose import RKNNLite, letterbox_resize
        pose_estimation_available = True
    except ImportError:
        logger.warning("姿态估计模块未加载，将不使用身体比例特征")
        # 直接返回原始结果，添加None作为身体比例
        return [(r[0], r[1], r[2], r[3], r[4], r[5], None) for r in detection_results]

    # 加载姿态估计模型（如果未加载）
    pose_model = None
    if not hasattr(extract_body_ratios_from_detections, 'pose_model'):
        try:
            logger.info("正在加载姿态估计模型...")
            pose_model = RKNNLite()
            pose_model_path = '/home/monster/桌面/project/model/yolov8_pose.rknn'
            ret = pose_model.load_rknn(pose_model_path)
            if ret != 0:
                logger.error(f"姿态估计模型加载失败: {ret}")
                pose_model = None
            else:
                ret = pose_model.init_runtime()
                if ret != 0:
                    logger.error(f"姿态估计模型运行时初始化失败: {ret}")
                    pose_model = None
                else:
                    logger.info("姿态估计模型加载成功")
            extract_body_ratios_from_detections.pose_model = pose_model
        except Exception as e:
            logger.error(f"加载姿态估计模型出错: {str(e)}")
            pose_model = None
    else:
        pose_model = extract_body_ratios_from_detections.pose_model

    # 如果模型不可用，直接返回原始结果
    if pose_model is None:
        return [(r[0], r[1], r[2], r[3], r[4], r[5], None) for r in detection_results]

    # 准备结果列表
    results_with_ratios = []

    # 计算每个检测结果对应的人体区域，并进行姿态估计
    from rknn_colour_detect import Determine_the_position_of_the_entire_body

    for detection in detection_results:
        # 解析检测结果
        upper_pos, lower_pos = detection[0], detection[1]
        upper_color, lower_color = detection[2] if len(detection) > 2 else None, detection[3] if len(
            detection) > 3 else None
        upper_conf = detection[4] if len(detection) > 4 else 0.5
        lower_conf = detection[5] if len(detection) > 5 else 0.5

        # 创建包含原始检测信息的结果
        result_with_ratios = list(detection)

        try:
            # 计算整体位置
            body_positions = Determine_the_position_of_the_entire_body(upper_pos, lower_pos, frame)

            # 如果计算出人体位置
            if body_positions and len(body_positions[0]) == 4:
                bbox = body_positions[0]  # [xmin, ymin, xmax, ymax]

                # 确保边界合法
                height, width = frame.shape[:2]
                xmin = max(0, bbox[0])
                ymin = max(0, bbox[1])
                xmax = min(width, bbox[2])
                ymax = min(height, bbox[3])

                # 检查边界框是否有效
                if xmax > xmin and ymax > ymin:
                    # 裁剪人体区域
                    person_img = frame[ymin:ymax, xmin:xmax].copy()

                    # 进行姿态估计
                    keypoints_list = get_keypoints(person_img, pose_model)

                    # 如果检测到关键点
                    if keypoints_list and len(keypoints_list) > 0:
                        # 选择第一组关键点（裁剪后的图像应该只有一个人）
                        keypoints = keypoints_list[0]

                        # 需要调整关键点坐标，因为它们是相对于裁剪图像的
                        for i in range(len(keypoints)):
                            if keypoints[i][2] > 0:  # 如果关键点有效
                                keypoints[i][0] += xmin  # 调整x坐标
                                keypoints[i][1] += ymin  # 调整y坐标

                        # 计算身体比例
                        body_ratios = calculate_body_ratios([keypoints])

                        # 添加身体比例到结果中
                        if body_ratios:
                            if len(result_with_ratios) > 6:  # 如果已有占位，替换
                                result_with_ratios[6] = body_ratios
                            else:  # 否则追加
                                result_with_ratios.append(body_ratios)

                            logger.debug(f"成功计算身体比例: {body_ratios[:3]}...")
                        else:
                            # 添加None作为占位
                            if len(result_with_ratios) <= 6:
                                result_with_ratios.append(None)
                                logger.debug(f"没有身体比例")
                    else:
                        # 添加None作为占位
                        if len(result_with_ratios) <= 6:
                            result_with_ratios.append(None)
                            logger.debug(f"没有检测到关键点")
                else:
                    # 边界框无效，添加None
                    if len(result_with_ratios) <= 6:
                        result_with_ratios.append(None)
                        logger.debug(f"边界框无效")
            else:
                # 没有有效的人体位置，添加None
                if len(result_with_ratios) <= 6:
                    result_with_ratios.append(None)
                    logger.debug(f"没有有效的人体位置")

        except Exception as e:
            logger.error(f"计算身体比例时出错: {str(e)}")
            # 确保结果有身体比例占位
            if len(result_with_ratios) <= 6:
                result_with_ratios.append(None)

        # 将结果添加到列表
        results_with_ratios.append(tuple(result_with_ratios))

    return results_with_ratios


if __name__ == "__main__":
    import argparse,time

    time.sleep(5)

    # 获取默认串口和可用串口列表
    default_port = "/dev/ttyACM0"
    available_ports = detect_serial_ports()

    # 命令行参数解析
    parser = argparse.ArgumentParser(description='融合颜色的人员跟踪系统')
    parser.add_argument('--video', type=str, help='输入视频路径，不指定则使用摄像头',
                        default=None)
    parser.add_argument('--output', type=str, help='输出视频路径', default="result_video.mp4")
    parser.add_argument('--mode', type=str, default='single', choices=['multi', 'single'],
                        help='跟踪模式: multi(多目标跟踪), single(单目标跟踪)')
    parser.add_argument('--target', type=str, help='目标特征文件(.xlsx)，单目标跟踪模式必需',
                        default="/home/monster/桌面/project/features-data/person.xlsx")
    parser.add_argument('--control-car', action='store_true', help='启用小车控制',default=True)
    parser.add_argument('--car-port', type=str, default=default_port,
                        help=f'小车串口端口 (可用端口: {", ".join(available_ports) if available_ports else "无"})')
    parser.add_argument('--list-ports', action='store_true', help='列出所有可用串口并退出',default=False)

    args = parser.parse_args()

    # 如果只是列出可用串口，则打印并退出
    if args.list_ports:
        print("可用串口设备:")
        if available_ports:
            for port in available_ports:
                print(f"  - {port}")
        else:
            print("  未检测到可用串口设备")
        exit(0)

    # 根据选择的模式启动不同的跟踪器
    if args.mode == 'single':
        if not args.target:
            logger.error("单目标跟踪模式需要提供目标特征文件(--target)")
            parser.print_help()
            exit(1)

        logger.info(f"启动单目标跟踪模式，目标特征文件: {args.target}")
        track_single_target(
            video_path=args.video,  # None表示使用摄像头
            target_features_xlsx=args.target,
            output_path=args.output,
            control_car=args.control_car,
            car_serial_port=args.car_port
        )
    else:
        # 多目标跟踪模式
        logger.info("启动多目标跟踪模式")
        track_with_color_features(
            video_path=args.video,
            output_path=args.output
        )