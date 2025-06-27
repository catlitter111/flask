#!/usr/bin/env python
# -*- coding: utf-8 -*-

import argparse
import cv2
import os
import time
import numpy as np
import logging
from datetime import datetime
from pathlib import Path
import traceback
import openpyxl
from PIL import Image, ImageDraw, ImageFont
import sys

# 导入RKNN服装检测模块和YOLO姿态估计模块
try:
    from rknn_colour_detect import Obtain_the_target_color, load_model as load_colour_model, get_model, detect_picture
    from yolov8_pose import letterbox_resize, DetectBox, NMS, process
    from rknnlite.api import RKNNLite
except ImportError as e:
    print(f"导入错误: {e}. 请确保已安装所有必要的依赖库。")
    exit(1)

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("feature_extraction.log")
    ]
)
logger = logging.getLogger("特征提取")

# 常量定义
OUTPUT_DIR = 'features-data'
DEFAULT_COLOUR_MODEL_PATH = 'best3.rknn'
DEFAULT_POSE_MODEL_PATH = '/home/monster/桌面/project/model/yolov8_pose.rknn'
DEFAULT_IMAGE_PATH = "target.png"
TARGET_SIZE = (640, 640)


def setup_output_directory(directory):
    """
    创建输出目录
    """
    os.makedirs(directory, exist_ok=True)
    logger.info(f"输出目录已设置: {directory}")


def load_models():
    """
    加载服装检测和姿态估计模型
    """
    logger.info("正在加载模型...")

    # 加载服装检测模型（使用rknn_colour中的函数）
    colour_model = get_model()
    if colour_model is None:
        logger.error("服装检测模型加载失败")
        return None, None

    # 加载姿态估计模型
    pose_model = RKNNLite()
    ret = pose_model.load_rknn(DEFAULT_POSE_MODEL_PATH)
    if ret != 0:
        logger.error(f"姿态估计模型加载失败: {ret}")
        return colour_model, None

    ret = pose_model.init_runtime()
    if ret != 0:
        logger.error(f"姿态估计模型运行时初始化失败: {ret}")
        return colour_model, None

    logger.info("模型加载成功")
    return colour_model, pose_model


def get_keypoints(image, pose_model):
    """
    使用YOLOv8 Pose模型获取人体关键点
    """
    if pose_model is None:
        logger.error("姿态估计模型未加载")
        return None

    try:
        # 使用letterbox_resize处理图像
        letterbox_img, aspect_ratio, offset_x, offset_y = letterbox_resize(image, TARGET_SIZE, 56)
        infer_img = letterbox_img[..., ::-1]  # BGR2RGB
        infer_img = np.expand_dims(infer_img, axis=0)  # 添加批次维度

        # 推理
        results = pose_model.inference(inputs=[infer_img])

        # 处理输出
        outputs = []
        keypoints = results[3]
        for i, x in enumerate(results[:3]):
            index, stride = 0, 0
            if x.shape[2] == 20:
                stride = 32
                index = 20 * 4 * 20 * 4 + 20 * 2 * 20 * 2
            if x.shape[2] == 40:
                stride = 16
                index = 20 * 4 * 20 * 4
            if x.shape[2] == 80:
                stride = 8
                index = 0
            feature = x.reshape(1, 65, -1)
            output = process(feature, keypoints, index, x.shape[3], x.shape[2], stride)
            outputs = outputs + output

        # 非极大值抑制
        pred_boxes = NMS(outputs)

        # 处理关键点
        all_keypoints = []
        for i in range(len(pred_boxes)):
            kpts = pred_boxes[i].keypoint.reshape(-1, 3)  # [x, y, conf]
            kpts[..., 0] = (kpts[..., 0] - offset_x) / aspect_ratio
            kpts[..., 1] = (kpts[..., 1] - offset_y) / aspect_ratio
            all_keypoints.append(kpts)

        return all_keypoints

    except Exception as e:
        logger.error(f"获取关键点时出错: {e}")
        logger.error(traceback.format_exc())
        return None


def calculate_body_ratios(keypoints):
    """
    根据人体关键点计算身体比例
    """
    if not keypoints or len(keypoints) == 0:
        logger.error("没有有效的关键点数据")
        return None

    # 使用第一个人的关键点（如果检测到多个人）
    kpts = keypoints[0]

    # 定义身体比例
    ratios = []

    try:
        # 常用的关键点索引（以YOLOv8 Pose的关键点顺序为准）:
        # 0: 鼻子, 1: 左眼, 2: 右眼, 3: 左耳, 4: 右耳, 5: 左肩, 6: 右肩
        # 7: 左肘, 8: 右肘, 9: 左腕, 10: 右腕, 11: 左髋, 12: 右髋
        # 13: 左膝, 14: 右膝, 15: 左踝, 16: 右踝

        # 检查关键点有效性的函数
        def is_valid(p1_idx, p2_idx):
            return (kpts[p1_idx][2] > 0.5 and kpts[p2_idx][2] > 0.5 and
                    kpts[p1_idx][0] != 0 and kpts[p1_idx][1] != 0 and
                    kpts[p2_idx][0] != 0 and kpts[p2_idx][1] != 0)

        # 计算两点之间的距离
        def distance(p1_idx, p2_idx):
            if not is_valid(p1_idx, p2_idx):
                return 0
            return np.sqrt(((kpts[p1_idx][0] - kpts[p2_idx][0]) ** 2) +
                           ((kpts[p1_idx][1] - kpts[p2_idx][1]) ** 2))

        # 计算一系列有意义的身体比例

        # 1. 上肢与下肢比例
        upper_limb = (distance(5, 7) + distance(7, 9) + distance(6, 8) + distance(8, 10)) / 4
        lower_limb = (distance(11, 13) + distance(13, 15) + distance(12, 14) + distance(14, 16)) / 4
        if upper_limb > 0 and lower_limb > 0:
            ratios.append(upper_limb / lower_limb)
        else:
            ratios.append(0)

        # 2. 躯干与身高比例
        torso_height = (distance(5, 11) + distance(6, 12)) / 2
        body_height = ((distance(0, 15) + distance(0, 16)) / 2)
        if torso_height > 0 and body_height > 0:
            ratios.append(torso_height / body_height)
        else:
            ratios.append(0)

        # 3. 肩宽与身高比例
        shoulder_width = distance(5, 6)
        if shoulder_width > 0 and body_height > 0:
            ratios.append(shoulder_width / body_height)
        else:
            ratios.append(0)

        # 4. 臀宽与肩宽比例
        hip_width = distance(11, 12)
        if hip_width > 0 and shoulder_width > 0:
            ratios.append(hip_width / shoulder_width)
        else:
            ratios.append(0)

        # 5. 头部与躯干比例
        head_height = (distance(0, 5) + distance(0, 6)) / 2
        if head_height > 0 and torso_height > 0:
            ratios.append(head_height / torso_height)
        else:
            ratios.append(0)

        # 6. 手臂与身高比例
        arm_length = (distance(5, 9) + distance(6, 10)) / 2
        if arm_length > 0 and body_height > 0:
            ratios.append(arm_length / body_height)
        else:
            ratios.append(0)

        # 7. 腿长与身高比例
        leg_length = (distance(11, 15) + distance(12, 16)) / 2
        if leg_length > 0 and body_height > 0:
            ratios.append(leg_length / body_height)
        else:
            ratios.append(0)

        # 8-16: 添加更多比例以达到16个比例
        # 例如：上臂与下臂比例、大腿与小腿比例等

        # 8. 上臂与下臂比例
        upper_arm = (distance(5, 7) + distance(6, 8)) / 2
        lower_arm = (distance(7, 9) + distance(8, 10)) / 2
        if upper_arm > 0 and lower_arm > 0:
            ratios.append(upper_arm / lower_arm)
        else:
            ratios.append(0)

        # 9. 大腿与小腿比例
        thigh = (distance(11, 13) + distance(12, 14)) / 2
        calf = (distance(13, 15) + distance(14, 16)) / 2
        if thigh > 0 and calf > 0:
            ratios.append(thigh / calf)
        else:
            ratios.append(0)

        # 10. 躯干与腿长比例
        if torso_height > 0 and leg_length > 0:
            ratios.append(torso_height / leg_length)
        else:
            ratios.append(0)

        # 11. 手臂与腿长比例
        if arm_length > 0 and leg_length > 0:
            ratios.append(arm_length / leg_length)
        else:
            ratios.append(0)

        # 12. 肩宽与髋宽比例（再次计算是因为这是重要的比例）
        if shoulder_width > 0 and hip_width > 0:
            ratios.append(shoulder_width / hip_width)
        else:
            ratios.append(0)

        # 13. 头围与身高比例（估算头围）
        head_width = (distance(3, 4) * 1.2)  # 估算头围
        if head_width > 0 and body_height > 0:
            ratios.append(head_width / body_height)
        else:
            ratios.append(0)

        # 14. 脚长与身高比例（估算脚长）
        foot_length = (distance(15, 16) * 0.7)  # 估算脚长
        if foot_length > 0 and body_height > 0:
            ratios.append(foot_length / body_height)
        else:
            ratios.append(0)

        # 15. 脚踝宽与身高比例
        ankle_width = (distance(15, 16))
        if ankle_width > 0 and body_height > 0:
            ratios.append(ankle_width / body_height)
        else:
            ratios.append(0)

        # 16. 腰围与身高比例（估算腰围）
        waist = (hip_width * 0.85)  # 估算腰围
        if waist > 0 and body_height > 0:
            ratios.append(waist / body_height)
        else:
            ratios.append(0)

        # 确保我们有16个比例
        while len(ratios) < 16:
            ratios.append(0)

        return ratios

    except Exception as e:
        logger.error(f"计算身体比例时出错: {e}")
        logger.error(traceback.format_exc())
        return None


def obtain_features(image, name):
    """
    从单个图像中提取特征，整合服装检测和姿态估计
    """
    start_time = time.time()
    logger.info(f"开始处理图像: {name}")

    try:
        # 确保图像格式正确
        if isinstance(image, str):
            img = cv2.imread(image)
            if img is None:
                raise ValueError(f"无法读取图像: {image}")
        else:
            img = image.copy()

        ration_img = img.copy()
        # img = cv2.resize(img,(640,640))

        # 创建结果图像的副本
        canvas = img.copy()

        # 加载模型
        colour_model, pose_model = load_models()

        # 获取目标颜色（服装检测）
        logger.info("正在获取目标颜色...")
        pairs = Obtain_the_target_color(img)
        logger.info(f"检测到 {len(pairs)} 对服装")

        # 提取服装颜色
        shirt_color = (0, 0, 0)  # 默认黑色
        pants_color = (0, 0, 0)  # 默认黑色

        for pair in pairs:
            # 检查是否有上衣颜色
            if len(pair) > 2 and pair[2]:
                shirt_color = pair[2]

            # 检查是否有裤子颜色
            if len(pair) > 3 and pair[3]:
                pants_color = pair[3]

            # 在图像上标记上衣和下装
            if len(pair[0]) > 1:  # 有上衣
                x1, y1, x2, y2 = pair[0][:4]
                cv2.rectangle(canvas, (x1, y1), (x2, y2), (255, 0, 0), 2)
                # cv2.putText(canvas, f"上衣: {shirt_color}", (x1, y1 - 10),
                #             cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 1)

            if len(pair[1]) > 1:  # 有下装
                x1, y1, x2, y2 = pair[1][:4]
                cv2.rectangle(canvas, (x1, y1), (x2, y2), (0, 255, 0), 2)
                # cv2.putText(canvas, f"下装: {pants_color}", (x1, y1 - 10),
                #             cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)

        # 获取人体关键点
        logger.info("正在获取人体关键点...")
        keypoints = get_keypoints(ration_img, pose_model)

        # 如果检测到关键点，在图像上标记
        if keypoints and len(keypoints) > 0:
            logger.info(f"检测到 {len(keypoints)} 个人体")

            # 在图像上绘制关键点和骨架
            pose_palette = np.array([[255, 128, 0], [255, 153, 51], [255, 178, 102], [230, 230, 0], [255, 153, 255],
                                     [153, 204, 255], [255, 102, 255], [255, 51, 255], [102, 178, 255], [51, 153, 255],
                                     [255, 153, 153], [255, 102, 102], [255, 51, 51], [153, 255, 153], [102, 255, 102],
                                     [51, 255, 51], [0, 255, 0], [0, 0, 255], [255, 0, 0], [255, 255, 255]],
                                    dtype=np.uint8)

            kpt_color = pose_palette[[16, 16, 16, 16, 16, 0, 0, 0, 0, 0, 0, 9, 9, 9, 9, 9, 9]]
            skeleton = [[16, 14], [14, 12], [17, 15], [15, 13], [12, 13], [6, 12], [7, 13], [6, 7], [6, 8],
                        [7, 9], [8, 10], [9, 11], [2, 3], [1, 2], [1, 3], [2, 4], [3, 5], [4, 6], [5, 7]]
            limb_color = pose_palette[[9, 9, 9, 9, 7, 7, 7, 0, 0, 0, 0, 0, 16, 16, 16, 16, 16, 16, 16]]

            for person_kpts in keypoints:
                # 绘制关键点
                for k, keypoint in enumerate(person_kpts):
                    x, y, conf = keypoint
                    color_k = [int(x) for x in kpt_color[k]]
                    if conf > 0.5 and x != 0 and y != 0:
                        cv2.circle(canvas, (int(x), int(y)), 5, color_k, -1, lineType=cv2.LINE_AA)

                # 绘制骨架
                for k, sk in enumerate(skeleton):
                    pos1 = (int(person_kpts[(sk[0] - 1), 0]), int(person_kpts[(sk[0] - 1), 1]))
                    pos2 = (int(person_kpts[(sk[1] - 1), 0]), int(person_kpts[(sk[1] - 1), 1]))

                    conf1 = person_kpts[(sk[0] - 1), 2]
                    conf2 = person_kpts[(sk[1] - 1), 2]

                    if conf1 > 0.5 and conf2 > 0.5 and pos1[0] != 0 and pos1[1] != 0 and pos2[0] != 0 and pos2[1] != 0:
                        cv2.line(canvas, pos1, pos2, [int(x) for x in limb_color[k]], thickness=2, lineType=cv2.LINE_AA)

        # 计算身体比例
        body_ratios = calculate_body_ratios(keypoints)

        if body_ratios:
            logger.info("身体比例计算成功")

            # 在图像上显示一些关键比例
            ratio_names = [
                "上肢/下肢", "躯干/身高", "肩宽/身高", "臀宽/肩宽",
                "头部/躯干", "手臂/身高", "腿长/身高", "上臂/下臂"
            ]

            # for i, (ratio, name) in enumerate(zip(body_ratios[:8], ratio_names)):
            #     if ratio > 0:
            #         cv2.putText(canvas, f"{name}: {ratio:.2f}", (10, 30 + i * 20),
            #                     cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1)
        else:
            logger.warning("身体比例计算失败，使用默认值")
            body_ratios = [0] * 16

        # 组合所有特征：身体比例+衣物颜色
        person_ratios = body_ratios + [shirt_color, pants_color]

        name = 'person'

        # 保存结果
        result_path = Path(OUTPUT_DIR) / f"{name}_result.jpg"
        cv2.imwrite(str(result_path), canvas)
        logger.info(f"结果图像已保存到: {result_path}")

        # 保存特征数据到Excel
        excel_path = Path(OUTPUT_DIR) / f"{name}.xlsx"
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sanitize_sheet_name(f'特征_{name}')

        # 添加比例数据 - 直接将数据保存到A列
        for i, ratio in enumerate(body_ratios):
            sheet.cell(row=i + 1, column=1).value = ratio

        # 添加颜色数据 - 直接添加到A17和A18单元格
        sheet.cell(row=17, column=1).value = str(shirt_color)
        sheet.cell(row=18, column=1).value = str(pants_color)

        wb.save(str(excel_path))
        logger.info(f"特征数据已保存到: {excel_path}")

        logger.info(f"图像处理完成，耗时: {time.time() - start_time:.2f}秒")
        return canvas, person_ratios

    except Exception as e:
        logger.error(f"处理图像时出错: {e}")
        logger.error(traceback.format_exc())
        return img, []


def sanitize_sheet_name(name, replacement='_'):
    """
    清理Excel工作表名称，替换不允许的字符

    Args:
        name: 原始名称
        replacement: 替换字符，默认为下划线

    Returns:
        清理后的名称
    """
    # Excel工作表名称不允许包含的字符
    invalid_chars = [':', '/', '\\', '?', '*', '[', ']', "'", '<', '>']

    # 替换所有不合法字符
    result = name
    for char in invalid_chars:
        result = result.replace(char, replacement)

    # 确保名称不超过31个字符（Excel工作表名称的最大长度）
    if len(result) > 31:
        result = result[:31]

    return result


def cv2_put_chinese_text(img, text, org, fontFace, fontScale, color, thickness=1, lineType=cv2.LINE_8,
                         bottomLeftOrigin=False):
    """
    在OpenCV图像上绘制中文文本，参数与cv2.putText一致
    """
    # 计算对应的PIL字体大小
    base_size = 20  # OpenCV的FONT_HERSHEY_SIMPLEX在fontScale=1时大约20像素
    font_size = int(base_size * fontScale)

    # 转换颜色格式，从BGR到RGB
    color_rgb = (color[2], color[1], color[0])

    # 将OpenCV图像转换为PIL图像
    img_pil = Image.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))

    # 创建绘图对象
    draw = ImageDraw.Draw(img_pil)

    # 尝试找到系统中的中文字体
    font = None

    font_paths = []
    if os.name == 'nt':  # Windows
        font_paths.extend([
            'C:/Windows/Fonts/simhei.ttf',  # 黑体
            'C:/Windows/Fonts/simsun.ttc',  # 宋体
            'C:/Windows/Fonts/msyh.ttc',  # 微软雅黑
            'C:/Windows/Fonts/arial.ttf'  # Arial (作为后备)
        ])
    elif sys.platform == 'darwin':  # macOS
        font_paths.extend([
            '/System/Library/Fonts/PingFang.ttc',
            '/Library/Fonts/Arial Unicode.ttf',
            '/System/Library/Fonts/AppleSDGothicNeo.ttc'
        ])
    else:  # Linux
        font_paths.extend([
            '/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf',
            '/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'
        ])

    # 尝试加载字体
    for path in font_paths:
        try:
            if os.path.exists(path):
                font = ImageFont.truetype(path, font_size)
                break
        except Exception:
            continue

    # 如果找不到中文字体，使用默认字体
    if font is None:
        font = ImageFont.load_default()
        font_size = 12  # 默认字体一般比较小，需要调整大小

    # 调整绘制位置（OpenCV是左下角，PIL是左上角）
    x, y = org
    if bottomLeftOrigin:
        # 如果基准点在左下角，需要上移文本高度
        try:
            # 新版本PIL
            text_bbox = draw.textbbox((0, 0), text, font=font)
            text_height = text_bbox[3] - text_bbox[1]
        except:
            # 旧版本PIL
            text_width, text_height = draw.textsize(text, font=font)

        # 调整y坐标，减去文本高度
        y -= text_height

    # 绘制文本
    draw.text((x, y), text, font=font, fill=color_rgb)

    # 转换回OpenCV图像
    img_result = cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)

    return img_result


def obtain_features_from_camera(camera_id=0, name="person", output_xlsx=None,
                                sampling_interval=15, max_frames=100,
                                display_preview=True):
    """
    从摄像头实时提取人物特征，进行数据融合并保存至Excel
    """
    start_time = time.time()
    logger.info(f"开始从摄像头(ID:{camera_id})提取特征")

    try:
        # 打开摄像头
        cap = cv2.VideoCapture(camera_id)
        if not cap.isOpened():
            logger.error(f"无法打开摄像头 ID: {camera_id}")
            return None

        # 获取摄像头参数
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        fps = cap.get(cv2.CAP_PROP_FPS)

        logger.info(f"摄像头信息: 分辨率 {width}x{height}, {fps:.2f}FPS")

        # 初始化数据收集
        all_ratios = []
        all_shirt_colors = []
        all_pants_colors = []

        # 加载模型
        colour_model, pose_model = load_models()

        # 初始化计数器
        frame_count = 0
        processed_frames = 0
        valid_frames = 0

        logger.info(f"开始采集，采样间隔: {sampling_interval}帧，最大采样数: {max_frames}帧")
        logger.info("按 'q' 键退出采集，按 's' 键立即保存当前结果")

        # 创建显示窗口
        if display_preview:
            cv2.namedWindow("摄像头预览", cv2.WINDOW_NORMAL)
            cv2.resizeWindow("摄像头预览", 800, 600)

        # 实时处理循环
        paused = False
        while processed_frames < max_frames:
            if not paused:
                ret, frame = cap.read()
                if not ret:
                    logger.error("无法从摄像头读取帧")
                    break

                frame_width = frame.shape[1]
                mid_x = frame_width // 2
                left_img = frame[:, :mid_x]
                right_img = frame[:, mid_x:]
                frame = left_img

                frame_count += 1

            # 处理当前帧并显示
            display_frame = frame.copy()

            # 添加状态信息到预览
            stats = {
                'processed': processed_frames,
                'valid': valid_frames,
                'max_frames': max_frames,
                'progress': processed_frames / max_frames * 100 if max_frames > 0 else 0,
                'elapsed_time': time.time() - start_time,
                'name': name,
                'processing': False
            }

            # 显示状态信息
            status_text = f"已采样: {processed_frames}/{max_frames} | 有效: {valid_frames}"
            # display_frame = cv2_put_chinese_text(display_frame, status_text, (10, 30),
            #                                      cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

            # 添加指导文本
            guide_text = "按 'q' 退出, 's' 保存, 'p' 暂停/继续"
            # display_frame = cv2_put_chinese_text(display_frame, guide_text, (10, height - 20),
            #                                      cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 255), 2)

            # 判断是否处理此帧
            if not paused and frame_count % sampling_interval == 0:
                # 在预览中标记正在处理
                stats['processing'] = True
                # display_frame = cv2_put_chinese_text(display_frame, "正在处理...", (width - 200, 30),
                #                                      cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)

                if display_preview:
                    cv2.imshow("摄像头预览", display_frame)
                    cv2.waitKey(1)

                try:
                    # 获取服装检测结果
                    pairs = Obtain_the_target_color(frame)

                    # 提取服装颜色
                    shirt_color = (0, 0, 0)  # 默认黑色
                    pants_color = (0, 0, 0)  # 默认黑色

                    for pair in pairs:
                        # 检查是否有上衣颜色
                        if len(pair) > 2 and pair[2]:
                            shirt_color = pair[2]

                        # 检查是否有裤子颜色
                        if len(pair) > 3 and pair[3]:
                            pants_color = pair[3]

                        # 在图像上标记上衣和下装
                        if len(pair[0]) > 1:  # 有上衣
                            x1, y1, x2, y2 = pair[0][:4]
                            cv2.rectangle(display_frame, (x1, y1), (x2, y2), (255, 0, 0), 2)

                        if len(pair[1]) > 1:  # 有下装
                            x1, y1, x2, y2 = pair[1][:4]
                            cv2.rectangle(display_frame, (x1, y1), (x2, y2), (0, 255, 0), 2)

                    # 获取人体关键点
                    keypoints = get_keypoints(frame, pose_model)

                    # 计算身体比例
                    body_ratios = calculate_body_ratios(keypoints)

                    processed_frames += 1

                    # 添加有效结果到数据集
                    if body_ratios and len(body_ratios) == 16:
                        all_ratios.append(body_ratios)
                        all_shirt_colors.append(shirt_color)
                        all_pants_colors.append(pants_color)
                        valid_frames += 1

                        logger.debug(f"第 {frame_count} 帧特征数据有效，当前有效帧数: {valid_frames}")

                        # 在图像上绘制关键点和骨架
                        if keypoints and len(keypoints) > 0:
                            pose_palette = np.array(
                                [[255, 128, 0], [255, 153, 51], [255, 178, 102], [230, 230, 0], [255, 153, 255],
                                 [153, 204, 255], [255, 102, 255], [255, 51, 255], [102, 178, 255], [51, 153, 255],
                                 [255, 153, 153], [255, 102, 102], [255, 51, 51], [153, 255, 153], [102, 255, 102],
                                 [51, 255, 51], [0, 255, 0], [0, 0, 255], [255, 0, 0], [255, 255, 255]], dtype=np.uint8)

                            kpt_color = pose_palette[[16, 16, 16, 16, 16, 0, 0, 0, 0, 0, 0, 9, 9, 9, 9, 9, 9]]
                            skeleton = [[16, 14], [14, 12], [17, 15], [15, 13], [12, 13], [6, 12], [7, 13], [6, 7],
                                        [6, 8],
                                        [7, 9], [8, 10], [9, 11], [2, 3], [1, 2], [1, 3], [2, 4], [3, 5], [4, 6],
                                        [5, 7]]
                            limb_color = pose_palette[[9, 9, 9, 9, 7, 7, 7, 0, 0, 0, 0, 0, 16, 16, 16, 16, 16, 16, 16]]

                            for person_kpts in keypoints:
                                # 绘制关键点
                                for k, keypoint in enumerate(person_kpts):
                                    x, y, conf = keypoint
                                    color_k = [int(x) for x in kpt_color[k]]
                                    if conf > 0.5 and x != 0 and y != 0:
                                        cv2.circle(display_frame, (int(x), int(y)), 5, color_k, -1,
                                                   lineType=cv2.LINE_AA)

                                # 绘制骨架
                                for k, sk in enumerate(skeleton):
                                    pos1 = (int(person_kpts[(sk[0] - 1), 0]), int(person_kpts[(sk[0] - 1), 1]))
                                    pos2 = (int(person_kpts[(sk[1] - 1), 0]), int(person_kpts[(sk[1] - 1), 1]))

                                    conf1 = person_kpts[(sk[0] - 1), 2]
                                    conf2 = person_kpts[(sk[1] - 1), 2]

                                    if conf1 > 0.5 and conf2 > 0.5 and pos1[0] != 0 and pos1[1] != 0 and pos2[
                                        0] != 0 and pos2[1] != 0:
                                        cv2.line(display_frame, pos1, pos2, [int(x) for x in limb_color[k]],
                                                 thickness=2, lineType=cv2.LINE_AA)

                except Exception as e:
                    logger.error(f"处理第 {frame_count} 帧时出错: {e}")
                    logger.debug(traceback.format_exc())

            # 更新状态显示
            stats['valid'] = valid_frames
            stats['processing'] = False

            status_text = f"processing: {processed_frames}/{max_frames} | USED: {valid_frames}"
            cv2.putText(display_frame, status_text, (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

            if paused:
                pause_text = "已暂停 - 按 'p' 继续"
                # display_frame = cv2_put_chinese_text(display_frame, pause_text, (width // 2 - 100, 30),
                #                                      cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)

            # 显示预览
            if display_preview:
                cv2.imshow("摄像头预览", display_frame)

            # 检查键盘输入
            key = cv2.waitKey(1) & 0xFF
            if key == ord('q'):
                logger.info("用户请求结束采集")
                break
            elif key == ord('s'):
                logger.info("用户请求立即保存当前结果")
                break
            elif key == ord('p'):
                paused = not paused
                logger.info(f"{'暂停' if paused else '继续'}采集")

        # 释放摄像头
        cap.release()

        # 关闭预览窗口
        if display_preview:
            cv2.destroyWindow("摄像头预览")

        # 检查是否有足够的数据进行融合
        if not all_ratios:
            logger.error("没有有效的特征数据，无法进行数据融合")
            return None

        logger.info(f"采集完成。总帧数: {frame_count}，采样: {processed_frames}帧，有效特征帧数: {valid_frames}")

        # 融合数据 - 使用正确的中位数方法
        fused_ratios = []
        all_ratios_array = np.array(all_ratios)
        for i in range(all_ratios_array.shape[1]):
            values = all_ratios_array[:, i]
            non_zero_values = values[values != 0]
            if len(non_zero_values) > 0:
                fused_ratios.append(float(np.median(non_zero_values)))
            else:
                fused_ratios.append(0.0)

        # 对于颜色，找出最频繁出现的颜色
        def most_common_color(colors):
            # 确保所有颜色都是元组形式
            color_tuples = []
            for color in colors:
                if isinstance(color, tuple):
                    color_tuples.append(color)
                elif isinstance(color, list) and len(color) == 3:
                    color_tuples.append(tuple(color))
                elif isinstance(color, str) and color.startswith('(') and color.endswith(')'):
                    # 处理字符串形式的元组 "(r,g,b)"
                    try:
                        color = eval(color)
                        color_tuples.append(color)
                    except:
                        continue

            if not color_tuples:
                return (0, 0, 0)  # 默认黑色

            # 使用Counter计数
            from collections import Counter
            counter = Counter(color_tuples)
            most_common = counter.most_common(1)

            if most_common:
                return most_common[0][0]
            else:
                return color_tuples[0]  # 返回第一个颜色

        # 融合颜色数据
        fused_shirt_color = most_common_color(all_shirt_colors)
        fused_pants_color = most_common_color(all_pants_colors)

        # 保存到Excel
        if output_xlsx:
            output_path = Path(output_xlsx)
            # 如果是目录，将文件名附加到目录路径
            if output_path.is_dir():
                output_path = output_path / f"{name}.xlsx"
            # 如果没有Excel扩展名，添加一个
            if not str(output_path).lower().endswith(('.xlsx', '.xls')):
                output_path = Path(f"{output_path}.xlsx")
        else:
            current_time = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            output_path = Path(OUTPUT_DIR) / f"person_{name}_ratios_{current_time}.xlsx"

        logger.info(f"正在保存数据到Excel: {output_path}")

        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = sanitize_sheet_name(f'特征_{name}')

        # 添加比例数据 - 直接将数据保存到A列
        for i, ratio in enumerate(fused_ratios):
            sheet.cell(row=i + 1, column=1).value = ratio

        # 添加颜色数据 - 直接添加到A17和A18单元格
        sheet.cell(row=17, column=1).value = str(fused_shirt_color)
        sheet.cell(row=18, column=1).value = str(fused_pants_color)

        # 确保目录存在
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info(f"Excel文件保存成功: {output_path}")

        total_time = time.time() - start_time
        logger.info(f"摄像头处理完成，总耗时: {total_time:.2f}秒 ({total_time / 60:.2f}分钟)")

        # 合并结果
        return fused_ratios + [fused_shirt_color, fused_pants_color]

    except Exception as e:
        logger.error(f"处理摄像头视频流时出错: {e}")
        logger.error(traceback.format_exc())
        return None


def main():
    """
    主函数，处理命令行参数并执行程序
    """
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='基于RKNN的特征提取工具')
    parser.add_argument('--image', type=str, help='输入图像路径',default="/home/monster/桌面/image/2.png")
    parser.add_argument('--camera', action='store_true', help='使用摄像头提取特征',default=True)
    parser.add_argument('--camera-id', type=int, default=1, help='摄像头ID（默认：0）')
    parser.add_argument('--output', type=str, default=OUTPUT_DIR, help='输出文件或目录路径')
    parser.add_argument('--name', type=str, default='person', help='人物标识名称')
    parser.add_argument('--display', action='store_true', help='显示处理结果')
    parser.add_argument('--interval', type=int, default=5, help='摄像头采样间隔（默认：15帧）')
    parser.add_argument('--max-frames', type=int, default=100, help='摄像头最大采样帧数（默认：100）')
    args = parser.parse_args()

    try:
        # 设置输出目录
        setup_output_directory(OUTPUT_DIR)

        # 处理图像
        if args.image:
            logger.info(f"开始处理图像: {args.image}")
            canvas, person_ratios = obtain_features(args.image, args.name)

            if args.display and canvas is not None:
                cv2.imshow("处理结果", canvas)
                cv2.waitKey(0)
                cv2.destroyAllWindows()

            if person_ratios:
                logger.info(f"图像处理完成，提取到特征数量: {len(person_ratios)}")
            else:
                logger.error("特征提取失败")

        # 处理摄像头
        elif args.camera:
            logger.info(f"开始从摄像头提取特征")
            features = obtain_features_from_camera(
                camera_id=args.camera_id,
                name=args.name,
                output_xlsx=args.output,
                sampling_interval=args.interval,
                max_frames=args.max_frames,
                display_preview=True
            )

            if features:
                logger.info("摄像头特征提取成功")
            else:
                logger.error("摄像头特征提取失败")

        # 如果没有输入
        else:
            logger.error("请提供--image或--camera参数")
            parser.print_help()

    except Exception as e:
        logger.error(f"程序执行出错: {e}")
        logger.error(traceback.format_exc())
    finally:
        # 关闭所有窗口
        cv2.destroyAllWindows()


if __name__ == '__main__':
    main()